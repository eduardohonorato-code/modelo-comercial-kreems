"""
Exportación de tablas: CSV y PNG de alta calidad (para enviar reportes por
WhatsApp sin sacar pantallazo).

El PNG se genera server-side con matplotlib (sin navegador → funciona en
Streamlit Cloud). matplotlib se importa DENTRO de tabla_png para no encarecer
la carga de páginas que no exportan.
"""
import io

NAVY = "#1B3A6B"
PINK = "#E62984"
PINK_TINT = "#FBEAF0"
ZEBRA = "#F5F5F9"
VERDE = "#1A7F4B"
AMBAR = "#D4881E"
ROJO = "#C0392B"

# Colores de las bandas de grupo (encabezados agrupados del PNG)
GRP_AZUL = "#1E5FA5"    # Facturación / Pedidos
GRP_VERDE = "#1A7F4B"   # Máquinas
GRP_NARANJO = "#C97A16"  # Efectividad visitas


def color_hex(pct, ok: float = 1.0, warn: float = 0.7):
    """Color de semáforo para un porcentaje (o None si no aplica)."""
    try:
        v = float(pct)
    except (TypeError, ValueError):
        return None
    if v != v:  # NaN
        return None
    if v >= ok:
        return VERDE
    if v >= warn:
        return AMBAR
    return ROJO


def to_csv(df) -> bytes:
    """DataFrame → CSV bytes (delimitador ';' y BOM, se abre bien en Excel)."""
    return df.to_csv(index=False, sep=";").encode("utf-8-sig")


def to_xlsx(df, hoja: str = "Datos") -> bytes:
    """
    DataFrame → archivo .xlsx nativo (bytes). Se abre limpio en Excel: sin
    problemas de acentos ni de delimitador (a diferencia del CSV). Encabezado
    navy en negrita, filtro automático y ancho de columna proporcional.
    openpyxl ya es dependencia del proyecto (lo usa el ETL / lectura de Excel).
    """
    import pandas as pd  # noqa: F401  (asegura que openpyxl esté disponible)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = hoja[:31]  # Excel limita el nombre de hoja a 31 chars

    cols = [str(c) for c in df.columns]
    hdr_fill = PatternFill("solid", fgColor="1B3A6B")
    hdr_font = Font(bold=True, color="FFFFFF")
    for j, name in enumerate(cols, 1):
        c = ws.cell(row=1, column=j, value=name)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")

    for i, (_, row) in enumerate(df.iterrows(), 2):
        for j, name in enumerate(df.columns, 1):
            ws.cell(row=i, column=j, value=row[name])

    for j, name in enumerate(cols, 1):
        ancho = max([len(name)] + [len(str(v)) for v in df.iloc[:, j - 1]]) + 2
        ws.column_dimensions[get_column_letter(j)].width = min(max(ancho, 8), 48)

    ws.freeze_panes = "A2"
    if len(df):
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(df) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def tabla_png(df, titulo: str, subtitulo: str = "", color_celdas: dict | None = None,
              resaltar_ultima: bool = False, col_labels=None, grupos=None,
              dpi: int = 200) -> bytes:
    """
    Render de un DataFrame de STRINGS ya formateados a PNG.

    - `subtitulo`: línea de contexto (ej. días del mes / última factura).
    - `color_celdas`: dict {(fila_idx, nombre_col): color_hex} para texto de celda
      (ej. % Cumpl en verde/ámbar/rojo). `fila_idx` es la posición 0-based en `df`.
    - `resaltar_ultima`: pinta la última fila como TOTAL (fondo rosado, negrita).
    - `col_labels`: lista opcional de encabezados a mostrar (default = df.columns).
    - `grupos`: lista opcional de (titulo, color_hex, col_ini, col_fin) — dibuja una
      banda de grupos coloreada SOBRE los encabezados (col_ini/col_fin inclusivos,
      0-based). Las columnas fuera de todo grupo quedan sin banda.
    """
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.patches import Rectangle

    color_celdas = color_celdas or {}
    df = df.astype(str)
    n_rows, n_cols = df.shape
    labels = list(col_labels) if col_labels else [str(c) for c in df.columns]

    widths = [max([len(labels[j])] + [len(v) for v in df.iloc[:, j]])
              for j in range(n_cols)]
    total_w = sum(widths)
    fig_w = min(max(total_w * 0.102 + 0.6, 6.0), 32.0)
    banda_in = 0.42 if grupos else 0.0
    # Encabezado: título navy a la izquierda + subtítulo (sin barra ni subrayado).
    L, R = 0.006, 0.994
    PAD, TITLE_H, GAP = 0.12, 0.46, 0.16
    SUB_H = 0.34 if subtitulo else 0.0
    header_in = PAD + TITLE_H + (GAP + SUB_H if subtitulo else 0.14) + 0.05
    fig_h = (n_rows + 1) * 0.34 + banda_in + header_in

    fig = plt.figure(figsize=(fig_w, fig_h), dpi=dpi, facecolor="white")

    def _f(inch):
        return inch / fig_h

    y_top = 1 - _f(PAD)
    fig.text(L, y_top, titulo, fontsize=20, fontweight="bold", color=NAVY,
             va="top", ha="left")
    y_after = y_top - _f(TITLE_H)
    if subtitulo:
        fig.text(L, y_after - _f(GAP), subtitulo, fontsize=10.5, color="#1A1A1A",
                 fontweight="bold", va="top")
        table_top = y_after - _f(GAP) - _f(SUB_H)
    else:
        table_top = y_after - _f(0.14)

    ax = fig.add_axes([L, 0.008, R - L, table_top - 0.008])
    ax.axis("off")
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)

    ax_h_in = (table_top - 0.008) * fig_h
    band_frac = (banda_in / ax_h_in) if grupos else 0.0

    tbl = ax.table(cellText=df.values.tolist(), colLabels=labels,
                   cellLoc="right", bbox=[0, 0, 1, 1 - band_frac])
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(9)

    # Ancho de columna proporcional al contenido (si no, matplotlib reparte igual
    # y trunca los nombres largos de vendedor).
    for (r, c), cell in tbl.get_celld().items():
        cell.set_width(widths[c] / total_w)
        cell.set_edgecolor("#E6E6EC")
        cell.set_linewidth(0.5)
        txt = cell.get_text()
        if r == 0:
            cell.set_facecolor(NAVY)
            txt.set_color("white")
            txt.set_fontweight("bold")
            txt.set_fontsize(8.5)
        else:
            i = r - 1
            if resaltar_ultima and i == n_rows - 1:
                cell.set_facecolor(PINK_TINT)
                txt.set_fontweight("bold")
            else:
                cell.set_facecolor(ZEBRA if (r % 2 == 0) else "white")
            col = df.columns[c]
            if (i, col) in color_celdas:
                txt.set_color(color_celdas[(i, col)])
                txt.set_fontweight("bold")
        if c == 0:
            txt.set_ha("left")

    # Banda de grupos sobre los encabezados
    if grupos:
        x_bounds = [0.0]
        for w in widths:
            x_bounds.append(x_bounds[-1] + w / total_w)
        y0 = 1 - band_frac
        for titulo_g, color_g, c0, c1 in grupos:
            gx0, gx1 = x_bounds[c0], x_bounds[c1 + 1]
            pad = 0.003
            ax.add_patch(Rectangle((gx0 + pad, y0 + 0.015),
                                   (gx1 - gx0) - 2 * pad, band_frac - 0.03,
                                   facecolor=color_g, edgecolor="none", clip_on=False))
            # Fuente que garantiza que el título quepa en el ancho del grupo
            # (0.72 ≈ ancho medio de carácter en negrita; 0.90 deja margen).
            avail_in = (gx1 - gx0) * (0.988 * fig_w)
            fs = max(6.5, min(11.0, avail_in * 0.90 * 72 / (0.72 * max(len(titulo_g), 1))))
            ax.text((gx0 + gx1) / 2, y0 + band_frac / 2, titulo_g,
                    ha="center", va="center", fontsize=fs, fontweight="bold",
                    color="white", clip_on=False)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight",
                pad_inches=0.12, facecolor="white")
    plt.close(fig)
    return buf.getvalue()


def bloque_descarga(disp, colores: dict, titulo: str, subtitulo: str,
                    nombre_base: str, col_labels=None, grupos=None, disp_csv=None):
    """
    Expander DISCRETO con descarga de imagen (PNG) y datos (CSV) de una tabla.
    El PNG se genera solo al pulsar el botón (no en cada carga de página).
    `disp` es el DataFrame de strings ya formateados (la última fila = TOTAL).
    `col_labels`/`grupos` se pasan tal cual a tabla_png (encabezados + banda de grupos).
    `disp_csv`: si se entrega, el CSV usa este DataFrame (más detallado) en vez de
    `disp` — así la imagen puede ser compacta y el CSV traer más columnas.
    """
    import streamlit as st
    csv_df = disp_csv if disp_csv is not None else disp
    with st.expander("📥 Descargar / Exportar (imagen PNG o datos)"):
        st.caption("La imagen PNG es ideal para enviar por WhatsApp (mejor calidad "
                   "que un pantallazo). El botón NO aparece dentro de la imagen.")
        c1, c2 = st.columns(2)
        with c1:
            if st.button("🖼️ Generar imagen PNG", key=f"png_{nombre_base}",
                         use_container_width=True):
                st.session_state[f"_png_{nombre_base}"] = tabla_png(
                    disp, titulo, subtitulo, color_celdas=colores, resaltar_ultima=True,
                    col_labels=col_labels, grupos=grupos)
            png = st.session_state.get(f"_png_{nombre_base}")
            if png:
                st.download_button("⬇️ Descargar PNG", png, f"{nombre_base}.png",
                                   "image/png", key=f"dl_png_{nombre_base}",
                                   use_container_width=True)
        with c2:
            st.download_button("📄 Descargar datos (CSV)", to_csv(csv_df),
                               f"{nombre_base}.csv", "text/csv",
                               key=f"dl_csv_{nombre_base}", use_container_width=True)
