"""
Informe de seguimiento de máquinas (comodato).

Responde, para el período filtrado: cuántas máquinas se instalaron, se
cambiaron y se retiraron, en qué estado quedó cada movimiento y cuáles siguen
sin confirmarse en terreno.

Cómo se arma (dos fuentes que hay que cruzar sí o sí):
  · **Qué movimiento fue** → Obuma. Cada máquina es una línea de flete de la
    categoría "Maquinas" en la factura: FL-4 instalación en cliente nuevo,
    FL-1/FL-3/FL-5 cambio, FL-2 retiro por término.
  · **Si se ejecutó o no** → Autoventa, "Detalle de despachos". Se cruza por
    número de documento (Obuma `N° DCTO` = Autoventa `Documento`) y el estado
    del despacho (Entregada / Rechazada / Pendiente) es lo que confirma que la
    máquina efectivamente se movió.

Dos advertencias que el informe hace explícitas en vez de esconder:
  · La fecha de ruta puede caer en un mes distinto al de la factura, así que el
    cruce se hace por documento sobre una ventana de despachos más ancha que el
    período, no mes contra mes.
  · Donde no hay despachos cargados (Acuña, y todo lo anterior a feb-2026) el
    movimiento no queda "pendiente": queda **Sin información**, que es distinto.
"""
import io
from datetime import date

import pandas as pd

from app.export_analisis import _escribir, _lbl_mes, _con_total, _FMT_NUM, _FMT_PCT

_FMT_FECHA = "dd/mm/yyyy"
_FMT_DEC = "#,##0.0"

# Movimiento según el código de flete de Obuma (etl/config.py TIPO_MOV_MAP).
MOV_LBL = {"nueva": "Nueva (FL-4)", "cambio": "Cambio (FL-1/3/5)",
           "retiro": "Retiro (FL-2)"}
FL_A_MOV = {"FL-4": "nueva", "FL-1": "cambio", "FL-3": "cambio",
            "FL-5": "cambio", "FL-2": "retiro"}

# Transportistas dedicados a máquinas ("Máquinas Externo RM", "Máquinas Kreems 1").
# Sirven para reconocer un despacho de máquina que TODAVÍA no tiene factura en
# Obuma, que es justo lo que hace diferir este conteo del de Autoventa. No es un
# filtro suficiente por sí solo: 40 movimientos de 2026 salieron con el
# transportista de la ruta normal (Cancino Temuco, sobre todo).
_TRANSP_MAQUINA = "quina"

ENTREGADA = "Entregada"
RECHAZADA = "Rechazada"
EN_RUTA = "Pendiente en ruta"
SIN_DESPACHO = "Sin despacho"
SIN_INFO = "Sin información"
ESTADOS = [ENTREGADA, RECHAZADA, EN_RUTA, SIN_DESPACHO, SIN_INFO]

# Prioridad al resolver un documento con varios despachos: si alguno quedó
# Entregada, el movimiento se ejecutó, aunque hubiera un intento rechazado antes.
_PRIO_DESP = {"entregada": 0, "rechazada": 1, "pendiente": 2}

# Meta comercial de gestiones semanales (definición de gerencia, ago-2026):
# una "gestión" es un flete de máquina CON DTE emitido —da lo mismo si fue
# instalación, cambio o retiro— y las 22 son del equipo completo, no por
# vendedor. Los pedidos que siguen "Sin DTE" no suman: son el pendiente.
META_GESTIONES_SEMANA = 22



def _norm_doc(s: pd.Series) -> pd.Series:
    return s.astype(str).str.strip().str.replace(r"\.0$", "", regex=True)


def _mes(f):
    try:
        return _lbl_mes(f.to_period("M"))
    except (AttributeError, ValueError):
        return "—"


def _pct(num, den):
    return (num / den) if den else 0.0


def _semanas_del_rango(f_ini, f_fin) -> list:
    """
    [(semana, días dentro del rango)] de lunes a domingo, incluidas las semanas
    sin ninguna gestión: una semana en cero es justo lo que hay que ver contra
    la meta. Se devuelven TODAS las que tocan el rango, para que el total de la
    hoja cuadre con el del informe; los días dentro del rango dicen cuáles
    quedaron cortadas y no se pueden juzgar contra la meta.
    """
    ini, fin = pd.Timestamp(f_ini), pd.Timestamp(f_fin)
    out = []
    for per in pd.period_range(ini, fin, freq="W-SUN"):
        dias = (min(per.end_time.normalize(), fin)
                - max(per.start_time.normalize(), ini)).days + 1
        out.append((per, int(dias)))
    return out


def _prep_pedidos(pedidos_fl: pd.DataFrame | None) -> pd.DataFrame:
    """
    Pedidos FL con la fecha de INGRESO resuelta y el tipo de movimiento.

    `fecha_pedido` (created_at de la línea en Autoventa) es cuándo el vendedor
    ingresó el pedido; `fecha` es la del DTE, o la de despacho pedida si todavía
    no factura. Si la base aún no tiene fecha_pedido —sql/036 sin correr, o un
    mes sin recargar— se cae a `fecha` para no dejar la fila fuera.
    """
    if pedidos_fl is None or pedidos_fl.empty:
        return pd.DataFrame()
    p = pedidos_fl.copy()
    p["fecha"] = pd.to_datetime(p["fecha"], errors="coerce")
    fp = (pd.to_datetime(p["fecha_pedido"], errors="coerce")
          if "fecha_pedido" in p.columns
          else pd.Series(pd.NaT, index=p.index))
    p["_ingreso"] = fp.fillna(p["fecha"])
    p["_cod"] = p["producto_codigo"].astype(str).str.upper().str.strip()
    p["_mov"] = p["_cod"].map(FL_A_MOV)
    p["_sin_dte"] = (p["doc_venta"].astype(str).str.strip().str.lower()
                     == "sin dte")
    p["_dias_a_dte"] = (p["fecha"] - p["_ingreso"]).dt.days.where(~p["_sin_dte"])
    # Un pedido sin `estado_pedido` es uno que la última recarga NO tocó: ya no
    # viene en la API. En la práctica significa anulado o vuelto a ingresar con
    # otro número, y si se deja pasar infla la cola de pendientes con trabajo
    # que en realidad no existe.
    p["_fantasma"] = (p["estado_pedido"].isna() if "estado_pedido" in p.columns
                      else pd.Series(False, index=p.index))
    return p


def _desc(ruts: pd.Series, clientes: pd.DataFrame | None, col: str) -> pd.Series:
    """Columna descriptiva de dim_cliente para una serie de RUT."""
    if clientes is None or clientes.empty or col not in clientes.columns:
        return pd.Series("(sin dato)", index=ruts.index)
    mapa = clientes.drop_duplicates("rut").set_index("rut")[col]
    return ruts.map(mapa).fillna("(sin dato)")


# ── Preparación del detalle ──────────────────────────────────────────────────

def preparar_movimientos(maquinas: pd.DataFrame,
                         despachos: pd.DataFrame | None = None,
                         lineas_fl: pd.DataFrame | None = None,
                         vendedores: pd.DataFrame | None = None,
                         clientes: pd.DataFrame | None = None,
                         sociedades: dict | None = None,
                         hoy: date | None = None) -> pd.DataFrame:
    """
    Una fila por movimiento de máquina, con el despacho que lo confirma (si lo
    hay) y todo lo descriptivo pegado. Es la base de todas las hojas.
    """
    hoy = hoy or date.today()
    m = maquinas.copy()
    m["fecha"] = pd.to_datetime(m["fecha"], errors="coerce")
    m["_doc"] = _norm_doc(m["documento"])
    m["_soc"] = pd.to_numeric(m.get("sociedad_id"), errors="coerce")

    # ── Código FL exacto, cantidad de máquinas y si el documento es NC ───────
    m["Código FL"] = ""
    m["Máquinas (cant.)"] = 1
    m["_nc"] = False
    if lineas_fl is not None and not lineas_fl.empty:
        fl = lineas_fl.copy()
        fl["_doc"] = _norm_doc(fl["n_dcto"])
        fl["_soc"] = pd.to_numeric(fl.get("sociedad_id"), errors="coerce")
        fl["producto_codigo"] = (fl["producto_codigo"].astype(str)
                                 .str.upper().str.strip())
        fl["_mov"] = fl["producto_codigo"].map(FL_A_MOV)
        fl["_cant"] = pd.to_numeric(fl["cantidad"], errors="coerce").abs().fillna(1)
        fl["_es_nc"] = (fl["tipo_dcto"].astype(str).str.upper()
                        .str.contains("CREDITO", na=False))
        agg = (fl.dropna(subset=["_mov"])
               .groupby(["_soc", "_doc", "cliente_rut", "_mov"])
               .agg(_cod=("producto_codigo", lambda s: " / ".join(sorted(set(s)))),
                    _cant=("_cant", "sum"),
                    _es_nc=("_es_nc", "any"))
               .reset_index())
        m = m.merge(agg, how="left",
                    left_on=["_soc", "_doc", "cliente_rut", "tipo_mov"],
                    right_on=["_soc", "_doc", "cliente_rut", "_mov"])
        m["Código FL"] = m["_cod"].fillna("")
        m["Máquinas (cant.)"] = (pd.to_numeric(m["_cant"], errors="coerce")
                                 .fillna(1).round().astype(int).clip(lower=1))
        m["_nc"] = m["_es_nc"].fillna(False).astype(bool)
        m = m.drop(columns=[c for c in ("_cod", "_cant", "_es_nc", "_mov")
                            if c in m.columns])

    # ── Despacho que confirma el movimiento (cruce por documento) ───────────
    cobertura = set()
    if despachos is not None and not despachos.empty:
        d = despachos.copy()
        d["fecha_ruta"] = pd.to_datetime(d["fecha_ruta"], errors="coerce")
        d["_doc"] = _norm_doc(d["documento"])
        d["_soc"] = pd.to_numeric(d.get("sociedad_id"), errors="coerce")
        d["_est"] = d["estado"].astype(str).str.strip().str.lower()
        # Meses/sociedades con despachos cargados: fuera de ahí, "sin despacho"
        # no significa nada (nunca se cargó el archivo).
        cobertura = set(zip(d["_soc"],
                            d["fecha_ruta"].dt.to_period("M").astype(str)))
        d["_prio"] = d["_est"].map(_PRIO_DESP).fillna(9)
        d = d.sort_values(["_prio", "fecha_ruta"])
        n_int = d.groupby("_doc").size().rename("_intentos")
        dd = (d.drop_duplicates("_doc")
              [["_doc", "fecha_ruta", "estado", "_est", "transportista",
                "devolucion", "peso"]]
              .rename(columns={"estado": "_estado_desp"}))
        dd = dd.merge(n_int, left_on="_doc", right_index=True, how="left")
        m = m.merge(dd, on="_doc", how="left")
    else:
        for c in ("_estado_desp", "_est", "transportista", "devolucion",
                  "peso", "_intentos"):
            m[c] = pd.NA
        m["fecha_ruta"] = pd.NaT

    # ── Estado de entrega ────────────────────────────────────────────────────
    mapa = {"entregada": ENTREGADA, "rechazada": RECHAZADA, "pendiente": EN_RUTA}
    estado = m["_est"].map(mapa)
    llave_cob = list(zip(m["_soc"], m["fecha"].dt.to_period("M").astype(str)))
    con_dato = pd.Series([k in cobertura for k in llave_cob], index=m.index)
    estado = estado.fillna(pd.Series(
        [SIN_DESPACHO if c else SIN_INFO for c in con_dato], index=m.index))
    m["Estado entrega"] = estado

    # ── Descriptivos ─────────────────────────────────────────────────────────
    nom_vend = {}
    if vendedores is not None and not vendedores.empty:
        nom_vend = dict(zip(vendedores["id"], vendedores["nombre_canonico"]))
    m["Vendedor"] = m["vendedor_id"].map(nom_vend).fillna("Sin asignar")
    m["Sociedad"] = m["_soc"].map(sociedades or {}).fillna("(sin dato)")

    for col in ("Cliente", "Comuna", "Región"):
        m[col] = "(sin dato)"
    if clientes is not None and not clientes.empty:
        c = clientes.drop_duplicates("rut").set_index("rut")
        for col, orig in (("Cliente", "razon_social"), ("Comuna", "comuna"),
                          ("Región", "region")):
            if orig in c.columns:
                m[col] = (m["cliente_rut"].map(c[orig])
                          .fillna("(sin dato)").replace("", "(sin dato)"))

    m["Movimiento"] = m["tipo_mov"].map(MOV_LBL).fillna(m["tipo_mov"])
    m["Mes"] = [_mes(f) for f in m["fecha"]]
    m["Tipo documento"] = m["_nc"].map({True: "Nota de crédito (anula)",
                                        False: "Factura"})
    m["Confirmado en Autoventa"] = m["Estado entrega"].isin(
        [ENTREGADA, RECHAZADA, EN_RUTA]).map({True: "Sí", False: "No"})
    m["Fecha ruta"] = pd.to_datetime(m.get("fecha_ruta"), errors="coerce")
    m["Días factura a ruta"] = (m["Fecha ruta"] - m["fecha"]).dt.days
    dias_abierto = (pd.Timestamp(hoy) - m["fecha"]).dt.days
    m["Días sin confirmar"] = dias_abierto.where(
        ~m["Estado entrega"].isin([ENTREGADA, RECHAZADA]))
    m["Devolución"] = (m["devolucion"].map({True: "Sí", False: "No"})
                       if "devolucion" in m.columns else "")
    m["Intentos de despacho"] = pd.to_numeric(m.get("_intentos"), errors="coerce")
    m["Transportista"] = m.get("transportista", pd.Series(index=m.index, dtype=str))
    m["Peso (Kg)"] = pd.to_numeric(m.get("peso"), errors="coerce")
    m["_entregada"] = m["Estado entrega"] == ENTREGADA
    m["_rechazada"] = m["Estado entrega"] == RECHAZADA
    m["_pendiente"] = m["Estado entrega"].isin([EN_RUTA, SIN_DESPACHO])
    m["_sin_info"] = m["Estado entrega"] == SIN_INFO
    return m.sort_values(["fecha", "_doc"]).reset_index(drop=True)


# ── Hojas ────────────────────────────────────────────────────────────────────

_COLS_DETALLE = [
    "Fecha", "Mes", "Documento", "Tipo documento", "Sociedad", "Movimiento",
    "Código FL", "Máquinas (cant.)", "Vendedor", "RUT", "Cliente", "Comuna",
    "Región", "Estado entrega", "Fecha ruta", "Días factura a ruta",
    "Transportista", "Intentos de despacho", "Devolución", "Peso (Kg)",
    "Confirmado en Autoventa", "Días sin confirmar",
]
_FMT_DETALLE = {
    "Fecha": _FMT_FECHA, "Fecha ruta": _FMT_FECHA,
    "Máquinas (cant.)": _FMT_NUM, "Días factura a ruta": _FMT_NUM,
    "Días sin confirmar": _FMT_NUM, "Peso (Kg)": _FMT_DEC,
    "Intentos de despacho": _FMT_NUM,
}


def _detalle(m: pd.DataFrame) -> pd.DataFrame:
    d = m.rename(columns={"fecha": "Fecha", "_doc": "Documento",
                          "cliente_rut": "RUT"}).copy()
    if d.empty:
        return pd.DataFrame(columns=_COLS_DETALLE)
    d["Fecha"] = d["Fecha"].dt.date
    d["Fecha ruta"] = d["Fecha ruta"].dt.date
    for c in _COLS_DETALLE:
        if c not in d.columns:
            d[c] = ""
    return d[_COLS_DETALLE]


_CLAVES_CONTEO = [
    "Nuevas (FL-4)", "Cambios (FL-1/3/5)", "Retiros (FL-2)", "Movimientos",
    "Máquinas (cant.)", "Parque neto (nuevas - retiros)", "Entregadas",
    "Rechazadas", "Pendientes", "Sin información", "% Entregado", "% Rechazo",
]


def _conteos(g: pd.DataFrame) -> dict:
    """Bloque de métricas común a todas las hojas agrupadas."""
    nuevas = int((g["tipo_mov"] == "nueva").sum())
    retiros = int((g["tipo_mov"] == "retiro").sum())
    ent = int(g["_entregada"].sum())
    rech = int(g["_rechazada"].sum())
    sin_info = int(g["_sin_info"].sum())
    con_info = len(g) - sin_info
    return {
        "Nuevas (FL-4)": nuevas,
        "Cambios (FL-1/3/5)": int((g["tipo_mov"] == "cambio").sum()),
        "Retiros (FL-2)": retiros,
        "Movimientos": int(len(g)),
        "Máquinas (cant.)": int(g["Máquinas (cant.)"].sum()),
        "Parque neto (nuevas - retiros)": nuevas - retiros,
        "Entregadas": ent,
        "Rechazadas": rech,
        "Pendientes": int(g["_pendiente"].sum()),
        "Sin información": sin_info,
        "% Entregado": _pct(ent, con_info),
        "% Rechazo": _pct(rech, con_info),
    }


_FMT_AGRUP = {k: _FMT_NUM for k in _CLAVES_CONTEO}
_FMT_AGRUP.update({"% Entregado": _FMT_PCT, "% Rechazo": _FMT_PCT,
                   "N° clientes": _FMT_NUM})
_NO_SUM = ("% Entregado", "% Rechazo", "N° clientes")


def _agrupado(m: pd.DataFrame, por, etiqueta: str,
              extra_clientes: bool = False) -> pd.DataFrame:
    por = [por] if isinstance(por, str) else list(por)
    filas = []
    for llave, g in m.groupby(por, dropna=False):
        llave = llave if isinstance(llave, tuple) else (llave,)
        fila = dict(zip(por, llave))
        fila.update(_conteos(g))
        if extra_clientes:
            fila["N° clientes"] = int(g["cliente_rut"].nunique())
        filas.append(fila)
    if not filas:
        return pd.DataFrame()
    return pd.DataFrame(filas).rename(columns={por[0]: etiqueta})


# ── Libro ────────────────────────────────────────────────────────────────────

def libro_maquinas(maquinas: pd.DataFrame, f_ini, f_fin, soc_lbl: str = "Ambas",
                   despachos: pd.DataFrame | None = None,
                   lineas_fl: pd.DataFrame | None = None,
                   pedidos_fl: pd.DataFrame | None = None,
                   maquinas_ctx: pd.DataFrame | None = None,
                   vendedores: pd.DataFrame | None = None,
                   clientes: pd.DataFrame | None = None,
                   sociedades: dict | None = None,
                   meta_semanal: int = META_GESTIONES_SEMANA,
                   hoy: date | None = None) -> bytes:
    """Devuelve el .xlsx del informe de seguimiento de máquinas."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    hoy = hoy or date.today()

    if maquinas is None or maquinas.empty:
        _escribir(wb, "Resumen", pd.DataFrame(),
                  nota="Sin movimientos de máquinas en el rango elegido.")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    m = preparar_movimientos(maquinas, despachos, lineas_fl, vendedores,
                             clientes, sociedades, hoy)

    tot = len(m)
    nuevas = int((m["tipo_mov"] == "nueva").sum())
    cambios = int((m["tipo_mov"] == "cambio").sum())
    retiros = int((m["tipo_mov"] == "retiro").sum())
    ent = int(m["_entregada"].sum())
    rech = int(m["_rechazada"].sum())
    en_ruta = int((m["Estado entrega"] == EN_RUTA).sum())
    sin_desp = int((m["Estado entrega"] == SIN_DESPACHO).sum())
    sin_info = int(m["_sin_info"].sum())
    con_info = tot - sin_info
    anuladas = int(m["_nc"].sum())
    nuevas_ent = int(((m["tipo_mov"] == "nueva") & m["_entregada"]).sum())
    nuevas_info = int(((m["tipo_mov"] == "nueva") & ~m["_sin_info"]).sum())
    retiros_ent = int(((m["tipo_mov"] == "retiro") & m["_entregada"]).sum())
    retiros_info = int(((m["tipo_mov"] == "retiro") & ~m["_sin_info"]).sum())
    dias_ent = m.loc[m["_entregada"], "Días factura a ruta"].dropna()
    _sem_rango = _semanas_del_rango(f_ini, f_fin)
    _cuenta_sem = m["fecha"].dt.to_period("W-SUN").value_counts()
    _completas = [p for p, dias in _sem_rango if dias == 7]
    n_semanas = len(_completas)
    semanas_ok = sum(1 for p in _completas
                     if int(_cuenta_sem.get(p, 0)) >= meta_semanal)
    _dias_rango = (pd.Timestamp(f_fin) - pd.Timestamp(f_ini)).days + 1
    ped = _prep_pedidos(pedidos_fl)
    # Ingresados en el período (por fecha de ingreso, no por la del DTE).
    ped_periodo = (ped[ped["_ingreso"].between(pd.Timestamp(f_ini),
                                               pd.Timestamp(f_fin))]
                   if not ped.empty else ped)
    sin_dte_n = int((ped["_sin_dte"] & ~ped["_fantasma"]).sum()) if not ped.empty else 0
    fantasmas_n = int((ped["_sin_dte"] & ped["_fantasma"]).sum()) if not ped.empty else 0
    ingresados = len(ped_periodo)
    dias_dte = (ped["_dias_a_dte"].dropna() if not ped.empty
                else pd.Series(dtype=float))

    # ── 1. Resumen ───────────────────────────────────────────────────────────
    ind = [
        ("Período del informe", f"{f_ini:%d/%m/%Y} a {f_fin:%d/%m/%Y}"),
        ("Sociedad", soc_lbl),
        ("Informe generado el", f"{hoy:%d/%m/%Y}"),
        ("", ""),
        ("MOVIMIENTOS DE MÁQUINA", ""),
        ("Movimientos totales", tot),
        ("Máquinas involucradas (cantidad facturada)",
         int(m["Máquinas (cant.)"].sum())),
        ("Nuevas - instalación cliente nuevo (FL-4)", nuevas),
        ("Cambios de máquina (FL-1 / FL-3 / FL-5)", cambios),
        ("Retiros por término (FL-2)", retiros),
        ("Parque neto del período (nuevas - retiros)", nuevas - retiros),
        ("Movimientos anulados por nota de crédito", anuladas),
        ("Pedidos de máquina ingresados en el período", ingresados),
        ("Pendientes de gestionar hoy (Sin DTE, todo el histórico)", sin_dte_n),
        ("Pendientes que ya no aparecen en Autoventa (revisar)", fantasmas_n),
        ("Días entre el ingreso del pedido y su DTE (mediana)",
         float(dias_dte.median()) if len(dias_dte) else ""),
        ("", ""),
        ("GESTIÓN SEMANAL (meta del equipo)", ""),
        ("Meta de gestiones por semana (equipo)", meta_semanal),
        ("Gestiones por semana (ritmo del período)",
         round(tot / (_dias_rango / 7), 1) if _dias_rango else 0),
        ("% de la meta al ritmo del período",
         _pct(tot / (_dias_rango / 7), meta_semanal) if _dias_rango else 0),
        ("Semanas completas en el rango", n_semanas),
        ("Semanas completas que alcanzaron la meta", semanas_ok),
        ("% de semanas en meta", _pct(semanas_ok, n_semanas)),
        ("", ""),
        ("ESTADO DE ENTREGA (cruce con despachos de Autoventa)", ""),
        ("Entregadas / ejecutadas confirmadas", ent),
        ("Rechazadas", rech),
        ("Pendientes en ruta", en_ruta),
        ("Sin despacho (el mes tiene despachos, este documento no aparece)",
         sin_desp),
        ("Sin información (mes/sociedad sin despachos cargados)", sin_info),
        ("% Entregado (sobre lo que sí tiene información)", _pct(ent, con_info)),
        ("% Rechazo (sobre lo que sí tiene información)", _pct(rech, con_info)),
        ("Cobertura del cruce (movimientos con información)", _pct(con_info, tot)),
        ("", ""),
        ("CONVERSIÓN Y TIEMPOS", ""),
        ("Instalaciones nuevas confirmadas en terreno", nuevas_ent),
        ("% Conversión gestionada a entregada (nuevas)",
         _pct(nuevas_ent, nuevas_info)),
        ("Retiros confirmados en terreno", retiros_ent),
        ("% Retiros ejecutados", _pct(retiros_ent, retiros_info)),
        ("Días promedio entre factura y ruta de entrega",
         float(dias_ent.mean()) if len(dias_ent) else ""),
        ("Días máximo entre factura y ruta de entrega",
         float(dias_ent.max()) if len(dias_ent) else ""),
        ("", ""),
        ("ALCANCE", ""),
        ("Clientes distintos con movimiento", int(m["cliente_rut"].nunique())),
        ("Vendedores con movimientos", int(m["Vendedor"].nunique())),
    ]
    res = pd.DataFrame(ind, columns=["Indicador", "Valor"])
    ws = _escribir(wb, "Resumen", res,
                   nota=("Seguimiento de máquinas en comodato. El movimiento sale "
                         "de Obuma (código FL de la factura); el estado de entrega, "
                         "del Detalle de despachos de Autoventa cruzado por número "
                         "de documento."))
    for fila in range(2, ws.max_row + 1):
        etq = str(ws.cell(row=fila, column=1).value or "")
        if etq.startswith("%") or etq.startswith("Cobertura"):
            ws.cell(row=fila, column=2).number_format = _FMT_PCT
        elif etq.startswith("Días"):
            ws.cell(row=fila, column=2).number_format = _FMT_DEC

    # ── 2. Mensual ───────────────────────────────────────────────────────────
    orden_mes = m.drop_duplicates("Mes").sort_values("fecha")["Mes"].tolist()
    mensual = _agrupado(m, "Mes", "Mes", extra_clientes=True)
    if not mensual.empty:
        mensual["Mes"] = pd.Categorical(mensual["Mes"], categories=orden_mes,
                                        ordered=True)
        mensual = mensual.sort_values("Mes")
        mensual["Parque neto acumulado"] = mensual[
            "Parque neto (nuevas - retiros)"].cumsum()
        mensual["Mes"] = mensual["Mes"].astype(str)
        mensual = _con_total(mensual, "Mes",
                             _NO_SUM + ("Parque neto acumulado",))
    fmt_mes = dict(_FMT_AGRUP)
    fmt_mes["Parque neto acumulado"] = _FMT_NUM
    _escribir(wb, "Mensual", mensual, fmt_mes,
              nota=("Movimientos por mes de FACTURA. El parque neto acumulado suma "
                    "el saldo mes a mes dentro del período; no es el parque total "
                    "instalado, porque las máquinas colocadas antes de que empezara "
                    "este registro no están en ninguna base."),
              total_ultima=not mensual.empty)

    # ── 2b. Semanal (seguimiento de la meta de gestiones) ────────────────────
    sem = m.copy()
    sem["_sem"] = sem["fecha"].dt.to_period("W-SUN")
    grupos = dict(tuple(sem.groupby("_sem")))
    # Todas las semanas del rango, incluidas las que no tuvieron ni una gestión:
    # una semana en cero es justamente lo que hay que ver contra la meta.
    ing_sem = {}
    if not ped.empty:
        ing_sem = (ped["_ingreso"].dt.to_period("W-SUN")
                   .value_counts().to_dict())
    vacio = m.iloc[0:0]
    filas_sem = []
    for per, dias in _semanas_del_rango(f_ini, f_fin):
        g = grupos.get(per, vacio)
        fila = {"Semana": f"{per.start_time:%d/%m} al {per.end_time:%d/%m}",
                "_ord": per.start_time, "Días en el rango": dias,
                "Pedidos ingresados": int(ing_sem.get(per, 0))}
        fila.update(_conteos(g))
        # Las semanas cortadas por el borde del rango van sin meta (None, que
        # sale como celda vacía): exigirles 22 sería comparar contra días que
        # el informe no está mirando.
        fila["Meta semanal"] = meta_semanal if dias == 7 else None
        fila["% Meta"] = _pct(len(g), meta_semanal) if dias == 7 else None
        fila["Sobre / bajo la meta"] = (len(g) - meta_semanal if dias == 7
                                        else None)
        filas_sem.append(fila)
    semanal = pd.DataFrame(filas_sem)
    if not semanal.empty:
        semanal = (semanal.sort_values("_ord").drop(columns="_ord")
                   .rename(columns={"Movimientos": "Gestiones (con DTE)"}))
        n_completas = int((semanal["Días en el rango"] == 7).sum())
        semanal = _con_total(semanal, "Semana",
                             _NO_SUM + ("Meta semanal", "% Meta",
                                        "Días en el rango"))
        # La meta del total es meta × semanas COMPLETAS: las cortadas por el
        # borde del rango no se pueden exigir enteras.
        fin_i = semanal.index[-1]
        for c in ("Meta semanal", "% Meta"):
            semanal[c] = semanal[c].astype(object)
        semanal.loc[fin_i, "Meta semanal"] = (meta_semanal * n_completas
                                              if n_completas else None)
        semanal.loc[fin_i, "% Meta"] = (
            _pct(semanal.loc[fin_i, "Gestiones (con DTE)"],
                 meta_semanal * n_completas) if n_completas else None)
    fmt_sem = dict(_FMT_AGRUP)
    fmt_sem.update({"Gestiones (con DTE)": _FMT_NUM, "Meta semanal": _FMT_NUM,
                    "Días en el rango": _FMT_NUM, "Pedidos ingresados": _FMT_NUM,
                    "% Meta": _FMT_PCT, "Sobre / bajo la meta": _FMT_NUM})
    _escribir(wb, "Semanal", semanal, fmt_sem,
              nota=(f"Seguimiento de la meta de {meta_semanal} gestiones "
                    "semanales del equipo. Una gestión = un flete de máquina con "
                    "DTE emitido (instalación, cambio o retiro), contado en la "
                    "semana del documento. Los pedidos que siguen 'Sin DTE' no "
                    "suman aquí: aparecen en 'Pedidos ingresados' y en su "
                    "propia hoja. Semanas de lunes a "
                    "domingo; se omiten las que quedan con menos de 4 días "
                    "dentro del rango elegido."),
              total_ultima=not semanal.empty)

    # ── 3. Por vendedor ──────────────────────────────────────────────────────
    vend = _agrupado(m, "Vendedor", "Vendedor", extra_clientes=True)
    if not vend.empty:
        vend = vend.sort_values("Movimientos", ascending=False)
        vend = _con_total(vend, "Vendedor", _NO_SUM)
    _escribir(wb, "Por vendedor", vend, _FMT_AGRUP,
              nota=("Vendedor con que quedó atribuida la máquina: para Gran Natural "
                    "es el de Autoventa (quien la gestionó en terreno), no el de la "
                    "factura de Obuma."),
              total_ultima=not vend.empty)

    # ── 4. Tipo × Estado ─────────────────────────────────────────────────────
    ct = pd.crosstab(m["Movimiento"], m["Estado entrega"])
    ct = ct.reindex(columns=[e for e in ESTADOS if e in ct.columns])
    ct["Total"] = ct.sum(axis=1)
    ct = _con_total(ct.reset_index(), "Movimiento")
    _escribir(wb, "Tipo x Estado", ct,
              {c: _FMT_NUM for c in ct.columns if c != "Movimiento"},
              nota="Matriz de control: qué pasó con cada tipo de movimiento.",
              total_ultima=not ct.empty)

    # ── 5. Geografía ─────────────────────────────────────────────────────────
    geo = _agrupado(m, ["Región", "Comuna"], "Región", extra_clientes=True)
    if not geo.empty:
        geo = geo.sort_values(["Región", "Movimientos"], ascending=[True, False])
        geo = _con_total(geo, "Región", _NO_SUM + ("Comuna",))
    _escribir(wb, "Geografía", geo, _FMT_AGRUP,
              nota="Región y comuna del cliente, según dim_cliente.",
              total_ultima=not geo.empty)

    # ── 6. Detalle de movimientos ────────────────────────────────────────────
    _escribir(wb, "Detalle movimientos", _detalle(m), _FMT_DETALLE,
              nota=("Una fila por movimiento de máquina. Un 'Días factura a ruta' "
                    "negativo o muy grande delata un desfase entre la facturación y "
                    "la ruta que conviene revisar."))

    # ── 7. Pendientes por confirmar ──────────────────────────────────────────
    pend = m[m["_pendiente"]].sort_values("Días sin confirmar", ascending=False)
    _escribir(wb, "Pendientes por confirmar", _detalle(pend), _FMT_DETALLE,
              nota=("Movimientos facturados que NO están confirmados como "
                    "entregados: o el despacho sigue pendiente en ruta, o el "
                    "documento no aparece en el detalle de despachos. Ordenados por "
                    "antigüedad: los de arriba llevan más días abiertos."))

    # ── 8. Rechazadas ────────────────────────────────────────────────────────
    rechz = m[m["_rechazada"]].sort_values("fecha", ascending=False)
    _escribir(wb, "Rechazadas", _detalle(rechz), _FMT_DETALLE,
              nota=("El despacho salió a ruta y volvió rechazado: la máquina no se "
                    "instaló, o no se pudo retirar. Son las que hay que reprogramar."))

    # ── 9. Retiros ───────────────────────────────────────────────────────────
    ret = m[m["tipo_mov"] == "retiro"].sort_values("fecha", ascending=False)
    _escribir(wb, "Retiros", _detalle(ret), _FMT_DETALLE,
              nota=("Todos los FL-2 del período. 'Entregada' en un retiro significa "
                    "que la máquina volvió: el retiro se ejecutó. Si queda sin "
                    "confirmar, la máquina sigue en poder del cliente."))

    # ── 10. Clientes ─────────────────────────────────────────────────────────
    filas_cli = []
    for rut, g in m.groupby("cliente_rut", dropna=False):
        ult = g.sort_values("fecha").iloc[-1]
        filas_cli.append({
            "RUT": rut,
            "Cliente": ult["Cliente"],
            "Comuna": ult["Comuna"],
            "Región": ult["Región"],
            "Vendedor": ult["Vendedor"],
            "Nuevas (FL-4)": int((g["tipo_mov"] == "nueva").sum()),
            "Cambios (FL-1/3/5)": int((g["tipo_mov"] == "cambio").sum()),
            "Retiros (FL-2)": int((g["tipo_mov"] == "retiro").sum()),
            "Saldo (nuevas - retiros)": int((g["tipo_mov"] == "nueva").sum()
                                            - (g["tipo_mov"] == "retiro").sum()),
            "Último movimiento": (ult["fecha"].date()
                                  if pd.notna(ult["fecha"]) else None),
            "Tipo último movimiento": ult["Movimiento"],
            "Estado último movimiento": ult["Estado entrega"],
        })
    cli = pd.DataFrame(filas_cli)
    if not cli.empty:
        cli = cli.sort_values("Último movimiento", ascending=False)
    _escribir(wb, "Clientes", cli,
              {"Nuevas (FL-4)": _FMT_NUM, "Cambios (FL-1/3/5)": _FMT_NUM,
               "Retiros (FL-2)": _FMT_NUM, "Saldo (nuevas - retiros)": _FMT_NUM,
               "Último movimiento": _FMT_FECHA},
              nota=("Un cliente por fila. El saldo es del período, no el parque "
                    "instalado: las máquinas colocadas antes del inicio del dato no "
                    "están registradas en ningún sistema."))

    # ── 11-12. Despachos de Autoventa ────────────────────────────────────────
    if despachos is not None and not despachos.empty:
        docs_mov = set(m["_doc"])
        d = despachos.copy()
        d["fecha_ruta"] = pd.to_datetime(d["fecha_ruta"], errors="coerce")
        # Las hojas de despachos se acotan al período del informe: la ventana
        # ancha existe solo para cruzar el movimiento con su ruta, y si se
        # dejara entrar entera, aparecerían despachos de máquinas facturadas
        # fuera del período marcados como si les faltara la factura.
        d = d[(d["fecha_ruta"] >= pd.Timestamp(f_ini))
              & (d["fecha_ruta"] <= pd.Timestamp(f_fin))]
        d["_doc"] = _norm_doc(d["documento"])
        d["Mes ruta"] = [_mes(f) for f in d["fecha_ruta"]]
        d["_es_maq_mov"] = d["_doc"].isin(docs_mov)
        marca_etl = (d["es_maquina"].fillna(False).astype(bool)
                     if "es_maquina" in d.columns
                     else pd.Series(False, index=d.index))
        # También los que salieron con transportista de máquinas aunque no
        # crucen con Obuma: son máquinas despachadas y todavía sin facturar.
        transp_maq = (d["transportista"].astype(str)
                      .str.contains(_TRANSP_MAQUINA, case=False, na=False)
                      if "transportista" in d.columns
                      else pd.Series(False, index=d.index))

        # Un documento puede estar facturado FUERA del período (se factura el 31
        # y sale a ruta el 3). Sin este contexto se le acusaría de no tener
        # factura, que es un problema muy distinto.
        ctx = m
        if maquinas_ctx is not None and not maquinas_ctx.empty:
            c = maquinas_ctx.copy()
            c["_doc"] = _norm_doc(c["documento"])
            c["fecha"] = pd.to_datetime(c["fecha"], errors="coerce")
            c["Movimiento"] = c["tipo_mov"].map(MOV_LBL).fillna(c["tipo_mov"])
            for col in ("Vendedor", "Cliente", "Comuna", "Región"):
                if col not in c.columns:
                    c[col] = None
            ctx = pd.concat([m, c], ignore_index=True).drop_duplicates("_doc")
        docs_facturados = set(ctx["_doc"])
        mov_por_doc = (ctx.drop_duplicates("_doc")
                       .set_index("_doc")[["Movimiento", "Vendedor", "Cliente",
                                           "Comuna", "Región", "fecha"]])
        dm = (d[d["_doc"].isin(docs_facturados) | marca_etl | transp_maq]
              .join(mov_por_doc, on="_doc"))
        dm["_sin_factura"] = ~dm["_doc"].isin(docs_facturados)
        dm["_otro_periodo"] = (~dm["_sin_factura"]) & (~dm["_es_maq_mov"])
        nom_vend = (dict(zip(vendedores["id"], vendedores["nombre_canonico"]))
                    if vendedores is not None and not vendedores.empty else {})
        desp_det = pd.DataFrame({
            "Fecha ruta": dm["fecha_ruta"].dt.date,
            "Mes ruta": dm["Mes ruta"],
            "Documento": dm["_doc"],
            "Movimiento": dm["Movimiento"].fillna(
                "AVISO: despachada y sin factura en Obuma"),
            "Vendedor": dm["Vendedor"].fillna(
                dm["vendedor_id"].map(nom_vend)).fillna("Sin asignar"),
            "Fecha factura": dm["fecha"].dt.date,
            "Factura": pd.Series(
                ["Sin factura (Sin DTE)" if sf else
                 ("De otro período" if op else "Del período")
                 for sf, op in zip(dm["_sin_factura"], dm["_otro_periodo"])],
                index=dm.index),
            "Estado despacho": dm["estado"],
            "Transportista": dm.get("transportista"),
            "Devolución": (dm["devolucion"].map({True: "Sí", False: "No"})
                           if "devolucion" in dm.columns else ""),
            "Peso (Kg)": pd.to_numeric(dm.get("peso"), errors="coerce"),
            "RUT": dm["cliente_rut"],
            "Cliente": dm["Cliente"].fillna(_desc(dm["cliente_rut"], clientes,
                                                  "razon_social")),
            "Comuna": dm["Comuna"].fillna(_desc(dm["cliente_rut"], clientes,
                                                "comuna")),
            "Región": dm["Región"].fillna(_desc(dm["cliente_rut"], clientes,
                                                "region")),
        }).sort_values("Fecha ruta", ascending=False)
        _escribir(wb, "Despachos de máquina", desp_det,
                  {"Fecha ruta": _FMT_FECHA, "Fecha factura": _FMT_FECHA,
                   "Peso (Kg)": _FMT_DEC},
                  nota=("Filas del Detalle de despachos de Autoventa que "
                        "corresponden a una máquina: las que cruzan con un "
                        "movimiento de Obuma, más las que salieron con un "
                        "transportista de máquinas. Las marcadas 'AVISO: "
                        "despachada y sin factura en Obuma' son máquinas que ya "
                        "salieron a ruta y cuyo flete todavía no se factura — son "
                        "la diferencia entre este informe y el conteo directo de "
                        "Autoventa. Ver la hoja 'Sin facturar (Autoventa)'."))

        # ── 11b. Conciliación con el conteo directo de Autoventa ────────────
        # La pregunta que aparece cada semana: "yo cuento N y el informe M".
        # Casi siempre son las mismas dos causas, así que van con su aritmética.
        n_mov = len(m)
        rutados_periodo = int(m["Fecha ruta"].between(
            pd.Timestamp(f_ini), pd.Timestamp(f_fin)).sum())
        n_otro = int(dm["_otro_periodo"].sum())
        n_sin_fact = int(dm["_sin_factura"].sum())
        conc = pd.DataFrame([
            ("Movimientos facturados con fecha de factura en el período",
             n_mov, "Es el número de la hoja 'Detalle movimientos'."),
            ("(−) De esos, sin despacho en el período (no salieron a ruta, o "
             "la ruta cae fuera)", n_mov - rutados_periodo,
             "Pendientes, sin despacho, o entregados en otro período."),
            ("(=) Facturados en el período y despachados en el período",
             rutados_periodo, ""),
            ("(+) Despachados en el período con factura de otro período",
             n_otro, "Se facturó antes (fin de mes) y salió a ruta ahora."),
            ("(+) Despachados en el período SIN factura en Obuma",
             n_sin_fact, "Pedido 'Sin DTE': ver hoja 'Sin facturar (Autoventa)'."),
            ("(=) Total de despachos de máquina del período",
             rutados_periodo + n_otro + n_sin_fact,
             "Esto es lo que cuenta el Detalle de despachos de Autoventa."),
        ], columns=["Concepto", "N° documentos", "Explicación"])
        _escribir(wb, "Conciliación Autoventa", conc,
                  {"N° documentos": _FMT_NUM},
                  nota=("Por qué este informe y el conteo directo de Autoventa "
                        "dan distinto. Los dos están bien: cuentan cosas "
                        "distintas. Este cuenta la máquina cuando Obuma factura "
                        "el flete; Autoventa, cuando sale a ruta. Ojo: Acuña no "
                        "pasa por Autoventa, así que sus movimientos entran en la "
                        "primera fila y no en el despacho."))

        def _cuenta(serie, valor):
            return int((serie.astype(str).str.lower() == valor).sum())

        gen = (d.groupby("Mes ruta")
               .agg(**{"Despachos": ("documento", "count"),
                       "Entregados": ("estado", lambda s: _cuenta(s, "entregada")),
                       "Rechazados": ("estado", lambda s: _cuenta(s, "rechazada")),
                       "Pendientes": ("estado", lambda s: _cuenta(s, "pendiente")),
                       "De máquina": ("_es_maq_mov", "sum")})
               .reset_index())
        gen["% Rechazo"] = gen["Rechazados"] / gen["Despachos"].replace(0, pd.NA)
        orden_r = (d.drop_duplicates("Mes ruta").sort_values("fecha_ruta")
                   ["Mes ruta"].tolist())
        gen["Mes ruta"] = pd.Categorical(gen["Mes ruta"], categories=orden_r,
                                         ordered=True)
        gen = gen.sort_values("Mes ruta")
        gen["Mes ruta"] = gen["Mes ruta"].astype(str)
        gen["De máquina"] = gen["De máquina"].astype(int)
        gen = _con_total(gen, "Mes ruta", ("% Rechazo",))
        _escribir(wb, "Despachos (contexto)", gen,
                  {"Despachos": _FMT_NUM, "Entregados": _FMT_NUM,
                   "Rechazados": _FMT_NUM, "Pendientes": _FMT_NUM,
                   "De máquina": _FMT_NUM, "% Rechazo": _FMT_PCT},
                  nota=("TODOS los despachos del período (no solo máquinas), "
                        "para poner en contexto el nivel de rechazo de las rutas."),
                  total_ultima=not gen.empty)

    # ── 12b. Pendientes de gestionar (ingresados y sin DTE) ──────────────────
    if not ped.empty:
        nom_v = (dict(zip(vendedores["id"], vendedores["nombre_canonico"]))
                 if vendedores is not None and not vendedores.empty else {})
        sin_dte = ped[ped["_sin_dte"]].copy()
        if not sin_dte.empty:
            tabla_sd = pd.DataFrame({
                "Fecha del pedido": sin_dte["_ingreso"].dt.date,
                "N° pedido": sin_dte["n_pedido"],
                "Código FL": sin_dte["_cod"],
                "Movimiento": sin_dte["_mov"].map(MOV_LBL).fillna("(otro)"),
                "Vendedor": sin_dte["vendedor_id"].map(nom_v).fillna("Sin asignar"),
                "RUT": sin_dte["cliente_rut"],
                "Cliente": _desc(sin_dte["cliente_rut"], clientes, "razon_social"),
                "Comuna": _desc(sin_dte["cliente_rut"], clientes, "comuna"),
                "Estado en Autoventa": sin_dte["estado_pedido"].fillna(
                    "no aparece en la API"),
                "Sigue vigente": sin_dte["_fantasma"].map({True: "revisar",
                                                           False: "sí"}),
                "Días sin gestionar": (pd.Timestamp(hoy)
                                       - sin_dte["_ingreso"]).dt.days,
            }).sort_values("Días sin gestionar", ascending=False)
            _escribir(wb, "Pendientes de gestionar", tabla_sd,
                      {"Fecha del pedido": _FMT_FECHA,
                       "Días sin gestionar": _FMT_NUM},
                      nota=("Pedidos de flete que el vendedor YA ingresó en "
                            "Autoventa y que todavía no tienen DTE emitido: la "
                            "gestión está pedida y no se ha concretado. No suman "
                            "a la meta semanal (esa cuenta gestiones con DTE), "
                            "pero son la cola de la que salen las próximas. "
                            "Salen todos los que siguen abiertos hoy, sin "
                            "importar el período del informe: un pedido de hace "
                            "tres meses sin gestionar es el que más urge. Los "
                            "marcados 'revisar' en Sigue vigente ya no vienen en "
                            "la API de Autoventa: lo más probable es que se "
                            "anularan o se reingresaran con otro número, así que "
                            "no cuentan como cola real."))

        # Gestión por vendedor: lo que pidió, lo que se le concretó y su cola.
        g_ing = ped.groupby("vendedor_id").size().rename("Pedidos ingresados")
        g_dte = (ped[~ped["_sin_dte"]].groupby("vendedor_id").size()
                 .rename("Gestionados (con DTE)"))
        g_pen = (ped[ped["_sin_dte"] & ~ped["_fantasma"]]
                 .groupby("vendedor_id").size().rename("Pendientes (Sin DTE)"))
        g_dias = (ped.groupby("vendedor_id")["_dias_a_dte"].median()
                  .rename("Días ingreso a DTE (mediana)"))
        gest = pd.concat([g_ing, g_dte, g_pen, g_dias], axis=1).fillna(0)
        gest.insert(0, "Vendedor",
                    pd.Series(gest.index, index=gest.index).map(nom_v)
                    .fillna("Sin asignar"))
        gest["% Concretado"] = gest["Gestionados (con DTE)"] / gest[
            "Pedidos ingresados"].replace(0, pd.NA)
        for c in ("Pedidos ingresados", "Gestionados (con DTE)",
                  "Pendientes (Sin DTE)"):
            gest[c] = gest[c].astype(int)
        gest = gest.reset_index(drop=True).sort_values("Pedidos ingresados",
                                                       ascending=False)
        gest = _con_total(gest, "Vendedor",
                          ("% Concretado", "Días ingreso a DTE (mediana)"))
        _escribir(wb, "Gestión por vendedor", gest,
                  {"Pedidos ingresados": _FMT_NUM,
                   "Gestionados (con DTE)": _FMT_NUM,
                   "Pendientes (Sin DTE)": _FMT_NUM,
                   "Días ingreso a DTE (mediana)": _FMT_DEC,
                   "% Concretado": _FMT_PCT},
                  nota=("Todos los pedidos de flete del histórico por vendedor: "
                        "cuántos ingresó, cuántos se concretaron con DTE, "
                        "cuántos siguen en cola y cuánto demora en promedio la "
                        "gestión. El % concretado no es solo mérito del "
                        "vendedor: depende de que se emita el documento."),
                  total_ultima=not gest.empty)

    # ── 13. Control del cruce ────────────────────────────────────────────────
    ctrl = []
    for (soc, mes_lbl), g in m.groupby(["Sociedad", "Mes"], dropna=False):
        con_i = int(len(g) - g["_sin_info"].sum())
        ctrl.append({
            "Sociedad": soc,
            "Mes": mes_lbl,
            "Movimientos": int(len(g)),
            "Con despacho cruzado": int(g["_entregada"].sum()
                                        + g["_rechazada"].sum()
                                        + (g["Estado entrega"] == EN_RUTA).sum()),
            "Sin despacho": int((g["Estado entrega"] == SIN_DESPACHO).sum()),
            "Sin información": int(g["_sin_info"].sum()),
            "% Cruce": _pct(con_i, len(g)),
        })
    ctrl = pd.DataFrame(ctrl)
    if not ctrl.empty:
        ctrl["Mes"] = pd.Categorical(ctrl["Mes"], categories=orden_mes,
                                     ordered=True)
        ctrl = ctrl.sort_values(["Sociedad", "Mes"])
        ctrl["Mes"] = ctrl["Mes"].astype(str)
        ctrl = _con_total(ctrl, "Sociedad", ("% Cruce", "Mes"))
    _escribir(wb, "Control del cruce", ctrl,
              {"Movimientos": _FMT_NUM, "Con despacho cruzado": _FMT_NUM,
               "Sin despacho": _FMT_NUM, "Sin información": _FMT_NUM,
               "% Cruce": _FMT_PCT},
              nota=("Calidad del cruce Obuma con Autoventa por mes y sociedad. "
                    "Acuña no pasa por Autoventa, así que sus movimientos salen "
                    "siempre 'Sin información': no es un error del dato, es que ese "
                    "despacho no existe."),
              total_ultima=not ctrl.empty)

    # ── 14. Definiciones ─────────────────────────────────────────────────────
    defs = pd.DataFrame([
        ("Movimiento de máquina",
         "Una línea de flete de la categoría 'Maquinas' en una factura de Obuma. "
         "El código FL dice qué pasó."),
        ("FL-4 · Nueva",
         "Instalación en cliente nuevo. Es la 'máquina gestionada' que se le "
         "cuenta al vendedor."),
        ("FL-1 / FL-3 / FL-5 · Cambio",
         "Cambio de máquina: FL-1 cambio, FL-3 por máquina mala, FL-5 por cambio "
         "de tamaño. El cliente sigue con máquina."),
        ("FL-2 · Retiro", "Retiro por término del comodato."),
        ("Parque neto",
         "Nuevas menos retiros del período. Es el crecimiento, NO el total de "
         "máquinas en calle: las colocadas antes de que empezara este registro no "
         "están en ninguna base."),
        (f"Estado {ENTREGADA}",
         "El documento aparece Entregada en el Detalle de despachos de Autoventa: "
         "el movimiento se ejecutó en terreno."),
        (f"Estado {RECHAZADA}",
         "Salió a ruta y volvió rechazado: no se instaló, o no se retiró."),
        (f"Estado {EN_RUTA}",
         "El despacho existe pero sigue marcado Pendiente."),
        (f"Estado {SIN_DESPACHO}",
         "Ese mes y sociedad SÍ tienen despachos cargados, pero este documento no "
         "aparece en ninguno. Es lo que hay que ir a buscar."),
        (f"Estado {SIN_INFO}",
         "No hay despachos cargados para ese mes y sociedad, así que el movimiento "
         "no se puede confirmar ni desmentir. Acuña nunca los tiene (no pasa por "
         "Autoventa) y antes de feb-2026 no existe el dato."),
        ("% Entregado",
         "Entregadas dividido por los movimientos con información. Deja fuera lo "
         "'Sin información' para no castigar al vendedor por un archivo que no se "
         "cargó."),
        ("Máquinas (cant.)",
         "Cantidad facturada en la línea FL: normalmente 1, pero hay documentos "
         "que instalan 2 máquinas."),
        ("Movimiento anulado",
         "El documento es una nota de crédito que revierte un flete de máquina. "
         "Sigue contando como movimiento en la base; la columna 'Tipo documento' "
         "del detalle los deja a la vista."),
        ("Gestionada en Autoventa sin factura",
         "El vendedor ingresó el flete en Autoventa (y la máquina puede haber "
         "salido a ruta), pero Obuma todavía no emite el documento. No cuenta "
         "como movimiento aquí, porque el movimiento se deriva de la factura. "
         "Es la razón habitual de que Autoventa muestre más máquinas que este "
         "informe; la hoja 'Sin facturar (Autoventa)' las lista una por una."),
        ("Llave del cruce",
         "Obuma 'N° DCTO' = Autoventa 'Documento'. Si un documento tiene varios "
         "despachos manda el mejor resultado (Entregada, luego Rechazada, luego "
         "Pendiente) y la columna 'Intentos de despacho' muestra cuántos hubo."),
        ("Vendedor",
         "Para Gran Natural, el vendedor de Autoventa (quien gestionó la máquina "
         "en terreno); Obuma suele dejar estos documentos sin vendedor asignado."),
    ], columns=["Concepto", "Definición"])
    _escribir(wb, "Definiciones", defs,
              nota="Las mismas definiciones que usa la app en pantalla (CLAUDE.md §3).")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
