"""Export Excel EXPLICATIVO de comisiones (simulador).

A diferencia de `export_comisiones.comisiones_a_excel`, que entrega los montos ya
calculados, este arma un libro donde **cada monto es una fórmula** colgada de sus
insumos. Gerencia puede cambiar un objetivo, una cartera o los días trabajados y
ver al instante cómo se mueve la comisión, sin tocar el sistema.

Tres hojas:
  · Guía       — qué es cada indicador, cómo se calcula y cómo usar el libro.
  · Simulador  — una fila por vendedor. Amarillo = editable, negro = fórmula.
  · Escalas    — los tramos y parámetros LEÍDOS DE LA BASE (no hardcodeados),
                 que es de donde tiran los INDEX/MATCH.

Las fórmulas replican `v_comision_vendedor_mes` (sql/006 + 009 + 011 + 014):
PNV redondea hacia abajo al 5% (piso 80%, techo 110%); Máquinas al 5% más
cercano (piso 25%, techo 140%); Efectividad al 10% más cercano (piso 30%, techo
60%) cruzada con el rango de cartera. El bono de reposición queda FUERA de la
base de la Semana Corrida, igual que en el sistema.
"""
import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MESES = {1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
         7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre",
         12: "Diciembre"}

_A = "Arial"
F_TXT = Font(name=_A, size=10)
F_IN = Font(name=_A, size=10, color="0000FF")
F_HD = Font(name=_A, size=9, bold=True, color="FFFFFF")
F_B = Font(name=_A, size=10, bold=True)
F_TIT = Font(name=_A, size=13, bold=True, color="1B3A6B")
F_SEC = Font(name=_A, size=11, bold=True, color="1B3A6B")
F_IT = Font(name=_A, size=9, italic=True, color="595959")
FILL_HD = PatternFill("solid", fgColor="1B3A6B")
FILL_IN = PatternFill("solid", fgColor="FFF2CC")     # celda editable
FILL_TOT = PatternFill("solid", fgColor="FCE4EC")
_S = Side(style="thin", color="D9D9D9")
BD = Border(left=_S, right=_S, top=_S, bottom=_S)
CLP = '$#,##0;($#,##0);-'
PC1 = '0.0%'
PC0 = '0%'

# (encabezado, clave del df, tipo, formato)
#   in  = insumo editable   ·   fx = fórmula   ·   txt = texto
COLUMNAS = [
    ("Vendedor",            "nombre_canonico",     "txt", None),
    ("Escala",              "_escala",             "txt", None),
    ("Plan",                "plan_id",             "in",  '0'),
    ("Objetivo de venta",   "obj_venta",           "in",  CLP),
    ("Fact-NC",             "fact_nc",             "in",  CLP),
    ("% PNV",               None,                  "fx",  PC1),
    ("Tramo PNV",           None,                  "fx",  PC0),
    ("Comisión PNV",        None,                  "fx",  CLP),
    ("Bono 4%",             None,                  "fx",  CLP),
    ("Obj. máquinas",       "obj_maquinas",        "in",  '0'),
    ("Máq. entregadas",     "maquinas_entregadas", "in",  '0'),
    ("% Máquinas",          None,                  "fx",  PC1),
    ("Tramo Máq.",          None,                  "fx",  PC0),
    ("Comisión Máquinas",   None,                  "fx",  CLP),
    ("Obj. visitas",        "obj_visitas",         "in",  '0'),
    ("N° facturas",         "n_facturas",          "in",  '0'),
    ("% Efectividad",       None,                  "fx",  PC1),
    ("Tramo Efec.",         None,                  "fx",  PC0),
    ("Cartera clientes",    "cartera_clientes",    "in",  '0'),
    ("Comisión Efectividad", None,                 "fx",  CLP),
    ("TOTAL COMISIÓN",      None,                  "fx",  CLP),
    ("Días trabajados",     "dias_trabajados",     "in",  '0'),
    ("INAB",                "inab",                "in",  '0'),
    ("Semana Corrida",      None,                  "fx",  CLP),
    ("Salas Ganga",         "salas_ganga",         "in",  '0'),
    ("Bono Reposición",     None,                  "fx",  CLP),
    ("TOTAL A PAGAR",       None,                  "fx",  CLP),
]
C = {h: get_column_letter(i) for i, (h, *_) in enumerate(COLUMNAS, start=1)}
ANCHOS = {"Vendedor": 32, "Escala": 12, "Plan": 6}


def _num(v, default=0):
    try:
        f = float(v)
        return default if pd.isna(f) else f
    except (TypeError, ValueError):
        return default


def _escalas(wb, t_pnv, t_maq, t_ef, params):
    """Escribe los tramos por plan y devuelve las referencias de rango."""
    ws = wb.create_sheet("Escalas")
    ws["A1"] = "Escalas y parámetros vigentes"
    ws["A1"].font = F_TIT
    ws["A2"] = ("Leídos de la base de datos al generar este archivo. Son la fuente de los "
                "INDEX/MATCH de la hoja Simulador: si los editas, el simulador cambia, "
                "pero el sistema NO. Para cambiarlos de verdad: Comisiones → Escalas y parámetros.")
    ws["A2"].font = F_IT

    ref = {}
    ws["A4"] = "Parámetros"
    ws["A4"].font = F_SEC
    fila = 5
    for clave, val in params.items():
        ws.cell(fila, 1, clave).font = F_TXT
        c = ws.cell(fila, 2, float(val))
        c.font = F_IN
        c.number_format = '0.0000' if float(val) < 100 else CLP
        ref[f"par_{clave}"] = f"Escalas!$B${fila}"
        fila += 1

    col = 4
    for plan in (1, 2):
        # ── PNV y Máquinas (logro → monto) ────────────────────────────────
        for nombre, tabla in (("PNV", t_pnv), ("MAQ", t_maq)):
            sub = tabla[tabla["plan_id"] == plan].sort_values("logro_pct")
            ws.cell(4, col, f"{nombre} · plan {plan}").font = F_SEC
            ws.cell(5, col, "Logro").font = F_B
            ws.cell(5, col + 1, "Monto").font = F_B
            for i, (_, r) in enumerate(sub.iterrows()):
                a = ws.cell(6 + i, col, float(r["logro_pct"]))
                a.number_format = PC0
                a.font = F_TXT
                b = ws.cell(6 + i, col + 1, float(r["monto"]))
                b.number_format = CLP
                b.font = F_TXT
            n = len(sub)
            L1, L2 = get_column_letter(col), get_column_letter(col + 1)
            ref[f"{nombre.lower()}{plan}_l"] = f"Escalas!${L1}$6:${L1}${5 + n}"
            ref[f"{nombre.lower()}{plan}_m"] = f"Escalas!${L2}$6:${L2}${5 + n}"
            col += 3

        # ── Efectividad (matriz cartera × %) ──────────────────────────────
        sub = t_ef[t_ef["plan_id"] == plan]
        carteras = sorted(sub["cartera_min"].unique())
        pcts = sorted(sub["efectividad_pct"].unique())
        ws.cell(4, col, f"EFECTIVIDAD · plan {plan}").font = F_SEC
        ws.cell(5, col, "Cartera desde").font = F_B
        for j, p in enumerate(pcts):
            c = ws.cell(5, col + 1 + j, float(p))
            c.number_format = PC0
            c.font = F_B
        piv = sub.pivot_table(index="cartera_min", columns="efectividad_pct", values="monto")
        for i, cm in enumerate(carteras):
            ws.cell(6 + i, col, int(cm)).font = F_TXT
            for j, p in enumerate(pcts):
                v = piv.loc[cm, p] if p in piv.columns else None
                c = ws.cell(6 + i, col + 1 + j, float(v) if pd.notna(v) else 0)
                c.number_format = CLP
                c.font = F_TXT
        Lc = get_column_letter(col)
        Lp1, Lp2 = get_column_letter(col + 1), get_column_letter(col + len(pcts))
        ref[f"ef{plan}_c"] = f"Escalas!${Lc}$6:${Lc}${5 + len(carteras)}"
        ref[f"ef{plan}_p"] = f"Escalas!${Lp1}$5:${Lp2}$5"
        ref[f"ef{plan}_m"] = f"Escalas!${Lp1}$6:${Lp2}${5 + len(carteras)}"
        col += len(pcts) + 3

    ws.column_dimensions["A"].width = 20
    for i in range(2, col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13
    return ref


def _guia(wb, anio, mes):
    ws = wb.create_sheet("Guía")
    ws["A1"] = f"Comisiones {MESES[mes]} {anio} — libro explicativo"
    ws["A1"].font = F_TIT
    bloques = [
        ("Para qué sirve",
         ["La hoja Simulador tiene una fila por vendedor con TODOS los insumos del cálculo y "
          "todos los montos como fórmulas.",
          "Cambia una celda amarilla y los montos de esa fila se recalculan solos. Sirve para "
          "responder «¿cuánto pagaría si le subo la meta?» o «¿cuánto cuesta bajar el piso?» "
          "sin tocar el sistema.",
          "IMPORTANTE: este archivo NO escribe en el sistema. Lo que edites aquí queda aquí."]),
        ("Cómo leerlo",
         ["Celda AMARILLA con número azul = insumo, lo puedes cambiar.",
          "Celda blanca con número negro = fórmula, no la toques (se recalcula sola).",
          "La hoja Escalas trae los tramos y parámetros tal como están hoy en la base de datos."]),
        ("Los cuatro componentes de la comisión",
         ["PNV — logro = Fact-NC ÷ objetivo de venta. El tramo se redondea HACIA ABAJO al 5%: "
          "con 104% pagas el tramo de 100%. Bajo 80% no paga; sobre 110% se congela en 110%.",
          "Bono 4% — solo si el logro llega a 110%: 4% de lo que exceda ese 110% del objetivo. "
          "No tiene tope.",
          "Máquinas — logro = entregadas ÷ objetivo. Se redondea al 5% MÁS CERCANO (no hacia "
          "abajo). Bajo 25% no paga; sobre 140% se congela en 140%.",
          "Efectividad — logro = N° de facturas ÷ objetivo de visitas, redondeado al 10% más "
          "cercano. Bajo 30% no paga y sobre 60% se congela. El monto además depende del rango "
          "de cartera: con menos de 81 clientes asignados no paga aunque el % alcance."]),
        ("Semana Corrida y Total a Pagar",
         ["Semana Corrida = TOTAL COMISIÓN ÷ días trabajados × INAB. Si la comisión es $0, la "
          "semana corrida también es $0: no hay remuneración variable que repartir.",
          "Días trabajados e INAB salen del calendario del mes, salvo que el vendedor haya "
          "trabajado solo parte del mes; en ese caso van sus días reales y los descansos de ese "
          "tramo (se cargan en Comisiones → Editar entradas del período → Mes parcial).",
          "Bono Reposición = $15.000 por sala Ganga. Va POR FUERA de la base de la Semana "
          "Corrida: se suma al final y no genera semana corrida (decisión de gerencia).",
          "TOTAL A PAGAR = Total Comisión + Semana Corrida + Bono Reposición."]),
        ("Dónde se cambia de verdad cada cosa",
         ["Objetivos de venta, máquinas y visitas → Panel Gerencia, editor de objetivos.",
          "Cartera, salas Ganga, mes parcial y forzado de tramos → Comisiones → Editar entradas "
          "del período.",
          "Las escalas y los parámetros (el 4%, el $15.000) → Comisiones → Escalas y parámetros."]),
    ]
    r = 3
    for titulo, lineas in bloques:
        ws.cell(r, 1, titulo).font = F_SEC
        r += 1
        for ln in lineas:
            c = ws.cell(r, 1, ln)
            c.font = F_TXT
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 30
            r += 1
        r += 1
    ws.column_dimensions["A"].width = 118
    return ws


def comisiones_simulador_xlsx(df: pd.DataFrame, anio: int, mes: int,
                              tramos_pnv: pd.DataFrame, tramos_maq: pd.DataFrame,
                              tramos_ef: pd.DataFrame, parametros: dict) -> bytes:
    """Libro explicativo/simulador. `parametros` = {clave: valor}."""
    wb = Workbook()
    wb.remove(wb.active)
    _guia(wb, anio, mes)
    ws = wb.create_sheet("Simulador")
    ref = _escalas(wb, tramos_pnv, tramos_maq, tramos_ef, parametros)

    ws["A1"] = f"Simulador de comisiones — {MESES[mes]} {anio}"
    ws["A1"].font = F_TIT
    ws["A2"] = ("Amarillo = insumo editable.  Blanco = fórmula.  Cambia un amarillo y la fila se "
                "recalcula.  Este archivo no modifica el sistema.")
    ws["A2"].font = F_IT

    hdr = 4
    for j, (h, _k, _t, _f) in enumerate(COLUMNAS, start=1):
        c = ws.cell(hdr, j, h)
        c.font = F_HD
        c.fill = FILL_HD
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        c.border = BD
    ws.row_dimensions[hdr].height = 34

    d = df.sort_values("nombre_canonico")
    r = hdr + 1
    for _, row in d.iterrows():
        plan = int(_num(row.get("plan_id"), 1)) or 1
        vals = {
            "nombre_canonico": row.get("nombre_canonico"),
            "_escala": "Macarena" if plan == 2 else "Normal",
            "plan_id": plan,
            "obj_venta": _num(row.get("obj_venta")),
            "fact_nc": _num(row.get("fact_nc")),
            "obj_maquinas": _num(row.get("obj_maquinas")),
            "maquinas_entregadas": _num(row.get("maquinas_entregadas")),
            "obj_visitas": _num(row.get("obj_visitas")),
            "n_facturas": _num(row.get("n_facturas")),
            "cartera_clientes": _num(row.get("cartera_clientes")),
            "dias_trabajados": _num(row.get("dias_trabajados")),
            "inab": _num(row.get("inab")),
            "salas_ganga": _num(row.get("salas_ganga")),
        }
        f = _formulas(r, plan, ref)
        for j, (h, k, tipo, fmt) in enumerate(COLUMNAS, start=1):
            cel = ws.cell(r, j)
            if tipo == "fx":
                cel.value = f[h]
                cel.font = F_B if h.startswith("TOTAL") else F_TXT
            else:
                cel.value = vals.get(k)
                cel.font = F_TXT if tipo == "txt" else F_IN
                if tipo == "in":
                    cel.fill = FILL_IN
            if fmt:
                cel.number_format = fmt
            cel.border = BD
            if h.startswith("TOTAL"):
                cel.fill = FILL_TOT
        r += 1

    # Fila de totales
    ult = r - 1
    ws.cell(r, 1, "TOTAL").font = F_B
    for h in ("Comisión PNV", "Bono 4%", "Comisión Máquinas", "Comisión Efectividad",
              "TOTAL COMISIÓN", "Semana Corrida", "Bono Reposición", "TOTAL A PAGAR",
              "Objetivo de venta", "Fact-NC"):
        col = C[h]
        cel = ws.cell(r, ws[f"{col}{hdr}"].column, f"=SUM({col}{hdr + 1}:{col}{ult})")
        cel.number_format = CLP
        cel.font = F_B
        cel.fill = FILL_TOT
        cel.border = BD

    for j, (h, *_rest) in enumerate(COLUMNAS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = ANCHOS.get(h, 13)
    ws.freeze_panes = ws.cell(hdr + 1, 2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _formulas(r, plan, ref):
    """Fórmulas de la fila `r` para un vendedor del plan `plan`."""
    o, f_ = f"{C['Objetivo de venta']}{r}", f"{C['Fact-NC']}{r}"
    pnv, tpnv = f"{C['% PNV']}{r}", f"{C['Tramo PNV']}{r}"
    mo, me = f"{C['Obj. máquinas']}{r}", f"{C['Máq. entregadas']}{r}"
    maq, tmaq = f"{C['% Máquinas']}{r}", f"{C['Tramo Máq.']}{r}"
    vo, nf = f"{C['Obj. visitas']}{r}", f"{C['N° facturas']}{r}"
    ef, tef = f"{C['% Efectividad']}{r}", f"{C['Tramo Efec.']}{r}"
    cart = f"{C['Cartera clientes']}{r}"
    tot = f"{C['TOTAL COMISIÓN']}{r}"
    dt, inab = f"{C['Días trabajados']}{r}", f"{C['INAB']}{r}"
    p = f"pnv{plan}"
    m = f"maq{plan}"
    e = f"ef{plan}"
    return {
        "% PNV": f"=IF({o}=0,\"\",{f_}/{o})",
        "Tramo PNV": f"=IF({pnv}=\"\",\"\",ROUND(MIN(1.1,FLOOR({pnv},0.05)),2))",
        "Comisión PNV": (f"=IF(OR({tpnv}=\"\",{tpnv}<0.8),0,"
                         f"INDEX({ref[p + '_m']},MATCH({tpnv},{ref[p + '_l']},1)))"),
        "Bono 4%": (f"=IF(AND({pnv}<>\"\",{pnv}>={ref['par_bono_umbral']}),"
                    f"ROUND({ref['par_bono_pct']}*MAX(0,{f_}-{ref['par_bono_umbral']}*{o}),0),0)"),
        "% Máquinas": f"=IF({mo}=0,\"\",{me}/{mo})",
        "Tramo Máq.": f"=IF({maq}=\"\",\"\",ROUND(MIN(1.4,ROUND({maq}/0.05,0)*0.05),2))",
        "Comisión Máquinas": (f"=IF(OR({tmaq}=\"\",{tmaq}<0.25),0,"
                              f"INDEX({ref[m + '_m']},MATCH({tmaq},{ref[m + '_l']},1)))"),
        "% Efectividad": f"=IF({vo}=0,\"\",{nf}/{vo})",
        "Tramo Efec.": f"=IF({ef}=\"\",\"\",ROUND(MIN(0.6,ROUND({ef}/0.1,0)*0.1),2))",
        "Comisión Efectividad": (f"=IF(OR({tef}=\"\",{tef}<0.3,{cart}<81),0,"
                                 f"INDEX({ref[e + '_m']},MATCH({cart},{ref[e + '_c']},1),"
                                 f"MATCH({tef},{ref[e + '_p']},1)))"),
        "TOTAL COMISIÓN": (f"={C['Comisión PNV']}{r}+{C['Bono 4%']}{r}"
                           f"+{C['Comisión Máquinas']}{r}+{C['Comisión Efectividad']}{r}"),
        "Semana Corrida": f"=IF({dt}=0,0,ROUND({tot}/{dt}*{inab},0))",
        "Bono Reposición": f"={ref['par_reposicion_monto']}*{C['Salas Ganga']}{r}",
        "TOTAL A PAGAR": (f"={tot}+{C['Semana Corrida']}{r}+{C['Bono Reposición']}{r}"),
    }
