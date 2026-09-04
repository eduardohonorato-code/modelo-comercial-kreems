"""
Informe de entregas: cuánto de lo que salió a ruta llegó, en pesos y en máquinas.

Mide la efectividad del despacho por transportista. Se apoya en dos números que
NO se pueden mezclar, porque miden cosas distintas:

  · **Helados y productos → en pesos.** Cada documento despachado vale lo que se
    facturó, así que el % de entrega es plata entregada sobre plata despachada.
  · **Máquinas → en unidades.** El flete de una máquina se factura a $1 nominal
    (36 líneas FL sumaron $38 en todo agosto), así que valorizarlas no dice
    nada: se cuentan por máquina movida.

Sobre las notas de crédito: un rechazo se termina acreditando, así que el monto
rechazado y las NC del mes deberían parecerse. El informe muestra los dos para
que ese cuadre se pueda verificar mes a mes, en vez de asumirlo.
"""
import io
from datetime import date

import pandas as pd

from app.export_analisis import _escribir, _con_total, _FMT_NUM, _FMT_PCT, _FMT_CLP
from app.export_maquinas import _norm_doc, _desc, _pct

_FMT_FECHA = "dd/mm/yyyy"

# Orden fijo para que las columnas de estado no bailen entre meses.
_ESTADOS = ["Entregada", "Rechazada", "Pendiente"]


def preparar_entregas(ventas: pd.DataFrame, despachos: pd.DataFrame,
                      maquinas: pd.DataFrame | None = None) -> pd.DataFrame:
    """
    Una fila por documento despachado, con lo que se facturó y cómo terminó.

    Un documento puede tener varias filas de despacho (varios intentos): se
    resuelve al mejor resultado, igual que en el resto del sistema — si alguna
    vez quedó Entregada, la entrega ocurrió.
    """
    if despachos is None or despachos.empty:
        return pd.DataFrame()

    v = ventas.copy()
    v["_doc"] = _norm_doc(v["n_dcto"])
    v["_nc"] = v["tipo_dcto"].astype(str).str.upper().str.contains("CREDITO", na=False)
    v["_fl"] = (v["producto_codigo"].astype(str).str.upper()
                .str.startswith("FL-"))
    v["neto"] = pd.to_numeric(v["neto"], errors="coerce").fillna(0)

    # Monto de PRODUCTO por documento: los fletes se excluyen porque valen $1 y
    # ensuciarían el monto sin aportar nada.
    prod = (v[~v["_nc"] & ~v["_fl"]].groupby("_doc")["neto"].sum()
            if not v.empty else pd.Series(dtype=float))

    d = despachos.copy()
    d["_doc"] = _norm_doc(d["documento"])
    d["_est"] = d["estado"].astype(str).str.strip().str.capitalize()
    prio = {"Entregada": 0, "Rechazada": 1, "Pendiente": 2}
    d["_prio"] = d["_est"].map(prio).fillna(9)
    d = (d.sort_values(["_prio", "fecha_ruta"])
         .drop_duplicates("_doc"))

    d["Monto facturado"] = d["_doc"].map(prod).fillna(0.0)
    d["Transportista"] = (d["transportista"].fillna("(sin transportista)")
                          if "transportista" in d.columns
                          else "(sin transportista)")

    # Marca de máquina: el documento tiene un movimiento FL derivado de Obuma.
    if maquinas is not None and not maquinas.empty:
        mm = maquinas.copy()
        mm["_doc"] = _norm_doc(mm["documento"])
        tipo = mm.drop_duplicates("_doc").set_index("_doc")["tipo_mov"]
        d["tipo_mov"] = d["_doc"].map(tipo)
    else:
        d["tipo_mov"] = None
    d["Es máquina"] = d["tipo_mov"].notna()
    return d


def _tabla_por(d: pd.DataFrame, col: str, valor: str,
               etiqueta: str) -> pd.DataFrame:
    """Matriz grupo × estado con su total y su % de entrega."""
    if d.empty:
        return pd.DataFrame()
    if valor == "monto":
        piv = d.pivot_table(index=col, columns="_est", values="Monto facturado",
                            aggfunc="sum", fill_value=0.0)
    else:
        piv = d.pivot_table(index=col, columns="_est", values="_doc",
                            aggfunc="count", fill_value=0)
    piv = piv.reindex(columns=[e for e in _ESTADOS if e in piv.columns],
                      fill_value=0)
    piv["Total"] = piv.sum(axis=1)
    piv["% de entrega"] = piv.get("Entregada", 0) / piv["Total"].replace(0, pd.NA)
    piv = piv.reset_index().rename(columns={col: etiqueta})
    piv = piv.sort_values("Total", ascending=False)
    tot = _con_total(piv, etiqueta, ("% de entrega",))
    if not tot.empty:
        fin = tot.index[-1]
        tot.loc[fin, "% de entrega"] = _pct(tot.loc[fin].get("Entregada", 0),
                                            tot.loc[fin, "Total"])
    return tot


def vista_por_factura(ventas: pd.DataFrame, despachos: pd.DataFrame,
                      f_ini, f_fin) -> pd.DataFrame:
    """
    De lo FACTURADO en el período, en qué terminó cada peso.

    Es la otra mitad de la historia. La vista por ruta mide al transportista
    (solo lo que llevó ese mes); esta mide el mes comercial completo, e incluye
    lo que se facturó a fin de mes y salió a ruta al mes siguiente. Los dos
    números son correctos y responden preguntas distintas.

    `despachos` tiene que venir de una ventana MÁS ANCHA que el período, o lo
    despachado al mes siguiente se vería como "sin despacho".
    """
    v = ventas.copy()
    v["_doc"] = _norm_doc(v["n_dcto"])
    v["_nc"] = v["tipo_dcto"].astype(str).str.upper().str.contains("CREDITO", na=False)
    v["_fl"] = v["producto_codigo"].astype(str).str.upper().str.startswith("FL-")
    v["neto"] = pd.to_numeric(v["neto"], errors="coerce").fillna(0)

    d = despachos.copy()
    d["_doc"] = _norm_doc(d["documento"])
    d["_est"] = d["estado"].astype(str).str.strip().str.capitalize()
    d["_prio"] = d["_est"].map({"Entregada": 0, "Rechazada": 1, "Pendiente": 2}).fillna(9)
    d["_mes_ruta"] = pd.to_datetime(d["fecha_ruta"], errors="coerce").dt.to_period("M")
    dd = d.sort_values("_prio").drop_duplicates("_doc").set_index("_doc")

    g = (v[~v["_nc"] & ~v["_fl"]].groupby("_doc")["neto"].sum().to_frame("Monto"))
    g["Estado"] = g.index.map(dd["_est"]).fillna("Sin despacho registrado")
    mes_ruta = g.index.map(dd["_mes_ruta"])
    en_periodo = pd.Series(
        [False if pd.isna(m) else
         (m.start_time.date() <= f_fin and m.end_time.date() >= f_ini)
         for m in mes_ruta], index=g.index)
    g["Salió a ruta"] = [
        "Sin despacho" if pd.isna(m) else ("En el período" if p else str(m))
        for m, p in zip(mes_ruta, en_periodo)]
    return g


def _puente(ventas, despachos, f_ini, f_fin) -> pd.DataFrame:
    """El paso a paso entre la facturación del mes y lo que salió a ruta."""
    g = vista_por_factura(ventas, despachos, f_ini, f_fin)
    if g.empty:
        return pd.DataFrame()
    v = ventas.copy()
    v["_nc"] = v["tipo_dcto"].astype(str).str.upper().str.contains("CREDITO", na=False)
    v["neto"] = pd.to_numeric(v["neto"], errors="coerce").fillna(0)
    v["_fl"] = v["producto_codigo"].astype(str).str.upper().str.startswith("FL-")
    bruta = float(v[~v["_nc"]]["neto"].sum())
    fletes = float(v[~v["_nc"] & v["_fl"]]["neto"].sum())
    otro_mes = float(g.loc[~g["Salió a ruta"].isin(["En el período", "Sin despacho"]),
                           "Monto"].sum())
    sin_desp = float(g.loc[g["Salió a ruta"] == "Sin despacho", "Monto"].sum())
    en_ruta = float(g.loc[g["Salió a ruta"] == "En el período", "Monto"].sum())
    nc = abs(float(v[v["_nc"]]["neto"].sum()))
    ent_fact = float(g.loc[g["Estado"] == "Entregada", "Monto"].sum())
    prod = bruta - fletes
    filas = [
        ("Facturación bruta del período", bruta,
         "Todo lo facturado, notas de crédito aparte"),
        ("(−) Fletes de máquina", -fletes,
         "Se facturan a $1: distorsionarían el monto sin aportar"),
        ("(−) Facturado que salió a ruta en OTRO mes", -otro_mes,
         "Se facturó a fin de mes y el camión salió al mes siguiente"),
        ("(−) Facturado sin despacho registrado", -sin_desp,
         "Retiro en local, o el despacho no llegó al sistema"),
        ("(=) Lo que salió a ruta EN el período", en_ruta,
         "Este es el denominador del % por transportista"),
        ("", None, ""),
        ("Notas de crédito del período", -nc,
         "OJO: pueden acreditar facturas de meses anteriores"),
        ("Facturación neta (bruta − NC)", bruta - nc,
         "No es comparable con lo despachado: mezcla meses"),
        ("", None, ""),
        ("Entregado de lo facturado en el período", ent_fact,
         "Incluye lo que salió a ruta el mes siguiente"),
        ("% Entregado sobre la facturación bruta", _pct(ent_fact, prod),
         "La lectura comercial del mes"),
    ]
    return pd.DataFrame(filas, columns=["Concepto", "Monto", "Qué significa"])


def libro_entregas(ventas: pd.DataFrame, despachos: pd.DataFrame,
                   f_ini, f_fin, maquinas: pd.DataFrame | None = None,
                   clientes: pd.DataFrame | None = None,
                   hoy: date | None = None) -> bytes:
    """Devuelve el .xlsx del informe de entregas del período."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    hoy = hoy or date.today()

    d = preparar_entregas(ventas, despachos, maquinas)
    if d.empty:
        _escribir(wb, "Resumen", pd.DataFrame(),
                  nota="Sin despachos cargados en el período elegido.")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    prods = d[~d["Es máquina"]]
    maqs = d[d["Es máquina"]]

    v = ventas.copy()
    v["_nc"] = v["tipo_dcto"].astype(str).str.upper().str.contains("CREDITO", na=False)
    v["neto"] = pd.to_numeric(v["neto"], errors="coerce").fillna(0)
    nc_mes = float(v[v["_nc"]]["neto"].sum())
    fact_bruta = float(v[~v["_nc"]]["neto"].sum())

    ent = float(prods.loc[prods["_est"] == "Entregada", "Monto facturado"].sum())
    rech = float(prods.loc[prods["_est"] == "Rechazada", "Monto facturado"].sum())
    pend = float(prods.loc[prods["_est"] == "Pendiente", "Monto facturado"].sum())
    desp = ent + rech + pend

    m_ent = int((maqs["_est"] == "Entregada").sum())
    m_rech = int((maqs["_est"] == "Rechazada").sum())
    m_pend = int((maqs["_est"] == "Pendiente").sum())
    m_tot = len(maqs)

    # ── 1. Resumen ───────────────────────────────────────────────────────────
    ind = [
        ("Período", f"{f_ini:%d/%m/%Y} a {f_fin:%d/%m/%Y}"),
        ("Informe generado el", f"{hoy:%d/%m/%Y}"),
        ("", ""),
        ("HELADOS Y PRODUCTOS · en pesos", ""),
        ("Facturado que salió a ruta", desp),
        ("Entregado", ent),
        ("Rechazado", rech),
        ("Pendiente en ruta", pend),
        ("% DE ENTREGA", _pct(ent, desp)),
        ("", ""),
        ("MÁQUINAS · en unidades", ""),
        ("Máquinas despachadas", m_tot),
        ("Entregadas", m_ent),
        ("Rechazadas", m_rech),
        ("Pendientes en ruta", m_pend),
        ("% DE ENTREGA", _pct(m_ent, m_tot)),
        ("", ""),
        ("CUADRE CON LAS NOTAS DE CRÉDITO", ""),
        ("Monto rechazado en el período", rech),
        ("Notas de crédito del período", abs(nc_mes)),
        ("Diferencia", rech - abs(nc_mes)),
        ("", ""),
        ("CONTEXTO", ""),
        ("Facturación bruta del período", fact_bruta),
        ("De eso, lo que salió a ruta", desp),
        ("% de la facturación con despacho", _pct(desp, fact_bruta)),
    ]
    res = pd.DataFrame(ind, columns=["Indicador", "Valor"])
    ws = _escribir(wb, "Resumen", res,
                   nota=("Cuánto de lo que salió a ruta llegó. Los productos se "
                         "miden en pesos y las máquinas en unidades: el flete de "
                         "una máquina se factura a $1 nominal, así que "
                         "valorizarla no diría nada."))
    for fila in range(2, ws.max_row + 1):
        etq = str(ws.cell(row=fila, column=1).value or "")
        celda = ws.cell(row=fila, column=2)
        if etq.startswith("%"):
            celda.number_format = _FMT_PCT
        elif any(etq.startswith(t) for t in
                 ("Facturado", "Entregado", "Rechazado", "Pendiente en ruta",
                  "Monto", "Notas", "Diferencia", "Facturación", "De eso")):
            celda.number_format = _FMT_CLP

    # ── 2. Productos por transportista ───────────────────────────────────────
    t_prod = _tabla_por(prods, "Transportista", "monto", "Transportista")
    fmt = {c: _FMT_CLP for c in t_prod.columns if c not in
           ("Transportista", "% de entrega")}
    fmt["% de entrega"] = _FMT_PCT
    _escribir(wb, "Productos por transportista", t_prod, fmt,
              nota=("Monto facturado de los documentos que llevó cada "
                    "transportista, según cómo terminó el despacho. El % de "
                    "entrega es plata entregada sobre plata despachada."),
              total_ultima=not t_prod.empty)

    # También en número de documentos: un transportista puede verse bien en
    # pesos y mal en documentos si lo que le rebotó fueron entregas chicas.
    t_docs = _tabla_por(prods, "Transportista", "docs", "Transportista")
    _escribir(wb, "Productos · N° documentos", t_docs,
              {c: _FMT_NUM for c in t_docs.columns if c != "Transportista"}
              | {"% de entrega": _FMT_PCT},
              nota=("Lo mismo contado por documento, no por monto. Comparar las "
                    "dos hojas dice si los rechazos son muchos y chicos o pocos "
                    "y grandes."),
              total_ultima=not t_docs.empty)

    # ── 3. Máquinas por transportista ────────────────────────────────────────
    if not maqs.empty:
        t_maq = _tabla_por(maqs, "Transportista", "docs", "Transportista")
        _escribir(wb, "Máquinas por transportista", t_maq,
                  {c: _FMT_NUM for c in t_maq.columns if c != "Transportista"}
                  | {"% de entrega": _FMT_PCT},
                  nota=("Máquinas movidas por cada transportista, en unidades. "
                        "No se valorizan: el flete se factura a $1."),
                  total_ultima=not t_maq.empty)

        maqs_tipo = maqs.copy()
        maqs_tipo["Movimiento"] = maqs_tipo["tipo_mov"].map(
            {"nueva": "Instalación", "cambio": "Cambio", "retiro": "Retiro"}
        ).fillna("(otro)")
        t_tipo = _tabla_por(maqs_tipo, "Movimiento", "docs", "Movimiento")
        _escribir(wb, "Máquinas por movimiento", t_tipo,
                  {c: _FMT_NUM for c in t_tipo.columns if c != "Movimiento"}
                  | {"% de entrega": _FMT_PCT},
                  nota="El mismo corte por tipo de movimiento.",
                  total_ultima=not t_tipo.empty)

    # ── 3b. Cuadre con la facturación ────────────────────────────────────────
    # Sin esto la pregunta "por qué no me calza con la facturación del mes"
    # queda sin respuesta, y es la primera que aparece al mirar el informe.
    pu = _puente(ventas, despachos, f_ini, f_fin)
    if not pu.empty:
        ws2 = _escribir(wb, "Cuadre con la facturación", pu,
                        nota=("De la facturación del mes a lo que salió a ruta, "
                              "paso a paso. Las dos miradas son válidas y "
                              "responden cosas distintas: el % por transportista "
                              "se mide sobre lo que salió a ruta ESE mes; el % "
                              "comercial, sobre lo facturado en el mes."))
        for fila in range(4, ws2.max_row + 1):
            etq = str(ws2.cell(row=fila, column=1).value or "")
            ws2.cell(row=fila, column=2).number_format = (
                _FMT_PCT if etq.startswith("%") else _FMT_CLP)

        vf = vista_por_factura(ventas, despachos, f_ini, f_fin)
        det = (vf.groupby(["Salió a ruta", "Estado"])["Monto"]
               .agg(["sum", "count"]).reset_index()
               .rename(columns={"sum": "Monto", "count": "N° documentos"})
               .sort_values("Monto", ascending=False))
        _escribir(wb, "Facturado · en qué terminó", det,
                  {"Monto": _FMT_CLP, "N° documentos": _FMT_NUM},
                  nota=("Cada peso facturado en el período, según cuándo salió a "
                        "ruta y cómo terminó. Lo que dice «2026-09» se facturó "
                        "este mes y se despachó el siguiente."))

    # ── 4. Detalle de los rechazos ───────────────────────────────────────────
    rechz = d[d["_est"] == "Rechazada"].copy()
    if not rechz.empty:
        det = pd.DataFrame({
            "Fecha ruta": pd.to_datetime(rechz["fecha_ruta"],
                                         errors="coerce").dt.date,
            "Documento": rechz["_doc"],
            "Qué es": rechz["Es máquina"].map({True: "Máquina",
                                               False: "Producto"}),
            "Monto facturado": rechz["Monto facturado"],
            "Transportista": rechz["Transportista"],
            "RUT": rechz["cliente_rut"],
            "Cliente": _desc(rechz["cliente_rut"], clientes, "razon_social"),
            "Comuna": _desc(rechz["cliente_rut"], clientes, "comuna"),
            "Comentario del repartidor": (rechz["comentario_entrega"]
                                          if "comentario_entrega" in rechz.columns
                                          else ""),
        }).sort_values("Monto facturado", ascending=False)
        _escribir(wb, "Rechazos · detalle", det,
                  {"Fecha ruta": _FMT_FECHA, "Monto facturado": _FMT_CLP},
                  nota=("Cada despacho que volvió sin entregar, del más caro al "
                        "más barato. El comentario es lo que escribió el "
                        "repartidor en la ruta."))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
