"""
Export Excel de Clientes y Sucursales (para el área comercial).

Un cliente grande no es un solo punto de venta: tiene una casa matriz y locales
en distintas direcciones, cada uno con su ruta y su facturación. Este libro abre
esa caja: por cliente, por sucursal, por ruta y por comuna, con la evolución
mensual y el estado comercial de cada local.

Hojas:
  1. Resumen           — totales, cobertura y clientes multi-sucursal.
  2. Clientes          — una fila por RUT (con su nº de sucursales).
  3. Sucursales        — una fila por local: dirección, ruta, Fact-NC, estado.
  4. Sucursales x Mes  — matriz local × mes (para pivotear).
  5. Rutas             — Fact-NC y locales por código de ruta.
  6. Comunas           — lo mismo por comuna.
  7. Sin sucursal      — lo que el ERP no permite atribuir (se declara, no se reparte).
"""
import io

import pandas as pd

_HDR_BG = "1B3A6B"
_FMT_CLP = '"$"#,##0;[Red]-"$"#,##0'
_FMT_PCT = "0.0%"


# ── Perfil por sucursal ─────────────────────────────────────────────────────
def _estado(recency, primera_ym, current_ym):
    if recency is None:
        return "Sin compras"
    if primera_ym == current_ym:
        return "Nueva"
    if recency <= 1:
        return "Activa"
    if recency == 2:
        return "Riesgo"
    return "Perdida"


def _meses_entre(ym_a: str, ym_b: str) -> int:
    a, b = (int(ym_a[:4]), int(ym_a[5:7])), (int(ym_b[:4]), int(ym_b[5:7]))
    return (b[0] - a[0]) * 12 + (b[1] - a[1])


def perfil_sucursales(hist: pd.DataFrame, current_ym: str) -> pd.DataFrame:
    """Una fila por sucursal, con su historia resumida y su estado comercial."""
    if hist.empty:
        return pd.DataFrame()
    h = hist[hist["ym"] <= current_ym].copy()
    if h.empty:
        return pd.DataFrame()

    ym_ant = (pd.Period(current_ym) - 1).strftime("%Y-%m")
    ATTR = ["cliente_rut", "razon_social", "nombre", "direccion", "comuna",
            "ciudad", "ruta", "es_principal", "origen", "region_cliente"]
    attrs = {c: (c, "first") for c in ATTR if c in h.columns}

    g = (h.groupby("direccion_id")
         .agg(fact_nc_total=("fact_nc", "sum"),
              n_facturas=("n_facturas", "sum"),
              meses_activos=("ym", "nunique"),
              primera_compra=("ym", "min"),
              ultima_compra=("ym", "max"),
              **attrs)
         .reset_index())

    mes = h[h["ym"] == current_ym].groupby("direccion_id")["fact_nc"].sum()
    ant = h[h["ym"] == ym_ant].groupby("direccion_id")["fact_nc"].sum()
    g["fact_nc_mes"] = g["direccion_id"].map(mes).fillna(0)
    g["fact_nc_mes_ant"] = g["direccion_id"].map(ant).fillna(0)
    g["variacion"] = [
        (m - a) / abs(a) if a else None
        for m, a in zip(g["fact_nc_mes"], g["fact_nc_mes_ant"])]

    g["recency_meses"] = [_meses_entre(u, current_ym) for u in g["ultima_compra"]]
    g["estado"] = [_estado(r, p, current_ym) for r, p in
                   zip(g["recency_meses"], g["primera_compra"])]
    g["ticket_promedio"] = [
        (t / n) if n else 0 for t, n in zip(g["fact_nc_total"], g["n_facturas"])]

    # Participación dentro de su cliente y sobre el total.
    tot_cli = g.groupby("cliente_rut")["fact_nc_total"].transform("sum")
    g["part_en_cliente"] = g["fact_nc_total"] / tot_cli.replace(0, pd.NA)
    total = g["fact_nc_total"].sum() or 1
    g["part_del_total"] = g["fact_nc_total"] / total

    # ABC sobre el acumulado de todas las sucursales.
    g = g.sort_values("fact_nc_total", ascending=False).reset_index(drop=True)
    acum = g["fact_nc_total"].clip(lower=0).cumsum() / max(
        g["fact_nc_total"].clip(lower=0).sum(), 1)
    g["abc"] = ["A" if a <= 0.8 else ("B" if a <= 0.95 else "C") for a in acum]

    if "es_principal" in g.columns:
        g["tipo"] = g["es_principal"].map({True: "Casa matriz", False: "Sucursal"})
    else:
        g["tipo"] = "Sucursal"

    n_suc = g.groupby("cliente_rut")["direccion_id"].transform("nunique")
    g["sucursales_del_cliente"] = n_suc
    return g


def _nombre_sucursal(r) -> str:
    nom = r.get("nombre")
    nom = "" if nom is None or pd.isna(nom) else str(nom).strip()
    calle = r.get("direccion")
    calle = "" if calle is None or pd.isna(calle) else str(calle).strip()
    if not nom or nom.lower().startswith("dirección principal"):
        return "Casa matriz" if r.get("es_principal") else (calle or "Sin nombre")
    return nom


# ── Escritura del libro ─────────────────────────────────────────────────────
def _celda(v):
    """Valor apto para openpyxl: None si es nulo (NaN, NaT o pd.NA — este último no
    lo atrapa `isinstance(v, float)`), y tipo nativo si viene de numpy."""
    try:
        if v is None or (not isinstance(v, (list, tuple, dict)) and pd.isna(v)):
            return None
    except (TypeError, ValueError):
        pass
    if hasattr(v, "item"):          # numpy/pandas escalar → tipo nativo de Python
        try:
            return v.item()
        except (AttributeError, ValueError):
            pass
    return v


def _escribir(wb, titulo: str, df: pd.DataFrame, fmt: dict | None = None,
              nota: str | None = None):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(titulo[:31])
    fila0 = 1
    if nota:
        ws.cell(row=1, column=1, value=nota).font = Font(italic=True, color="64748B")
        fila0 = 3

    if df.empty:
        ws.cell(row=fila0, column=1, value="(sin datos)")
        return ws

    cols = [str(c) for c in df.columns]
    for j, name in enumerate(cols, 1):
        c = ws.cell(row=fila0, column=j, value=name)
        c.fill = PatternFill("solid", fgColor=_HDR_BG)
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, (_, row) in enumerate(df.iterrows(), fila0 + 1):
        for j, name in enumerate(df.columns, 1):
            v = _celda(row[name])
            celda = ws.cell(row=i, column=j, value=v)
            if fmt and fmt.get(name):
                celda.number_format = fmt[name]

    for j, name in enumerate(cols, 1):
        vals = [len(str(v)) for v in df.iloc[:, j - 1].head(200)]
        ancho = max([len(name)] + vals) + 2
        ws.column_dimensions[get_column_letter(j)].width = min(max(ancho, 10), 46)

    ws.freeze_panes = ws.cell(row=fila0 + 1, column=1)
    ws.auto_filter.ref = (f"A{fila0}:{get_column_letter(len(cols))}"
                          f"{fila0 + len(df)}")
    return ws


def libro_clientes_sucursales(hist_suc: pd.DataFrame, perfil_cli: pd.DataFrame,
                              current_ym: str) -> bytes:
    """
    Arma el Excel completo. `hist_suc` viene de data.get_sucursales_historia y
    `perfil_cli` del perfil que ya construye la página de Clientes.
    """
    from openpyxl import Workbook

    p = perfil_sucursales(hist_suc, current_ym)
    if not p.empty:
        p["sucursal"] = p.apply(_nombre_sucursal, axis=1)

    wb = Workbook()
    wb.remove(wb.active)

    # ── 2. Sucursales ──
    COLS_SUC = {
        "razon_social": "Cliente", "cliente_rut": "RUT", "sucursal": "Sucursal",
        "tipo": "Tipo", "direccion": "Dirección", "comuna": "Comuna",
        "ciudad": "Región", "ruta": "Ruta", "sucursales_del_cliente": "N° sucursales del cliente",
        "fact_nc_total": "Fact-NC acumulado", "part_en_cliente": "% del cliente",
        "part_del_total": "% del total", "abc": "Clase ABC",
        "fact_nc_mes": f"Fact-NC {current_ym}", "fact_nc_mes_ant": "Fact-NC mes anterior",
        "variacion": "Variación vs mes anterior", "n_facturas": "N° facturas",
        "ticket_promedio": "Ticket promedio", "meses_activos": "Meses con compra",
        "primera_compra": "Primera compra", "ultima_compra": "Última compra",
        "recency_meses": "Meses sin comprar", "estado": "Estado",
    }
    fmt_suc = {"Fact-NC acumulado": _FMT_CLP, f"Fact-NC {current_ym}": _FMT_CLP,
               "Fact-NC mes anterior": _FMT_CLP, "Ticket promedio": _FMT_CLP,
               "% del cliente": _FMT_PCT, "% del total": _FMT_PCT,
               "Variación vs mes anterior": _FMT_PCT}
    suc = (p[[c for c in COLS_SUC if c in p.columns]].rename(columns=COLS_SUC)
           if not p.empty else pd.DataFrame())

    # ── 3. Clientes (con su nº de sucursales) ──
    cli = perfil_cli.copy() if perfil_cli is not None else pd.DataFrame()
    if not cli.empty and not p.empty:
        nsuc = p.groupby("cliente_rut")["direccion_id"].nunique()
        cli["n_sucursales"] = cli["cliente_rut"].map(nsuc).fillna(0).astype(int)
    COLS_CLI = {
        "razon_social": "Cliente", "cliente_rut": "RUT", "comuna": "Comuna",
        "region": "Región", "n_sucursales": "N° sucursales",
        "ventas_total": "Fact-NC acumulado", "part": "% del total", "abc": "Clase ABC",
        "segmento": "Segmento", "ticket": "Ticket promedio",
        "frecuencia": "Meses con compra", "last_ym": "Última compra",
        "recency": "Meses sin comprar", "estado": "Estado",
    }
    fmt_cli = {"Fact-NC acumulado": _FMT_CLP, "Ticket promedio": _FMT_CLP,
               "% del total": _FMT_PCT}
    cli_out = (cli[[c for c in COLS_CLI if c in cli.columns]].rename(columns=COLS_CLI)
               .sort_values("Fact-NC acumulado", ascending=False)
               if not cli.empty else pd.DataFrame())

    # ── 4. Sucursales × Mes ──
    if not hist_suc.empty and not p.empty:
        base = hist_suc[hist_suc["ym"] <= current_ym].merge(
            p[["direccion_id", "sucursal"]], on="direccion_id", how="left")
        piv = (base.pivot_table(index=["razon_social", "cliente_rut", "sucursal"],
                                columns="ym", values="fact_nc", aggfunc="sum")
               .fillna(0).reset_index())
        piv.columns.name = None
        piv = piv.rename(columns={"razon_social": "Cliente", "cliente_rut": "RUT",
                                  "sucursal": "Sucursal"})
        fmt_piv = {c: _FMT_CLP for c in piv.columns if c not in
                   ("Cliente", "RUT", "Sucursal")}
    else:
        piv, fmt_piv = pd.DataFrame(), {}

    # ── 5. Rutas / 6. Comunas ──
    def _agrupado(col, etiqueta):
        if p.empty or col not in p.columns:
            return pd.DataFrame(), {}
        d = p.copy()
        d[col] = d[col].fillna("(sin dato)")
        g = (d.groupby(col)
             .agg(**{"N° clientes": ("cliente_rut", "nunique"),
                     "N° sucursales": ("direccion_id", "nunique"),
                     "Fact-NC acumulado": ("fact_nc_total", "sum"),
                     f"Fact-NC {current_ym}": ("fact_nc_mes", "sum"),
                     "N° facturas": ("n_facturas", "sum")})
             .reset_index().rename(columns={col: etiqueta}))
        tot = g["Fact-NC acumulado"].sum() or 1
        g["% del total"] = g["Fact-NC acumulado"] / tot
        g = g.sort_values("Fact-NC acumulado", ascending=False)
        return g, {"Fact-NC acumulado": _FMT_CLP, f"Fact-NC {current_ym}": _FMT_CLP,
                   "% del total": _FMT_PCT}

    rutas, fmt_rut = _agrupado("ruta", "Ruta")
    comunas, fmt_com = _agrupado("comuna", "Comuna")

    # ── 7. Sin sucursal ──
    sin = hist_suc.attrs.get("sin_sucursal", pd.DataFrame())
    if isinstance(sin, pd.DataFrame) and not sin.empty:
        sin = sin[sin["ym"] <= current_ym]
        sin = (sin.groupby("cliente_rut")["fact_nc"].sum()
               .reset_index(name="Fact-NC sin sucursal"))
        if not cli.empty:
            nom = dict(zip(cli["cliente_rut"], cli["razon_social"]))
            sin.insert(0, "Cliente", sin["cliente_rut"].map(nom))
        sin = (sin.rename(columns={"cliente_rut": "RUT"})
               .sort_values("Fact-NC sin sucursal", ascending=False))
    fmt_sin = {"Fact-NC sin sucursal": _FMT_CLP}

    # ── 1. Resumen ──
    multi = (p.groupby("cliente_rut")["direccion_id"].nunique() > 1).sum() if not p.empty else 0
    total_suc = float(p["fact_nc_total"].sum()) if not p.empty else 0.0
    total_sin = float(sin["Fact-NC sin sucursal"].sum()) if isinstance(sin, pd.DataFrame) and not sin.empty else 0.0
    resumen = pd.DataFrame({
        "Indicador": [
            "Período del informe", "Clientes", "Clientes con más de una sucursal",
            "Sucursales con facturación", "Fact-NC atribuido a sucursales",
            "Fact-NC sin sucursal", "% atribuido",
            "Sucursales activas (compraron el último mes)",
            "Sucursales en riesgo (2 meses sin comprar)",
            "Sucursales perdidas (3+ meses sin comprar)",
        ],
        "Valor": [
            current_ym,
            int(p["cliente_rut"].nunique()) if not p.empty else 0,
            int(multi),
            int(len(p)),
            total_suc,
            total_sin,
            (total_suc / (total_suc + total_sin)) if (total_suc + total_sin) else 0,
            int((p["estado"].isin(["Activa", "Nueva"])).sum()) if not p.empty else 0,
            int((p["estado"] == "Riesgo").sum()) if not p.empty else 0,
            int((p["estado"] == "Perdida").sum()) if not p.empty else 0,
        ],
    })

    NOTA = ("Fact-NC = facturas menos notas de crédito. La sucursal sale de la "
            "dirección de despacho del documento en el ERP. Lo que el ERP no "
            "permite atribuir (notas de crédito de anulación) va en la hoja "
            "'Sin sucursal': se declara, no se reparte.")

    _escribir(wb, "Resumen", resumen, nota=NOTA)
    _escribir(wb, "Sucursales", suc, fmt_suc)
    _escribir(wb, "Clientes", cli_out, fmt_cli)
    _escribir(wb, "Sucursales x Mes", piv, fmt_piv,
              nota="Fact-NC por sucursal y mes. Listo para tabla dinámica.")
    _escribir(wb, "Rutas", rutas, fmt_rut)
    _escribir(wb, "Comunas", comunas, fmt_com)
    _escribir(wb, "Sin sucursal", sin if isinstance(sin, pd.DataFrame) else pd.DataFrame(),
              fmt_sin, nota="Facturación que el ERP no permite asignar a un local.")

    # El formato $ del resumen se aplica solo a las filas de monto.
    ws = wb["Resumen"]
    for fila in (8, 9):        # 'Fact-NC atribuido' y 'Fact-NC sin sucursal'
        ws.cell(row=fila, column=2).number_format = _FMT_CLP
    ws.cell(row=10, column=2).number_format = _FMT_PCT   # '% atribuido'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
