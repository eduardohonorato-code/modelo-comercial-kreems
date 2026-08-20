"""
Informe Excel completo de la sección Análisis.

Un solo libro con todo lo que la página muestra en pantalla —y bastante más—
para el rango de fechas, sociedad y categorías que estén filtrados arriba:
cuánto se vendió por región, por mes, por SKU, por caja, por centro de
distribución, por categoría, por sucursal, por vendedor y por cliente.

Las hojas que terminan en "x Mes" son matrices listas para pegar en un gráfico;
la hoja "Detalle largo" es el mismo dato en formato plano, que es el que sirve
para armar tablas dinámicas propias.

Definiciones (idénticas a las de la app, ver CLAUDE.md §3):
  Fact-NC   = suma del neto; las notas de crédito ya entran con signo negativo.
  Cajas     = líneas cuya unidad de medida es CAJA, más Galletas NY (el ERP las
              tipifica como UN pero se venden por display de 12).
  Unidades  = cantidad facturada; las líneas de Servicios se neutralizan porque
              el ERP les manda una cantidad basura (el monto sí es real).
"""
import io

import pandas as pd

_HDR_BG = "1B3A6B"      # navy de marca
_TOT_BG = "FBEAF0"      # rosado suave para la fila TOTAL
_FMT_CLP = '"$"#,##0;[Red]-"$"#,##0'
_FMT_PCT = "0.0%"
_FMT_NUM = "#,##0"

MESES_C = {1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun",
           7: "Jul", 8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic"}


# ── Utilidades de escritura ──────────────────────────────────────────────────

def _celda(v):
    """Valor apto para openpyxl: None si es nulo, tipo nativo si viene de numpy."""
    try:
        if v is None or (not isinstance(v, (list, tuple, dict)) and pd.isna(v)):
            return None
    except (TypeError, ValueError):
        pass
    if hasattr(v, "item"):
        try:
            return v.item()
        except (AttributeError, ValueError):
            pass
    return v


def _escribir(wb, titulo: str, df: pd.DataFrame, fmt: dict | None = None,
              nota: str | None = None, total_ultima: bool = False):
    """Vuelca un DataFrame con el estilo de la casa: encabezado navy, panel
    congelado, autofiltro, ancho proporcional y (opcional) fila TOTAL resaltada."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(titulo[:31])
    fila0 = 1
    if nota:
        c = ws.cell(row=1, column=1, value=nota)
        c.font = Font(italic=True, color="64748B")
        fila0 = 3

    if df is None or df.empty:
        ws.cell(row=fila0, column=1, value="(sin datos para los filtros elegidos)")
        return ws

    cols = [str(c) for c in df.columns]
    for j, name in enumerate(cols, 1):
        c = ws.cell(row=fila0, column=j, value=name)
        c.fill = PatternFill("solid", fgColor=_HDR_BG)
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)

    for i, (_, row) in enumerate(df.iterrows(), fila0 + 1):
        for j, name in enumerate(df.columns, 1):
            celda = ws.cell(row=i, column=j, value=_celda(row[name]))
            if fmt and fmt.get(name):
                celda.number_format = fmt[name]

    if total_ultima:
        fila = fila0 + len(df)
        for j in range(1, len(cols) + 1):
            ws.cell(row=fila, column=j).font = Font(bold=True)
            ws.cell(row=fila, column=j).fill = PatternFill("solid", fgColor=_TOT_BG)

    for j, name in enumerate(cols, 1):
        vals = [len(str(v)) for v in df.iloc[:, j - 1].head(300)]
        ancho = max([len(name)] + vals) + 2
        ws.column_dimensions[get_column_letter(j)].width = min(max(ancho, 10), 46)

    ws.freeze_panes = ws.cell(row=fila0 + 1, column=1)
    n_filas = len(df) - (1 if total_ultima else 0)
    if n_filas > 0:
        ws.auto_filter.ref = (f"A{fila0}:{get_column_letter(len(cols))}"
                              f"{fila0 + n_filas}")
    return ws


# ── Métricas ─────────────────────────────────────────────────────────────────

def _prep(df: pd.DataFrame) -> pd.DataFrame:
    """Columnas auxiliares que usan todas las hojas."""
    d = df.copy()
    d["fecha"] = pd.to_datetime(d["fecha"], errors="coerce")
    d["_nc"] = (d["tipo_dcto"].astype(str).str.upper().str.contains("CREDITO")
                if "tipo_dcto" in d.columns else False)
    d["_ym"] = d["fecha"].dt.to_period("M")
    d["Mes"] = [_lbl_mes(p) for p in d["_ym"]]
    for c in ("cantidad", "neto"):
        d[c] = pd.to_numeric(d.get(c), errors="coerce").fillna(0)
    if "es_caja" not in d.columns:
        d["es_caja"] = False
    d["_cajas"] = d["cantidad"].where(d["es_caja"], 0)
    for c in ("region", "comuna", "categoria", "subcategoria", "sucursal", "cd",
              "nombre", "fabricante", "unidad_medida"):
        if c in d.columns:
            d[c] = d[c].fillna("(sin dato)").astype(str).str.strip()
            d.loc[d[c] == "", c] = "(sin dato)"
    return d


def _lbl_mes(p) -> str:
    """Period(M) → 'Jul-26' (ordenable con la lista de meses del rango)."""
    try:
        return f"{MESES_C.get(p.month, p.month)}-{str(p.year)[2:]}"
    except AttributeError:
        return "—"


def _metricas(d: pd.DataFrame, keys, total_general: float | None = None,
              etiquetas: dict | None = None) -> pd.DataFrame:
    """Bloque de métricas estándar agrupado por `keys`."""
    if d.empty:
        return pd.DataFrame()
    keys = [keys] if isinstance(keys, str) else list(keys)
    g = d.groupby(keys, dropna=False)
    out = g.agg(**{
        "Fact-NC": ("neto", "sum"),
        "Unidades": ("cantidad", "sum"),
        "Cajas": ("_cajas", "sum"),
        "N° clientes": ("cliente_rut", "nunique"),
        "N° SKUs": ("producto_codigo", "nunique"),
    })
    fac = d[~d["_nc"]]
    nc = d[d["_nc"]]
    out["Facturación bruta"] = fac.groupby(keys)["neto"].sum()
    out["Notas de crédito"] = nc.groupby(keys)["neto"].sum()
    out["N° facturas"] = fac.groupby(keys)["n_dcto"].nunique()
    out["N° NC"] = nc.groupby(keys)["n_dcto"].nunique()
    out = out.fillna(0).reset_index()

    out["Ticket promedio"] = [
        (v / n) if n else 0 for v, n in zip(out["Fact-NC"], out["N° facturas"])]
    tot = total_general if total_general is not None else out["Fact-NC"].sum()
    out["% del total"] = out["Fact-NC"] / tot if tot else 0

    out = out.sort_values("Fact-NC", ascending=False)
    orden = keys + ["Fact-NC", "% del total", "Facturación bruta",
                    "Notas de crédito", "Cajas", "Unidades", "N° facturas",
                    "N° NC", "N° clientes", "N° SKUs", "Ticket promedio"]
    # Contar únicos de lo que ya es la llave del grupo da siempre 1: sobra.
    sobran = set()
    if "cliente_rut" in keys:
        sobran.add("N° clientes")
    if "producto_codigo" in keys:
        sobran.add("N° SKUs")
    out = out[[c for c in orden if c in out.columns and c not in sobran]]
    # Cajas y unidades quedan tal cual (hay cantidades fraccionarias en las NC):
    # el formato #,##0 las muestra redondeadas y los totales siguen cuadrando.
    for c in ("N° facturas", "N° NC", "N° clientes", "N° SKUs"):
        if c in out.columns:
            out[c] = out[c].round().astype("int64")
    if etiquetas:
        out = out.rename(columns=etiquetas)
    return out


_FMT_METRICAS = {
    "Fact-NC": _FMT_CLP, "Facturación bruta": _FMT_CLP,
    "Notas de crédito": _FMT_CLP, "Ticket promedio": _FMT_CLP,
    "% del total": _FMT_PCT, "Cajas": _FMT_NUM, "Unidades": _FMT_NUM,
    "N° facturas": _FMT_NUM, "N° NC": _FMT_NUM, "N° clientes": _FMT_NUM,
    "N° SKUs": _FMT_NUM,
}


def _con_total(df: pd.DataFrame, etiqueta_col: str,
               no_sumar: tuple = ()) -> pd.DataFrame:
    """Agrega la fila TOTAL al pie (suma de las numéricas; el resto en blanco).

    Ojo con los conteos de únicos (clientes, SKUs, documentos): sumarlos por
    grupo duplicaría a quien aparece en más de uno, así que se dejan en blanco
    salvo que el llamador pase el valor real.
    """
    if df.empty:
        return df
    fila = {}
    for c in df.columns:
        if c in no_sumar or not pd.api.types.is_numeric_dtype(df[c]):
            fila[c] = ""
        else:
            fila[c] = df[c].sum()
    fila[etiqueta_col] = "TOTAL"
    if "% del total" in df.columns and "% del total" not in no_sumar:
        fila["% del total"] = df["% del total"].sum()
    if "Ticket promedio" in df.columns and "N° facturas" in df.columns:
        nf = df["N° facturas"].sum()
        fila["Ticket promedio"] = (df["Fact-NC"].sum() / nf) if nf else 0
    return pd.concat([df, pd.DataFrame([fila])], ignore_index=True)


def _meses_ordenados(d: pd.DataFrame) -> list:
    return [_lbl_mes(p) for p in sorted(p for p in d["_ym"].dropna().unique())]


def _pivot(d: pd.DataFrame, idx, val: str, cols: str = "Mes",
           orden_cols: list | None = None,
           etiquetas: dict | None = None) -> pd.DataFrame:
    """Matriz idx × cols con columna Total y fila TOTAL."""
    if d.empty:
        return pd.DataFrame()
    idx = [idx] if isinstance(idx, str) else list(idx)
    piv = d.pivot_table(index=idx, columns=cols, values=val,
                        aggfunc="sum", fill_value=0.0)
    if orden_cols:
        piv = piv.reindex(columns=[c for c in orden_cols if c in piv.columns],
                          fill_value=0.0)
    piv.columns = [str(c) for c in piv.columns]
    num_cols = list(piv.columns)
    piv["Total"] = piv[num_cols].sum(axis=1)
    piv = piv.sort_values("Total", ascending=False).reset_index()
    if etiquetas:
        piv = piv.rename(columns=etiquetas)
    return _con_total(piv, piv.columns[0])


def _fmt_pivot(df: pd.DataFrame, n_idx: int, money: bool) -> dict:
    if df is None or df.empty:
        return {}
    f = _FMT_CLP if money else _FMT_NUM
    return {c: f for c in df.columns[n_idx:]}


# ── Libro ────────────────────────────────────────────────────────────────────

def libro_analisis(df: pd.DataFrame, f_ini, f_fin, soc_lbl: str,
                   cats_sel: list | None = None,
                   df_prev: pd.DataFrame | None = None,
                   maquinas: pd.DataFrame | None = None,
                   vendedores: pd.DataFrame | None = None,
                   clientes: pd.DataFrame | None = None,
                   sociedades: dict | None = None) -> bytes:
    """
    Arma el informe completo. `df` son las líneas de venta ya enriquecidas por la
    página de Análisis (producto, región, sucursal, CD, marca de caja).
    """
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    if df is None or df.empty:
        _escribir(wb, "Resumen", pd.DataFrame(),
                  nota="Sin ventas en el rango de fechas elegido.")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    d = _prep(df)
    meses = _meses_ordenados(d)
    total = float(d["neto"].sum())
    cats_sel = cats_sel or []

    nom_vend = {}
    if vendedores is not None and not vendedores.empty:
        nom_vend = dict(zip(vendedores["id"], vendedores["nombre_canonico"]))
    if "vendedor_id" in d.columns:
        d["Vendedor"] = d["vendedor_id"].map(nom_vend).fillna("(sin vendedor)")
    if "sociedad_id" in d.columns:
        d["Sociedad"] = d["sociedad_id"].map(sociedades or {}).fillna("(sin dato)")

    _ren = {"region": "Región", "comuna": "Comuna", "categoria": "Categoría",
            "subcategoria": "Subcategoría", "sucursal": "Sucursal",
            "cd": "Centro distribución", "producto_codigo": "Código",
            "nombre": "Producto", "fabricante": "Fabricante",
            "unidad_medida": "Unidad", "cliente_rut": "RUT"}

    # ── 1. Resumen ───────────────────────────────────────────────────────────
    fac, nc = d[~d["_nc"]], d[d["_nc"]]
    bruto = float(fac["neto"].sum())
    n_fac = int(fac["n_dcto"].nunique())
    prev_total = (float(pd.to_numeric(df_prev["neto"], errors="coerce").sum())
                  if df_prev is not None and not df_prev.empty else None)

    def _top(col):
        if col not in d.columns:
            return "—"
        s = d.groupby(col)["neto"].sum().sort_values(ascending=False)
        if not len(s):
            return "—"
        return f"{s.index[0]} ({s.iloc[0]:,.0f})".replace(",", ".")

    ind = [
        ("Período del informe", f"{f_ini:%d/%m/%Y} → {f_fin:%d/%m/%Y}"),
        ("Días del rango", (f_fin - f_ini).days + 1),
        ("Sociedad", soc_lbl),
        ("Categorías filtradas", ", ".join(cats_sel) if cats_sel else "Todas"),
        ("Meses incluidos", len(meses)),
        ("", ""),
        ("Fact-NC (facturas − notas de crédito)", total),
        ("Facturación bruta", bruto),
        ("Notas de crédito", float(nc["neto"].sum())),
        ("% de NC sobre la facturación bruta",
         (abs(float(nc["neto"].sum())) / bruto) if bruto else 0),
        ("N° de facturas", n_fac),
        ("N° de notas de crédito", int(nc["n_dcto"].nunique())),
        ("Ticket promedio por factura", (total / n_fac) if n_fac else 0),
        ("", ""),
        ("Cajas vendidas", float(d["_cajas"].sum())),
        ("Unidades vendidas", float(d["cantidad"].sum())),
        ("SKUs distintos vendidos", int(d["producto_codigo"].nunique())),
        ("Clientes distintos facturados", int(d["cliente_rut"].nunique())),
        ("Regiones con venta",
         int(d["region"].nunique()) if "region" in d.columns else 0),
        ("Comunas con venta",
         int(d["comuna"].nunique()) if "comuna" in d.columns else 0),
        ("Sucursales / centros de distribución",
         int(d["sucursal"].nunique()) if "sucursal" in d.columns else 0),
        ("", ""),
        ("Mejor región", _top("region")),
        ("Mejor categoría", _top("categoria")),
        ("Mejor SKU", _top("nombre")),
        ("Mejor centro de distribución", _top("cd")),
    ]
    if prev_total is not None:
        var = (total - prev_total) / abs(prev_total) if prev_total else None
        ind += [
            ("", ""),
            ("Fact-NC del período anterior (mismo nº de días)", prev_total),
            ("Variación vs período anterior", total - prev_total),
            ("Variación % vs período anterior", var if var is not None else "—"),
        ]
    resumen = pd.DataFrame(ind, columns=["Indicador", "Valor"])

    NOTA = ("Fact-NC = facturas menos notas de crédito (las NC ya vienen con signo "
            "negativo). Cajas = líneas con unidad de medida CAJA más Galletas NY "
            "(display de 12). Las cantidades de la categoría Servicios se "
            "neutralizan: el ERP las manda corruptas, el monto sí es real. "
            "Todas las hojas respetan el filtro de fecha, sociedad y categoría.")
    ws = _escribir(wb, "Resumen", resumen, nota=NOTA)
    for i, (etiq, val) in enumerate(ind, 4):   # nota en 1, encabezado en 3
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            es_pct = "%" in etiq
            es_plata = any(k in etiq for k in
                           ("Fact-NC", "Facturación", "crédito", "Ticket",
                            "Variación vs"))
            ws.cell(row=i, column=2).number_format = (
                _FMT_PCT if es_pct else (_FMT_CLP if es_plata else _FMT_NUM))

    # ── 2. Evolución mensual ─────────────────────────────────────────────────
    mens = _metricas(d, "Mes", total)
    if not mens.empty:
        mens["Mes"] = pd.Categorical(mens["Mes"], categories=meses, ordered=True)
        mens = mens.sort_values("Mes")
        mens["Var. % vs mes anterior"] = mens["Fact-NC"].pct_change()
        mens["Mes"] = mens["Mes"].astype(str)
        mens = _con_total(mens, "Mes",
                          no_sumar=("Var. % vs mes anterior", "N° clientes",
                                    "N° SKUs"))
    _escribir(wb, "Mensual", mens,
              {**_FMT_METRICAS, "Var. % vs mes anterior": _FMT_PCT},
              nota="Una fila por mes del rango. Var. % compara con el mes "
                   "anterior de esta misma tabla.",
              total_ultima=not mens.empty)

    _NO_SUM = ("N° clientes", "N° SKUs")

    # ── 3-5. Geografía ───────────────────────────────────────────────────────
    reg = _con_total(_metricas(d, "region", total, _ren), "Región", _NO_SUM)
    _escribir(wb, "Región", reg, _FMT_METRICAS,
              nota="Venta por región del cliente (dim_cliente).",
              total_ultima=not reg.empty)

    p_reg = _pivot(d, "region", "neto", orden_cols=meses, etiquetas=_ren)
    _escribir(wb, "Región x Mes", p_reg, _fmt_pivot(p_reg, 1, True),
              nota="Fact-NC por región y mes. Filas = región, columnas = mes.",
              total_ultima=not p_reg.empty)

    p_regc = _pivot(d, "region", "_cajas", orden_cols=meses, etiquetas=_ren)
    _escribir(wb, "Región x Mes (cajas)", p_regc, _fmt_pivot(p_regc, 1, False),
              nota="Cajas por región y mes.", total_ultima=not p_regc.empty)

    com = _con_total(_metricas(d, ["region", "comuna"], total, _ren),
                     "Comuna", _NO_SUM)
    _escribir(wb, "Comuna", com, _FMT_METRICAS,
              nota="Venta por comuna, con su región.", total_ultima=not com.empty)

    # ── 6-11. Producto ───────────────────────────────────────────────────────
    llaves_sku = [c for c in ("producto_codigo", "nombre", "categoria",
                              "subcategoria", "fabricante", "unidad_medida")
                  if c in d.columns]
    sku = _metricas(d, llaves_sku, total, _ren)
    if not sku.empty:
        acum = (sku["Fact-NC"].clip(lower=0).cumsum()
                / max(sku["Fact-NC"].clip(lower=0).sum(), 1))
        sku["% acumulado"] = acum
        sku["Clase ABC"] = ["A" if a <= 0.8 else ("B" if a <= 0.95 else "C")
                            for a in acum]
        # En blanco para lo que no se vende en caja: un "$0" se lee como precio.
        sku["Precio promedio por caja"] = [
            (v / c) if c else None for v, c in zip(sku["Fact-NC"], sku["Cajas"])]
        sku = _con_total(sku, "Código",
                         no_sumar=("% acumulado", "Precio promedio por caja",
                                   "N° clientes"))
    _escribir(wb, "SKU", sku,
              {**_FMT_METRICAS, "% acumulado": _FMT_PCT,
               "Precio promedio por caja": _FMT_CLP},
              nota="Una fila por SKU. Clase ABC: A = 80% de la venta, B = hasta "
                   "el 95%, C = la cola. El precio por caja solo tiene sentido "
                   "en los SKUs que se venden en caja.",
              total_ultima=not sku.empty)

    p_sku = _pivot(d, ["producto_codigo", "nombre"], "neto", orden_cols=meses,
                   etiquetas=_ren)
    _escribir(wb, "SKU x Mes", p_sku, _fmt_pivot(p_sku, 2, True),
              nota="Fact-NC por SKU y mes.", total_ultima=not p_sku.empty)

    p_skuc = _pivot(d, ["producto_codigo", "nombre"], "_cajas",
                    orden_cols=meses, etiquetas=_ren)
    _escribir(wb, "SKU x Mes (cajas)", p_skuc, _fmt_pivot(p_skuc, 2, False),
              nota="Cajas por SKU y mes.", total_ultima=not p_skuc.empty)

    if "region" in d.columns:
        p_sr = _pivot(d, ["producto_codigo", "nombre"], "neto", cols="region",
                      etiquetas=_ren)
        _escribir(wb, "SKU x Región", p_sr, _fmt_pivot(p_sr, 2, True),
                  nota="Fact-NC por SKU y región del cliente.",
                  total_ultima=not p_sr.empty)

    if "cd" in d.columns:
        caja = d[d["es_caja"]]
        p_cd = _pivot(caja, ["producto_codigo", "nombre", "categoria"], "_cajas",
                      cols="cd", etiquetas=_ren)
        if not p_cd.empty:
            monto = caja.groupby("producto_codigo")["neto"].sum()
            p_cd["Monto $"] = p_cd["Código"].map(monto).fillna(0)
            p_cd.loc[p_cd.index[-1], "Monto $"] = float(caja["neto"].sum())
        _escribir(wb, "SKU x CD (cajas)", p_cd,
                  {**_fmt_pivot(p_cd, 3, False), "Monto $": _FMT_CLP},
                  nota="Cajas por SKU y centro de distribución (solo las líneas "
                       "que cuentan como caja). Monto $ = Fact-NC de esas líneas.",
                  total_ultima=not p_cd.empty)

        p_cdm = _pivot(d, "cd", "neto", orden_cols=meses, etiquetas=_ren)
        _escribir(wb, "CD x Mes", p_cdm, _fmt_pivot(p_cdm, 1, True),
                  nota="Fact-NC por centro de distribución y mes.",
                  total_ultima=not p_cdm.empty)

        p_cdc = _pivot(d, "cd", "_cajas", orden_cols=meses, etiquetas=_ren)
        _escribir(wb, "CD x Mes (cajas)", p_cdc, _fmt_pivot(p_cdc, 1, False),
                  nota="Cajas por centro de distribución y mes.",
                  total_ultima=not p_cdc.empty)

    # ── 12-15. Categoría / sucursal ──────────────────────────────────────────
    cat = _con_total(_metricas(d, "categoria", total, _ren), "Categoría", _NO_SUM)
    _escribir(wb, "Categoría", cat, _FMT_METRICAS,
              nota="Venta por categoría de producto (las categorías duplicadas "
                   "del catálogo ya vienen unificadas).",
              total_ultima=not cat.empty)

    p_cat = _pivot(d, "categoria", "neto", orden_cols=meses, etiquetas=_ren)
    _escribir(wb, "Categoría x Mes", p_cat, _fmt_pivot(p_cat, 1, True),
              nota="Fact-NC por categoría y mes.", total_ultima=not p_cat.empty)

    if "subcategoria" in d.columns:
        sub = _con_total(_metricas(d, ["categoria", "subcategoria"], total, _ren),
                         "Subcategoría", _NO_SUM)
        _escribir(wb, "Subcategoría", sub, _FMT_METRICAS,
                  total_ultima=not sub.empty)

    suc = _con_total(_metricas(d, "sucursal", total, _ren), "Sucursal", _NO_SUM)
    _escribir(wb, "Sucursal", suc, _FMT_METRICAS,
              nota="Sucursal emisora del documento en el ERP.",
              total_ultima=not suc.empty)

    # ── 16-17. Vendedor ──────────────────────────────────────────────────────
    if "Vendedor" in d.columns:
        vend = _con_total(_metricas(d, "Vendedor", total), "Vendedor", _NO_SUM)
        _escribir(wb, "Vendedor", vend, _FMT_METRICAS,
                  nota="Atribución por documento, igual que el panel de gerencia.",
                  total_ultima=not vend.empty)

        p_v = _pivot(d, "Vendedor", "neto", orden_cols=meses)
        _escribir(wb, "Vendedor x Mes", p_v, _fmt_pivot(p_v, 1, True),
                  nota="Fact-NC por vendedor y mes.", total_ultima=not p_v.empty)

    # ── 18. Clientes ─────────────────────────────────────────────────────────
    cli = _metricas(d, "cliente_rut", total, _ren)
    if not cli.empty:
        # Región y comuna salen de las propias ventas (ya normalizadas), no del
        # catálogo: si no, esta hoja mostraría "Biobio" y las demás "Biobío".
        for col, etiqueta in (("region", "Región"), ("comuna", "Comuna")):
            if col in d.columns:
                cli[etiqueta] = cli["RUT"].map(d.groupby("cliente_rut")[col].first())
        if clientes is not None and not clientes.empty and "rut" in clientes.columns:
            info = clientes.drop_duplicates("rut").set_index("rut")
            for col, etiqueta in (("razon_social", "Cliente"), ("tipo", "Tipo")):
                if col in info.columns:
                    cli[etiqueta] = cli["RUT"].map(info[col])
        orden = [c for c in ("Cliente", "RUT", "Tipo", "Región", "Comuna")
                 if c in cli.columns]
        cli = cli[orden + [c for c in cli.columns if c not in orden]]
        # dim_cliente no siempre trae 'tipo': no ensuciar con una columna vacía.
        vacias = [c for c in ("Tipo", "Región", "Comuna")
                  if c in cli.columns and cli[c].isna().all()]
        cli = cli.drop(columns=vacias)
        acum = (cli["Fact-NC"].clip(lower=0).cumsum()
                / max(cli["Fact-NC"].clip(lower=0).sum(), 1))
        cli["Clase ABC"] = ["A" if a <= 0.8 else ("B" if a <= 0.95 else "C")
                            for a in acum]
        cli = _con_total(cli, cli.columns[0], _NO_SUM)
    _escribir(wb, "Clientes", cli, _FMT_METRICAS,
              nota="Una fila por cliente facturado en el período.",
              total_ultima=not cli.empty)

    # ── 19-20. Máquinas ──────────────────────────────────────────────────────
    if maquinas is not None and not maquinas.empty:
        m = maquinas.copy()
        m["fecha"] = pd.to_datetime(m["fecha"], errors="coerce")
        m["Mes"] = [_lbl_mes(f.to_period("M")) if pd.notna(f) else "—"
                    for f in m["fecha"]]
        m["Movimiento"] = m["tipo_mov"].map(
            {"nueva": "Nueva (FL-4)", "cambio": "Cambio (FL-1/3/5)",
             "retiro": "Retiro (FL-2)"}).fillna(m["tipo_mov"])
        m["Estado"] = m["estado"].map(
            {"entregada": "Entregada", "gestionada": "Pendiente",
             "rechazada": "Rechazada"}).fillna(m["estado"])
        m["Vendedor"] = m["vendedor_id"].map(nom_vend).fillna("(sin vendedor)")

        maq = (m.groupby(["Mes", "Movimiento", "Estado"])
               .agg(**{"N° máquinas": ("documento", "count"),
                       "N° clientes": ("cliente_rut", "nunique")})
               .reset_index())
        maq["Mes"] = pd.Categorical(maq["Mes"], categories=meses + ["—"],
                                    ordered=True)
        maq = maq.sort_values(["Mes", "Movimiento", "Estado"])
        maq["Mes"] = maq["Mes"].astype(str)
        maq = _con_total(maq, "Mes", _NO_SUM)
        _escribir(wb, "Máquinas", maq,
                  {"N° máquinas": _FMT_NUM, "N° clientes": _FMT_NUM},
                  nota="Movimientos de máquinas en comodato. El tipo sale de "
                       "Obuma (código FL) y el estado de entrega, de los "
                       "despachos de Autoventa. El filtro de categoría no aplica.",
                  total_ultima=not maq.empty)

        mv = (m.pivot_table(index="Vendedor", columns="Movimiento",
                            values="documento", aggfunc="count", fill_value=0)
              .reset_index())
        if not mv.empty:
            mv.columns = [str(c) for c in mv.columns]
            cols_mov = [c for c in mv.columns if c != "Vendedor"]
            mv["Total"] = mv[cols_mov].sum(axis=1)
            mv = _con_total(mv.sort_values("Total", ascending=False), "Vendedor")
        _escribir(wb, "Máquinas por vendedor", mv,
                  {c: _FMT_NUM for c in mv.columns if c != "Vendedor"},
                  nota="Cuenta de movimientos por vendedor y tipo.",
                  total_ultima=not mv.empty)

    # ── 21. Detalle largo ────────────────────────────────────────────────────
    llaves = [c for c in ("Mes", "Sociedad", "region", "comuna", "sucursal", "cd",
                          "Vendedor", "categoria", "subcategoria",
                          "producto_codigo", "nombre", "unidad_medida")
              if c in d.columns]
    largo = (d.groupby(llaves, dropna=False)
             .agg(**{"Fact-NC": ("neto", "sum"),
                     "Cajas": ("_cajas", "sum"),
                     "Unidades": ("cantidad", "sum"),
                     "N° documentos": ("n_dcto", "nunique"),
                     "N° clientes": ("cliente_rut", "nunique")})
             .reset_index().rename(columns=_ren))
    if "Mes" in largo.columns:
        largo["Mes"] = pd.Categorical(largo["Mes"], categories=meses, ordered=True)
        largo = largo.sort_values(["Mes", "Fact-NC"], ascending=[True, False])
        largo["Mes"] = largo["Mes"].astype(str)
    else:
        largo = largo.sort_values("Fact-NC", ascending=False)
    _escribir(wb, "Detalle largo", largo,
              {"Fact-NC": _FMT_CLP, "Cajas": _FMT_NUM, "Unidades": _FMT_NUM,
               "N° documentos": _FMT_NUM, "N° clientes": _FMT_NUM},
              nota="Formato plano (una fila por combinación) para armar tablas "
                   "dinámicas propias. Es la misma data de las demás hojas.")

    # ── 22. Definiciones ─────────────────────────────────────────────────────
    defs = pd.DataFrame([
        ("Fact-NC", "Suma del neto: facturas menos notas de crédito. Las NC ya "
                    "vienen con signo negativo desde el ERP."),
        ("Facturación bruta", "Solo documentos de tipo factura, sin descontar NC."),
        ("Notas de crédito", "Solo documentos de tipo nota de crédito (negativo)."),
        ("N° facturas", "Facturas distintas, contadas por número de documento."),
        ("Ticket promedio", "Fact-NC dividido por el N° de facturas."),
        ("Cajas", "Líneas cuya unidad de medida es CAJA, más Galletas NY, que el "
                  "ERP tipifica como UN pero se vende por display de 12."),
        ("Unidades", "Cantidad facturada. Las líneas de Servicios quedan en 0: el "
                     "ERP manda una cantidad corrupta (el monto sí es real)."),
        ("Región / Comuna", "Del cliente, según dim_cliente. Los nombres se "
                            "unifican para que una misma región no aparezca dos "
                            "veces por acentos mal codificados."),
        ("Sucursal / Centro distribución",
         "Sucursal emisora del documento en Obuma. C. Matriz (Acuña) se cuenta "
         "como Concepción."),
        ("Clase ABC", "A = SKUs o clientes que acumulan el 80% de la venta, "
                      "B = hasta el 95%, C = la cola."),
        ("Máquinas", "FL-4 = instalación en cliente nuevo, FL-1/3/5 = cambio, "
                     "FL-2 = retiro. El estado de entrega viene de despachos."),
        ("Fila TOTAL", "Los conteos de únicos (clientes, SKUs) se dejan en blanco "
                       "en el total: sumarlos por grupo duplicaría a quien "
                       "aparece en más de uno. El total real está en la hoja "
                       "Resumen."),
    ], columns=["Concepto", "Definición"])
    _escribir(wb, "Definiciones", defs,
              nota="Mismas definiciones que usa la app en pantalla.")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
