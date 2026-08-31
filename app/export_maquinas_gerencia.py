"""
Informe de máquinas para gerencia: cinco hojas y el flujo a la vista.

El informe completo (19 hojas) sigue existiendo para quien tiene que ir al
detalle. Este es el otro extremo: lo que se abre en una reunión y se entiende
sin explicación previa, contando siempre el mismo recorrido —

    pedido ingresado  →  sin gestionar  →  DTE emitido  →  despacho  →  entrega

Las hojas son:
  1. Tablero            · cada indicador contra su meta, con semáforo
  2. Flujo mensual      · el recorrido completo, mes a mes
  3. Sin gestionar      · pedidos ingresados que siguen sin DTE, por antigüedad
  4. Despachos y estado · qué pasó con cada máquina que salió a ruta
  5. Rechazos           · por qué volvieron, y de quién es cada motivo
"""
import io
from datetime import date

import pandas as pd

from app.export_analisis import _escribir, _con_total, _FMT_NUM, _FMT_PCT
from app.export_maquinas import (ENTREGADA, RECHAZADA, EN_RUTA, SIN_DESPACHO,
                                 SIN_INFO, MOV_LBL, _detalle, _desc, _mes,
                                 _FMT_FECHA, _FMT_DEC, _pct)
from app.kpis_maquinas import calcular_kpis, calcular_flujo, semanal

_SEMAFORO = {True: "✔ en meta", False: "✘ fuera de meta", None: "— sin meta"}


def libro_gerencia(mov: pd.DataFrame, ped: pd.DataFrame, f_ini, f_fin,
                   metas: dict, soc_lbl: str = "Ambas",
                   clientes: pd.DataFrame | None = None,
                   hoy: date | None = None) -> bytes:
    """Devuelve el .xlsx corto para gerencia."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    hoy = hoy or date.today()

    if mov is None or mov.empty:
        _escribir(wb, "Tablero", pd.DataFrame(),
                  nota="Sin movimientos de máquinas en el período elegido.")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    kpis = calcular_kpis(mov, ped, f_ini, f_fin, metas, hoy)
    flujo = calcular_flujo(mov, ped, f_ini, f_fin)

    # ── 1. Tablero ───────────────────────────────────────────────────────────
    tab = pd.DataFrame([{
        "Etapa": k["grupo"],
        "Indicador": k["nombre"],
        "Resultado": k["valor_txt"],
        "Meta": k["meta_txt"],
        "Estado": _SEMAFORO[k["cumple"]],
        "Responsable": k["responsable"],
        "Detalle": k["detalle"],
    } for k in kpis])
    ws = _escribir(wb, "Tablero", tab,
                   nota=(f"Período {f_ini:%d/%m/%Y} a {f_fin:%d/%m/%Y} · "
                         f"Sociedad: {soc_lbl} · Generado el {hoy:%d/%m/%Y}. "
                         "Las metas se editan en la app, sección Control de "
                         "Máquinas."))
    _pintar_estado(ws, tab, "Estado")

    # Debajo del tablero, el recorrido en cuatro líneas: es lo que explica los
    # números de arriba y evita tener que saltar a otra hoja para entenderlos.
    fl = pd.DataFrame([{
        "Etapa del recorrido": f["etapa"],
        "Cuántas": f["valor"],
        "Responsable": f["responsable"],
        "Qué mirar": f["detalle"],
    } for f in flujo])
    _escribir(wb, "Flujo del período", fl, {"Cuántas": _FMT_NUM},
              nota=("El mismo movimiento contado en cada etapa: entra como "
                    "pedido, se emite el documento, sale a ruta y se entrega. "
                    "La caída entre dos filas es dónde se está perdiendo."))

    # ── 2. Flujo mensual ─────────────────────────────────────────────────────
    m = mov.copy()
    m["_ym"] = m["fecha"].dt.to_period("M")
    ing_mes = (ped["_ingreso"].dt.to_period("M").value_counts().to_dict()
               if not ped.empty else {})
    sindte_mes = (ped[ped["_sin_dte"] & ~ped["_fantasma"]]["_ingreso"]
                  .dt.to_period("M").value_counts().to_dict()
                  if not ped.empty else {})
    filas = []
    for per in pd.period_range(pd.Timestamp(f_ini), pd.Timestamp(f_fin), freq="M"):
        g = m[m["_ym"] == per]
        con_info = int((~g["_sin_info"]).sum())
        nuevas = int((g["tipo_mov"] == "nueva").sum())
        retiros = int((g["tipo_mov"] == "retiro").sum())
        filas.append({
            "Mes": _mes(per.to_timestamp()),
            "Pedidos ingresados": int(ing_mes.get(per, 0)),
            "Sin gestionar (sin DTE)": int(sindte_mes.get(per, 0)),
            "Gestiones con DTE": len(g),
            "Instalaciones": nuevas,
            "Cambios": int((g["tipo_mov"] == "cambio").sum()),
            "Retiros": retiros,
            "Parque neto": nuevas - retiros,
            "Con despacho": con_info,
            "Entregadas": int(g["_entregada"].sum()),
            "Rechazadas": int(g["_rechazada"].sum()),
            "Sin confirmar": int(g["_pendiente"].sum()),
            "% Entregado": _pct(int(g["_entregada"].sum()), con_info) if con_info else None,
        })
    mes_df = _con_total(pd.DataFrame(filas), "Mes", ("% Entregado",))
    if not mes_df.empty:
        fin_i = mes_df.index[-1]
        tot_ent = mes_df.loc[fin_i, "Entregadas"]
        tot_ci = mes_df.loc[fin_i, "Con despacho"]
        mes_df.loc[fin_i, "% Entregado"] = _pct(tot_ent, tot_ci) if tot_ci else None
    fmt_mes = {c: _FMT_NUM for c in mes_df.columns if c != "Mes"}
    fmt_mes["% Entregado"] = _FMT_PCT
    _escribir(wb, "Flujo mensual", mes_df, fmt_mes,
              nota=("Cada mes de punta a punta. 'Sin gestionar' son pedidos "
                    "ingresados ESE mes que hoy siguen sin documento, así que "
                    "los meses viejos con número alto son deuda vieja, no "
                    "actividad reciente."),
              total_ultima=not mes_df.empty)

    # ── 3. Semana a semana contra la meta ────────────────────────────────────
    sem = semanal(mov, ped, f_ini, f_fin, metas.get("meta_gestiones_semana"))
    _escribir(wb, "Semana a semana", sem,
              {"Días en el rango": _FMT_NUM, "Pedidos ingresados": _FMT_NUM,
               "Gestiones con DTE": _FMT_NUM, "Meta": _FMT_NUM,
               "% Meta": _FMT_PCT, "Entregadas": _FMT_NUM,
               "Rechazadas": _FMT_NUM},
              nota=("Semanas de lunes a domingo. Las que quedan cortadas por el "
                    "borde del período salen sin meta: no se les puede exigir "
                    "una semana entera."))

    # ── 4. Sin gestionar ─────────────────────────────────────────────────────
    if not ped.empty:
        cola = ped[ped["_sin_dte"] & ~ped["_fantasma"]].copy()
        if not cola.empty:
            tabla = pd.DataFrame({
                "Fecha del pedido": cola["_ingreso"].dt.date,
                "Días esperando": (pd.Timestamp(hoy) - cola["_ingreso"]).dt.days,
                "N° pedido": cola["n_pedido"],
                "Movimiento": cola["_mov"].map(MOV_LBL).fillna("(otro)"),
                "Vendedor que lo ingresó": cola["vendedor_id"].map(
                    _mapa_vendedor(mov)).fillna("Sin asignar"),
                "RUT": cola["cliente_rut"],
                "Cliente": _desc(cola["cliente_rut"], clientes, "razon_social"),
                "Comuna": _desc(cola["cliente_rut"], clientes, "comuna"),
            }).sort_values("Días esperando", ascending=False)
            _escribir(wb, "Sin gestionar", tabla,
                      {"Fecha del pedido": _FMT_FECHA, "Días esperando": _FMT_NUM},
                      nota=("Pedidos que el vendedor ya ingresó y que siguen sin "
                            "DTE. El vendedor hizo su parte: esto es cola de "
                            "emisión. Salen todos los abiertos, sin importar el "
                            "período, porque el de hace tres meses es el que "
                            "más urge."))

    # ── 5. Despachos y estado ────────────────────────────────────────────────
    est_orden = [ENTREGADA, RECHAZADA, EN_RUTA, SIN_DESPACHO, SIN_INFO]
    ct = pd.crosstab(mov["Movimiento"], mov["Estado entrega"])
    ct = ct.reindex(columns=[e for e in est_orden if e in ct.columns], fill_value=0)
    ct["Total"] = ct.sum(axis=1)
    ct = _con_total(ct.reset_index(), "Movimiento")
    _escribir(wb, "Despachos y estado", ct,
              {c: _FMT_NUM for c in ct.columns if c != "Movimiento"},
              nota=("Qué pasó con cada tipo de movimiento. 'Sin despacho' es que "
                    "ese mes SÍ tiene despachos cargados y este documento no "
                    "aparece: hay que ir a buscarlo. 'Sin información' es que no "
                    "hay despachos de ese mes y sociedad — Acuña nunca los "
                    "tiene, no pasa por Autoventa."),
              total_ultima=not ct.empty)

    # ── 6. Rechazos ──────────────────────────────────────────────────────────
    rech = mov[mov["_rechazada"]]
    if not rech.empty:
        mot = (rech.groupby("Motivo del rechazo")
               .agg(**{"Rechazos": ("_doc", "count"),
                       "Instalaciones": ("tipo_mov",
                                         lambda s: int((s == "nueva").sum())),
                       "Retiros": ("tipo_mov",
                                   lambda s: int((s == "retiro").sum())),
                       "Clientes": ("cliente_rut", "nunique")})
               .reset_index().sort_values("Rechazos", ascending=False))
        mot["% del total"] = mot["Rechazos"] / mot["Rechazos"].sum()
        mot = _con_total(mot, "Motivo del rechazo", ("Clientes",))
        _escribir(wb, "Rechazos", mot,
                  {"Rechazos": _FMT_NUM, "Instalaciones": _FMT_NUM,
                   "Retiros": _FMT_NUM, "Clientes": _FMT_NUM,
                   "% del total": _FMT_PCT},
                  nota=("Por qué volvieron. El motivo se lee del comentario que "
                        "escribe el repartidor: el campo 'Motivo rechazo' del "
                        "ERP llega vacío. Ojo con la lectura: que el cliente no "
                        "entregue la máquina o se arrepienta de recibirla es "
                        "coordinación comercial, no un problema de ruta."),
                  total_ultima=not mot.empty)

        _escribir(wb, "Rechazos · detalle", _detalle(rech.sort_values("fecha")),
                  nota="Una fila por rechazo, con el comentario textual.")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _mapa_vendedor(mov: pd.DataFrame) -> dict:
    """id → nombre, reutilizando lo que ya trae el detalle de movimientos."""
    if mov.empty or "Vendedor" not in mov.columns:
        return {}
    return dict(mov.drop_duplicates("vendedor_id")
                .set_index("vendedor_id")["Vendedor"])


def _pintar_estado(ws, df: pd.DataFrame, col: str):
    """Semáforo en la columna de estado del tablero."""
    from openpyxl.styles import Font, PatternFill

    if df.empty or col not in df.columns:
        return
    j = list(df.columns).index(col) + 1
    colores = {"✔": ("E7F3EC", "1A7F4B"), "✘": ("FAEAE8", "B2332A"),
               "—": ("F1F1F4", "64748B")}
    for i in range(len(df)):
        celda = ws.cell(row=4 + i, column=j)
        fondo, letra = colores.get(str(celda.value or "")[:1], (None, None))
        if fondo:
            celda.fill = PatternFill("solid", fgColor=fondo)
            celda.font = Font(bold=True, color=letra)
