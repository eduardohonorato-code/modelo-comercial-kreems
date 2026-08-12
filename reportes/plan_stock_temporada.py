# -*- coding: utf-8 -*-
"""Plan de stock TEMPORADA ALTA: 16 meses (ene-2026 a abr-2027).
Demanda en cajas por SKU + stock minimo/ideal + grillas y $ por CD.

REFRESCO MENSUAL: el corte entre meses REALES y PROYECTADOS se detecta solo a
partir de la ultima fecha cargada en Supabase (ultimo mes COMPLETO = real).
Basta con volver a correr este script despues de cargar el mes que cerro.
"""
import os
import sys
import calendar
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from etl.db import get_client
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.errors import IgnoredError, IgnoredErrors


def sin_triangulos(ws, rango):
    """Apaga el aviso de Excel 'fórmula distinta a las vecinas' (triángulo verde).
    En estas grillas hay filas que legítimamente usan otra fórmula (galletas, y las
    celdas de 2027 reestimadas por quiebre), así que el aviso es ruido."""
    ws.ignored_errors = IgnoredErrors(
        ignoredError=[IgnoredError(sqref=rango, formula=True, formulaRange=True)])

sb = get_client()
ANIO = 2026

# Dónde se guarda el archivo. Si la carpeta de Drive no está disponible (o el
# archivo está abierto en Excel), cae a reportes/ del propio repo.
SALIDA_DRIVE = Path(r"G:\Mi unidad\Reportes Financieros\Entregables Valorizacion"
                    r"\Entregables Due Diligence\Proyecciones\plan_stock_temporada_2026_2027.xlsx")
SALIDA_LOCAL = Path(__file__).resolve().parent / "plan_stock_temporada_2026_2027.xlsx"

def fetch(table, select, filt=None, orders=("fecha",)):
    rows, page = [], 0
    while True:
        q = sb.table(table).select(select)
        if filt: q = filt(q)
        for o in orders: q = q.order(o)
        r = q.range(page * 1000, page * 1000 + 999).execute()
        rows.extend(r.data or [])
        if not r.data or len(r.data) < 1000: break
        page += 1
    return pd.DataFrame(rows)

# ── datos ──────────────────────────────────────────────────────────────────
prod = fetch("dim_producto", "codigo,nombre,categoria", orders=("codigo",)).set_index("codigo")

v = fetch("fact_ventas", "fecha,tipo_dcto,producto_codigo,cantidad,neto,sucursal,sociedad_id",
          lambda q: q.gte("fecha", f"{ANIO}-01-01").lte("fecha", f"{ANIO}-12-31"),
          orders=("fecha", "n_dcto", "linea"))
v["fecha"] = pd.to_datetime(v["fecha"]); v["mes"] = v["fecha"].dt.month
# Defensa: las NC cargadas por API traían cantidad POSITIVA (sumaban cajas en vez
# de restarlas). Corregido en el loader + sql/035; esto deja el archivo correcto
# aunque la reparación aún no se haya corrido. Si ya se corrió, es un no-op.
_nc_mal = v["tipo_dcto"].str.contains("CREDITO", na=False) & (v["cantidad"] > 0)
if _nc_mal.any():
    print(f"  [aviso] {_nc_mal.sum()} líneas de NC con cantidad positiva -> corregidas en memoria "
          f"({v.loc[_nc_mal, 'cantidad'].sum():.0f} cajas). Correr sql/035 para arreglarlo en la base.")
    v.loc[_nc_mal, "cantidad"] = -v.loc[_nc_mal, "cantidad"]
v = v.join(prod, on="producto_codigo")
v = v[~v["categoria"].isin(["Servicios", "Maquinas"])].copy()

# CORTE AUTOMATICO: ultimo mes COMPLETO, exigiendo que las sociedades ACTIVAS
# esten al dia. Gran Natural entra por API (siempre al dia); Acuña se carga a
# mano desde Excel, asi que un mes suyo a medio cargar no debe tomarse como real.
SOC = {1: "Acuña (Excel manual)", 2: "Gran Natural (API)"}

# Sociedades DADAS DE BAJA: dejaron de facturar de verdad, no es falta de carga.
# Acuña quedo sin operacion por la transicion a Gran Natural; su ultima venta
# real es del 13-jul-2026. No debe bloquear el cierre de ningun mes (si lo
# hiciera, el plan quedaria esperando datos que nunca van a llegar).
SOCIEDAD_BAJA = {1: date(2026, 7, 13)}

def ultimo_mes_completo(fechas):
    f = fechas.max().date()
    return f, (f.month if f.day >= calendar.monthrange(f.year, f.month)[1] else f.month - 1)

print("Cobertura por sociedad:")
cortes = {}
for sid, nom in SOC.items():
    fs = v.loc[v["sociedad_id"] == sid, "fecha"]
    if fs.empty:
        continue
    f, u = ultimo_mes_completo(fs)
    if sid in SOCIEDAD_BAJA:
        baja = SOCIEDAD_BAJA[sid]
        if f > baja:
            # volvio a facturar: ya no esta de baja, hay que actualizar la constante
            print(f"   {nom:26s} última venta {f}  [!] FACTURÓ DESPUÉS DE LA BAJA ({baja})")
            print(f"      -> vuelve a operar: actualizar SOCIEDAD_BAJA en este script.")
            cortes[sid] = u
        else:
            print(f"   {nom:26s} última venta {f}  (sociedad de baja: no limita el corte)")
        continue
    cortes[sid] = u
    print(f"   {nom:26s} última venta {f} -> completo hasta mes {u:02d}")

ULT_FECHA = v["fecha"].max().date()
U = min(cortes.values()) if cortes else 0
rezagadas = [SOC[s] for s, u in cortes.items() if u < max(cortes.values())]
if rezagadas:
    print(f"\n  [!] CORTE LIMITADO POR: {', '.join(rezagadas)}")
    print(f"      El mes {max(cortes.values()):02d} NO se toma como real porque esa fuente no está")
    print(f"      cargada hasta fin de mes. Corre 'Cargar mes.bat' y vuelve a intentarlo.\n")
print(f"Meses REALES: ene..{U:02d}/{ANIO} | proyectados: {U+1:02d}..12 + ene-abr 2027")

vr = v[v["mes"] <= U].copy()          # solo meses completos
es_gall_cat = lambda c: c == "Galletas"

# estacionalidad: promedio de las curvas de share 2024+2025 por GRUPO estable
vh = fetch("fact_ventas", "fecha,producto_codigo,neto",
           lambda q: q.gte("fecha", "2024-01-01").lte("fecha", "2025-12-31"),
           orders=("fecha", "n_dcto", "linea"))
vh["fecha"] = pd.to_datetime(vh["fecha"])
vh["anio"] = vh["fecha"].dt.year; vh["mes"] = vh["fecha"].dt.month
vh = vh.join(prod, on="producto_codigo")
# Entre 2024 y 2025 la empresa renombro categorias ('Paletas'->'Helados Paletas')
# y recodifico SKUs ('CR-6'->'CR-6x20'); se agrupa por familia estable.
GRUPO = {"Paletas": "PALETAS", "Helados Paletas": "PALETAS", "MULTIPACK": "PALETAS",
         "Helados multipack": "PALETAS", "Helados Pote": "POTES", "POTE": "POTES",
         "Helados Bacha": "BACHA", "BACHA": "BACHA"}
vh["grupo"] = vh["categoria"].map(GRUPO).fillna("OTROS")

def perfil(df):
    profs = []
    for _, g in df.groupby("anio"):
        p = g.groupby("mes")["neto"].sum().reindex(range(1, 13), fill_value=0.0).clip(lower=0)
        if p.sum() > 0: profs.append(p / p.sum())
    return sum(profs) / len(profs) if profs else None

perf_global = perfil(vh)
perf_grupo = {g: perfil(sub) for g, sub in vh.groupby("grupo") if perfil(sub) is not None}

# ── plan oficial 2026 (curva re-derivada) ──────────────────────────────────
TEND = [177000516, 169841061, 132491389, 71954908, 61668854, 60337588,
        45699497, 78442259, 96730126, 148376029, 151767585, 155763863]
SERV = [0, 3914133, 1479846, 1480201, 1421384, 1349166] + [1400000] * 6
GALL_PLAN_ORIG = [0]*5 + [1563035] + [2097703 * (1.15 ** k) for k in range(6)]
plan_total = [TEND[i] - SERV[i] - (1563035 if i == 5 else 0) + GALL_PLAN_ORIG[i] for i in range(12)]
plan_helados = [plan_total[i] - GALL_PLAN_ORIG[i] for i in range(12)]
META_HELADOS = round(sum(plan_helados))
CREC_REAL_2027 = 0.0511
print(f"Plan oficial: total ${sum(plan_total)/1e6:,.1f}M | helados ${META_HELADOS/1e6:,.1f}M | galletas ${sum(GALL_PLAN_ORIG)/1e6:,.1f}M")

# ── mes en curso (parcial): solo informativo, NO entra al modelo ───────────
# Un mes a medias es muy ruidoso para proyectar (el % facturado al mismo dia
# oscila entre 22% y 50% segun el mes), asi que no se mezcla con la demanda.
# Pero sirve como semaforo: se muestra en Parametros para decidir a tiempo.
curso = {}
if U < 12:
    vc = v[v["mes"] == U + 1]
    if not vc.empty:
        dia = int(vc["fecha"].max().day)
        facturado = float(vc["neto"].sum())
        shares = []
        for m in range(1, U + 1):
            sub = v[v["mes"] == m]
            tot = float(sub["neto"].sum())
            if tot > 0:
                shares.append(float(sub.loc[sub["fecha"].dt.day <= dia, "neto"].sum()) / tot)
        sh = sum(shares) / len(shares) if shares else None
        curso = {"mes": U + 1, "dia": dia, "facturado": facturado,
                 "share": sh, "cierre": (facturado / sh) if sh else None,
                 "min_share": min(shares) if shares else None,
                 "max_share": max(shares) if shares else None}
        print(f"Mes en curso ({U+1:02d}, al día {dia}): facturado ${facturado/1e6:,.1f}M | "
              f"al ritmo histórico cerraría en ${curso['cierre']/1e6:,.1f}M")

# ── base por SKU ───────────────────────────────────────────────────────────
real = vr.pivot_table(index="producto_codigo", columns="mes", values="cantidad",
                      aggfunc="sum", fill_value=0.0)
# Ingreso REAL facturado por SKU/mes (y por CD): en los meses cerrados las hojas
# de $ muestran lo efectivamente facturado, no cajas × precio promedio. El precio
# por caja se movió fuerte (jul $20.883 vs $23.703 ene–may), así que un promedio
# único distorsionaría meses que ya son un hecho.
neto_real = vr.pivot_table(index="producto_codigo", columns="mes", values="neto",
                           aggfunc="sum", fill_value=0.0)
neto_r = vr.groupby("producto_codigo")["neto"].sum()
cajas_r = real.sum(axis=1)
precio = (neto_r / cajas_r.replace(0, np.nan)).rename("precio_caja")
skus = [c for c in cajas_r.index
        if cajas_r[c] > 10 and neto_r.get(c, 0) > 200000 and not pd.isna(precio.get(c))]
es_gall = {c for c in skus if es_gall_cat(prod.loc[c, "categoria"])}
helados = [c for c in skus if c not in es_gall]
print(f"SKUs: {len(skus)} ({len(helados)} helados + {len(es_gall)} galletas)")

dem = pd.DataFrame(0.0, index=skus, columns=range(1, 13))
for c in skus:
    for m in range(1, U + 1):
        dem.loc[c, m] = max(0.0, real.loc[c, m] if m in real.columns else 0.0)

# HELADOS: forma estacional del grupo aplicada al acumulado real
for c in helados:
    p = perf_grupo.get(GRUPO.get(prod.loc[c, "categoria"], "OTROS"))
    if p is None: p = perf_global
    base = p[list(range(1, U + 1))].sum()
    if base <= 0.02:
        p = perf_global; base = p[list(range(1, U + 1))].sum()
    acum = dem.loc[c, 1:U].sum()
    for m in range(U + 1, 13):
        dem.loc[c, m] = acum * p[m] / base

# Calce de helados al presupuesto restante (META - real acumulado).
# El factor resulta ser el MISMO para todos los meses proyectados
# (obj_m = resto * peso_m/total  =>  f = obj_m/peso_m = resto/total), asi que se
# puede dejar como UNA celda en Parametros y que las celdas proyectadas la
# multipliquen: cambiar la META en el Excel recalcula todo sin regenerar nada.
dem_pre = dem.copy()                      # proyeccion sin escalar
hel_real = sum(sum(dem.loc[c, m] * precio[c] for m in range(1, U + 1)) for c in helados)
resto = META_HELADOS - hel_real
BASE_H2 = sum(dem_pre.loc[c, m] * precio[c] for c in helados for m in range(U + 1, 13))
FACTOR = resto / BASE_H2 if BASE_H2 > 0 else 1.0
for m in range(U + 1, 13):
    for c in helados:
        dem.loc[c, m] = dem_pre.loc[c, m] * FACTOR
print(f"Helados: real ene-{U:02d} ${hel_real/1e6:,.1f}M | resto {U+1}-12 ${resto/1e6:,.1f}M "
      f"| factor de calce {FACTOR:.4f}")

# GALLETAS: proyeccion desde el ultimo mes real con rampa (tienen historia propia
# desde jun-2026, no hay estacionalidad historica que aplicar todavia).
gal_real = {m: sum(dem.loc[c, m] * precio[c] for c in es_gall) for m in range(1, U + 1)}
base_gal = gal_real.get(U, 0.0)
RAMPA_GALL = 0.08
mix_gal = {}
if es_gall:
    ult = {c: dem.loc[c, U] * precio[c] for c in es_gall}
    s = sum(ult.values())
    mix_gal = {c: (ult[c] / s if s > 0 else 1 / len(es_gall)) for c in es_gall}
    for k, m in enumerate(range(U + 1, 13), start=1):
        objetivo = base_gal * ((1 + RAMPA_GALL) ** k)
        for c in es_gall:
            dem.loc[c, m] = objetivo * mix_gal[c] / precio[c]
print("Galletas reales por mes ($):", {m: round(x) for m, x in gal_real.items() if x > 0})
print(f"Galletas: base {U:02d} ${base_gal/1e6:,.2f}M -> dic ${base_gal*(1+RAMPA_GALL)**(12-U)/1e6:,.2f}M (rampa {RAMPA_GALL:.0%}/mes)")

# ── base ene-abr 2027: tapar huecos de quiebre ─────────────────────────────
grupo_de = {c: GRUPO.get(prod.loc[c, "categoria"], "OTROS") for c in skus}
grupo_mes = {}
for g in set(grupo_de[c] for c in helados):
    miembros = [c for c in helados if grupo_de[c] == g]
    grupo_mes[g] = {m: sum(dem.loc[c, m] for c in miembros) for m in range(1, 13)}
base27, huecos = {}, []
for c in helados:
    g = grupo_de[c]
    meses_ok = [m for m in range(1, 13) if dem.loc[c, m] > 0]
    den = sum(grupo_mes[g][m] for m in meses_ok)
    cuota = (sum(dem.loc[c, m] for m in meses_ok) / den) if den > 0 else 0.0
    for k in range(1, 5):
        est = grupo_mes[g][k] * cuota
        if est > 0 and dem.loc[c, k] < 0.25 * est:
            base27[(c, k)] = est
            huecos.append((c, k, round(dem.loc[c, k], 1), round(est, 1)))
if huecos:
    print(f"Huecos de quiebre tapados en la base 2027: {len(huecos)} celdas")

# ── split por CD (incluye galletas) ────────────────────────────────────────
MAP_CD = {"CONCEPCION": "CONCEPCION", "SANTIAGO": "SANTIAGO", "TEMUCO": "TEMUCO",
          "C. Matriz": "CONCEPCION", "VALPARAISO": "SANTIAGO"}
v["cd"] = v["sucursal"].map(MAP_CD).fillna("CONCEPCION")
cd_share = (v[v["producto_codigo"].isin(skus)]
            .pivot_table(index="producto_codigo", columns="cd", values="neto", aggfunc="sum", fill_value=0.0))
for col in ["CONCEPCION", "SANTIAGO", "TEMUCO"]:
    if col not in cd_share: cd_share[col] = 0.0
cd_share = cd_share[["CONCEPCION", "SANTIAGO", "TEMUCO"]].clip(lower=0)
glob = cd_share.sum() / cd_share.sum().sum()
cd_share = cd_share.div(cd_share.sum(axis=1), axis=0).fillna(glob)
print("Split global CD:", (glob * 100).round(1).to_dict())

# Ingreso real facturado por SKU × CD × mes (meses cerrados)
neto_real_cd = (v[v["mes"] <= U]
                .pivot_table(index=["producto_codigo", "cd"], columns="mes",
                             values="neto", aggfunc="sum", fill_value=0.0))

# ── layout ─────────────────────────────────────────────────────────────────
MESES = [f"{n}-26" for n in ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]] + \
        [f"{n}-27" for n in ["Ene","Feb","Mar","Abr"]]
NM = 16
COL_TOT26, COL_TEMP, COL_PRECIO = 21, 22, 24
TEMP_MESES = list(range(9, 17))
ARIAL = "Arial"
F_TITLE = Font(name=ARIAL, size=13, bold=True); F_HDR = Font(name=ARIAL, size=10, bold=True)
F_TXT = Font(name=ARIAL, size=10); F_NOTE = Font(name=ARIAL, size=9, italic=True, color="666666")
F_INPUT = Font(name=ARIAL, size=10, color="0000FF")
F_HDR_T = Font(name=ARIAL, size=10, bold=True, color="7A0040")
F_FIX = Font(name=ARIAL, size=10, color="B36B00")
F_REAL = Font(name=ARIAL, size=9, bold=True, color="1B5E20")
F_PROY = Font(name=ARIAL, size=9, bold=True, color="8A6D00")
FILL_HDR = PatternFill("solid", fgColor="F2D5E4"); FILL_TOT = PatternFill("solid", fgColor="E8E8E8")
FILL_YEL = PatternFill("solid", fgColor="FFFF00"); FILL_TEMP = PatternFill("solid", fgColor="FCE4EC")
FILL_FIX = PatternFill("solid", fgColor="FFF2CC"); FILL_REAL = PatternFill("solid", fgColor="E8F5E9")
CAJ = "#,##0;(#,##0);-"; CLP = "$#,##0;($#,##0);-"; PCT = "0.0%"

orden = [c for c in neto_r.loc[skus].sort_values(ascending=False).index]
wb = Workbook()

# ── Parametros ─────────────────────────────────────────────────────────────
ws = wb.active; ws.title = "Parametros"; ws.sheet_view.showGridLines = False
ws["B1"] = "Plan de Stock TEMPORADA ALTA (ene-2026 → abr-2027) — Parámetros"; ws["B1"].font = F_TITLE
ws["B2"] = (f"DATOS AL {ULT_FECHA.strftime('%d-%m-%Y')} · Meses REALES: enero a {MESES[U-1]} · "
            f"Proyectados: {MESES[U]} en adelante.")
ws["B2"].font = Font(name=ARIAL, size=10, bold=True, color="1B5E20")
params = [("Semanas de stock de SEGURIDAD (mínimo)", 2.0, "0.0"),
          ("Semanas de ciclo de REPOSICIÓN", 4.0, "0.0"),
          ("Crecimiento REAL de cajas 2027 vs 2026", CREC_REAL_2027, PCT),
          ("Ajuste de PRECIO 2027 (inflación) — solo hojas de $", 0.03, PCT)]
for i, (lbl, val, fmt) in enumerate(params):
    ws.cell(row=4 + i, column=2, value=lbl).font = F_TXT
    cc = ws.cell(row=4 + i, column=3, value=val)
    cc.font = F_INPUT; cc.fill = FILL_YEL; cc.number_format = fmt
ws["B8"] = "META de venta HELADOS 2026 ($) — reparte los meses proyectados"; ws["B8"].font = F_TXT
ws["C8"] = META_HELADOS; ws["C8"].font = F_INPUT; ws["C8"].fill = FILL_YEL; ws["C8"].number_format = CLP
ws["B9"] = f"Rampa mensual de GALLETAS ({MESES[U]} en adelante)"; ws["B9"].font = F_TXT
ws["C9"] = RAMPA_GALL; ws["C9"].font = F_INPUT; ws["C9"].fill = FILL_YEL; ws["C9"].number_format = PCT
# Cadena de calce: las celdas proyectadas de la hoja Demanda multiplican por F7,
# asi que editar la META (C8) recalcula demanda, stock, CD e ingresos en cascada.
sh_ytd = perf_global[list(range(1, U + 1))].sum()
ritmo = hel_real / sh_ytd if sh_ytd > 0 else 0
ws["E4"] = f"Helados REALES ene–{MESES[U-1]} ($)"; ws["F4"] = round(hel_real)
ws["E5"] = "Resto del año a repartir ($) = META − real"; ws["F5"] = "=C8-F4"
ws["E6"] = f"Proyección {MESES[U]}–Dic sin escalar ($)"; ws["F6"] = round(BASE_H2)
ws["E7"] = "FACTOR DE CALCE = resto ÷ proyección sin escalar"; ws["F7"] = "=F5/F6"
ws["E8"] = "Referencia: año implícito al RITMO REAL ($)"; ws["F8"] = round(ritmo)
for cc in ("F4", "F5", "F6", "F8"): ws[cc].font = F_TXT; ws[cc].number_format = CLP
ws["F7"].font = Font(name=ARIAL, size=10, bold=True); ws["F7"].number_format = "0.0000"
ws["F8"].font = Font(name=ARIAL, size=10, bold=True, color="B36B00")
for cc in ("E4", "E5", "E6", "E7", "E8"): ws[cc].font = F_NOTE
ws.column_dimensions["B"].width = 118; ws.column_dimensions["C"].width = 14
ws.column_dimensions["E"].width = 48; ws.column_dimensions["F"].width = 17
# ── panel del mes en curso (semáforo, no entra al modelo) ──────────────────
if curso:
    F_AL = Font(name=ARIAL, size=10, bold=True, color="B36B00")
    plan_curso = plan_total[curso["mes"] - 1]
    desv = curso["cierre"] / plan_curso - 1 if curso["cierre"] else None
    ws["E9"] = f"MES EN CURSO — {MESES[curso['mes']-1]} (al día {curso['dia']})"
    ws["E9"].font = F_HDR; ws["E9"].fill = FILL_YEL
    filas_curso = [
        ("Facturado hasta hoy", curso["facturado"], CLP, F_TXT),
        ("Plan del mes", plan_curso, CLP, F_TXT),
        (f"% del mes que suele ir al día {curso['dia']}", curso["share"], PCT, F_TXT),
        ("Cierre estimado a ese ritmo", curso["cierre"], CLP, F_AL),
        ("Desvío estimado vs plan del mes", desv, PCT, F_AL),
    ]
    for i, (lbl, val, fmt, fnt) in enumerate(filas_curso):
        ws.cell(row=10 + i, column=5, value=lbl).font = F_NOTE
        cc = ws.cell(row=10 + i, column=6, value=val)
        cc.font = fnt; cc.number_format = fmt
    ws["E15"] = (f"Referencia, NO entra en la proyección: el % facturado al día {curso['dia']} "
                 f"osciló entre {curso['min_share']:.0%} y {curso['max_share']:.0%} en los meses cerrados,")
    ws["E16"] = ("así que el cierre estimado es orientativo. Sirve para ver a tiempo si el mes se está "
                 "desviando del plan;")
    ws["E17"] = ("si la desviación se repite mes a mes, bajar la META en vez de exigirle la recuperación "
                 "a los meses que quedan.")
    for c in ("E15", "E16", "E17"):
        ws[c].font = F_NOTE

ws["B10"] = "Cómo se actualiza (proceso mensual):"; ws["B10"].font = F_HDR
notas = [
    f"1) Cargar en Supabase el mes que cerró.  2) Volver a correr el generador.  3) Listo: el corte real/proyectado se mueve solo.",
    f"   Hoy el corte quedó en {MESES[U-1]} (último mes completo). Las columnas REALES van marcadas en verde en cada hoja.",
    "",
    f"REAL (ene–{MESES[U-1]}): cajas efectivamente facturadas, netas de notas de crédito. No se tocan.",
    f"PROYECTADO ({MESES[U]}–Dic-26): forma estacional validada 2024+2025 por grupo (paletas/potes/bacha) aplicada al",
    "   acumulado real de cada SKU, y escalada para que el año de helados cierre en la META (celda C8).",
    "   → Si un mes real viene bajo el plan, su diferencia se reparte en los meses que quedan. Ver 'año implícito al",
    "     ritmo real' arriba: si se aleja mucho de la META, conviene bajar la META en vez de exigir la recuperación.",
    "GALLETAS: producto nuevo (lanzado jun-2026). No tiene estacionalidad histórica, así que se proyecta desde el último",
    f"   mes real con una rampa de {RAMPA_GALL:.0%} mensual. Ya van por encima del plan original, por eso suman sobre la META",
    "   de helados en vez de competir con las paletas. Revisar la rampa cada mes contra lo que vaya pasando.",
    "ENE–ABR 2027: mes equivalente de 2026 × (1 + crecimiento REAL de cajas). Solo real: la inflación sube el $, no las cajas.",
    "   Celdas NARANJAS: el mes base de 2026 fue quiebre de stock (venta ~0 por falta de producto). Se reemplaza por la",
    "   'cuota de familia' del SKU dentro de su grupo; copiar el cero habría dejado el SKU sin stock en el peak.",
    "",
    "Stock MÍNIMO fin de mes m = demanda del mes m+1 × (semanas seguridad ÷ 4,33). Piso: bajo esto hay riesgo de quiebre.",
    "Stock IDEAL fin de mes m = demanda del mes m+1 × (seguridad + reposición) ÷ 4,33. 4,33 = 52 semanas ÷ 12 meses.",
    "Para abr-2027 la demanda de may-2027 se estima como may-2026 × (1 + crecimiento real).",
    "La frecuencia real de envío la define logística; este archivo entrega NIVELES objetivo, no el calendario.",
    "",
    "Split por CD: venta 2026 por sucursal. 'C. Matriz' = CONCEPCION. Temuco es el complemento (=100%−CCP−STGO).",
    "Hojas de $: ingreso = cajas × precio/caja del SKU (promedio real de los meses cerrados). 2027 lleva el ajuste de precio.",
]
for i, n in enumerate(notas):
    ws.cell(row=11 + i, column=2, value=n).font = F_NOTE

# ── helper de grilla ───────────────────────────────────────────────────────
def hoja_grid(nombre, titulo, nota, con_totales=True, ancho_mes=9):
    ws = wb.create_sheet(nombre); ws.sheet_view.showGridLines = False
    ws["B1"] = titulo; ws["B1"].font = F_TITLE
    ws["B2"] = nota; ws["B2"].font = F_NOTE
    for j in range(NM):
        et = ws.cell(row=3, column=5 + j, value="REAL" if (j + 1) <= U else "PROY")
        et.font = F_REAL if (j + 1) <= U else F_PROY
        et.alignment = __import__("openpyxl").styles.Alignment(horizontal="center")
        if (j + 1) <= U: et.fill = FILL_REAL
    ws["B3"] = f"Datos al {ULT_FECHA.strftime('%d-%m-%Y')}"; ws["B3"].font = F_NOTE
    ws["B4"] = "Código"; ws["C4"] = "Producto"; ws["D4"] = "Categoría"
    for j, mm in enumerate(MESES):
        c = ws.cell(row=4, column=5 + j, value=mm)
        c.font = F_HDR_T if (j + 1) in TEMP_MESES else F_HDR
        c.fill = FILL_TEMP if (j + 1) in TEMP_MESES else FILL_HDR
    if con_totales:
        ws.cell(row=4, column=COL_TOT26, value="Total 2026").font = F_HDR
        ws.cell(row=4, column=COL_TOT26).fill = FILL_HDR
        ws.cell(row=4, column=COL_TEMP, value="TEMPORADA sep26–abr27").font = F_HDR_T
        ws.cell(row=4, column=COL_TEMP).fill = FILL_TEMP
    for col in range(2, 5):
        ws.cell(row=4, column=col).font = F_HDR; ws.cell(row=4, column=col).fill = FILL_HDR
    ws.column_dimensions["B"].width = 12; ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 17
    for j in range(NM): ws.column_dimensions[get_column_letter(5 + j)].width = ancho_mes
    ws.column_dimensions[get_column_letter(COL_TOT26)].width = 12
    ws.column_dimensions[get_column_letter(COL_TEMP)].width = 14
    ws.freeze_panes = "E5"
    return ws

# ── Demanda ────────────────────────────────────────────────────────────────
ws = hoja_grid("Demanda", "Demanda por SKU en CAJAS — ene-2026 a abr-2027",
               f"Verde = meses REALES (ene–{MESES[U-1]}). Rosa = temporada alta. Naranja = base 2027 reestimada por quiebre.")
r = 5
fila_sku = {}
for c in orden:
    fila_sku[c] = r
    ws.cell(row=r, column=2, value=c).font = F_TXT
    ws.cell(row=r, column=3, value=prod.loc[c, "nombre"]).font = F_TXT
    ws.cell(row=r, column=4, value=prod.loc[c, "categoria"]).font = F_TXT
    for m in range(1, NM + 1):
        col = 4 + m
        if m <= U:                       # mes cerrado: cajas reales, valor fijo
            cc = ws.cell(row=r, column=col, value=round(float(dem.loc[c, m]), 2))
            cc.font = F_INPUT
        elif m <= 12:                    # proyeccion 2026: colgada de Parametros
            if c in es_gall:             # galletas: ultimo mes real x rampa
                k = m - U
                cc = ws.cell(row=r, column=col,
                             value=f"={get_column_letter(4 + U)}{r}*(1+Parametros!$C$9)^{k}")
            else:                        # helados: proyeccion x factor de calce (META)
                cc = ws.cell(row=r, column=col,
                             value=f"={round(float(dem_pre.loc[c, m]), 2)}*Parametros!$F$7")
            cc.font = F_TXT
        else:
            k = m - 12
            if c in es_gall:                       # galletas: nivel dic-26 con crecimiento
                cc = ws.cell(row=r, column=col, value=f"=$P{r}*(1+Parametros!$C$6)")
            elif (c, k) in base27:
                cc = ws.cell(row=r, column=col,
                             value=f"={round(base27[(c, k)], 2)}*(1+Parametros!$C$6)")
                cc.font = F_FIX
            else:
                cc = ws.cell(row=r, column=col,
                             value=f"={get_column_letter(4 + k)}{r}*(1+Parametros!$C$6)")
            if not isinstance(cc.font, Font) or cc.font.color is None: cc.font = F_TXT
        cc.number_format = CAJ
        if m <= U: cc.fill = FILL_REAL
        elif m in TEMP_MESES: cc.fill = FILL_TEMP
        if m > 12 and (c, m - 12) in base27: cc.fill = FILL_FIX
    t1 = ws.cell(row=r, column=COL_TOT26, value=f"=SUM(E{r}:P{r})"); t1.font = F_TXT; t1.number_format = CAJ
    t2 = ws.cell(row=r, column=COL_TEMP, value=f"=SUM(M{r}:T{r})")
    t2.font = F_HDR; t2.number_format = CAJ; t2.fill = FILL_TEMP
    pc = ws.cell(row=r, column=COL_PRECIO, value=round(float(precio[c]), 0))
    pc.font = F_INPUT; pc.number_format = CLP
    r += 1
ws.cell(row=4, column=COL_PRECIO, value="Precio/caja").font = F_HDR
ws.column_dimensions[get_column_letter(COL_PRECIO)].width = 13
rt = r
ws.cell(row=rt, column=2, value="TOTAL CAJAS").font = F_HDR
ws.cell(row=rt, column=2).fill = FILL_TOT
for m in range(1, NM + 1):
    L = get_column_letter(4 + m)
    cc = ws.cell(row=rt, column=4 + m, value=f"=SUM({L}5:{L}{rt-1})")
    cc.font = F_HDR; cc.number_format = CAJ
    cc.fill = FILL_REAL if m <= U else (FILL_TEMP if m in TEMP_MESES else FILL_TOT)
for col in (COL_TOT26, COL_TEMP):
    L = get_column_letter(col)
    cc = ws.cell(row=rt, column=col, value=f"=SUM({L}5:{L}{rt-1})")
    cc.font = F_HDR; cc.number_format = CAJ; cc.fill = FILL_TOT
sin_triangulos(ws, f"E5:{get_column_letter(COL_TEMP)}{rt}")
n_skus = len(orden)

# ── Stock minimo / ideal ───────────────────────────────────────────────────
def hoja_stock(nombre, titulo, mult):
    ws = hoja_grid(nombre, titulo,
                   "Fin de mes, en cajas. Mira la demanda del mes SIGUIENTE, por eso no calza con la demanda del mes.",
                   con_totales=False)
    r = 5
    for c in orden:
        rd = fila_sku[c]
        ws.cell(row=r, column=2, value=c).font = F_TXT
        ws.cell(row=r, column=3, value=prod.loc[c, "nombre"]).font = F_TXT
        ws.cell(row=r, column=4, value=prod.loc[c, "categoria"]).font = F_TXT
        for m in range(1, NM + 1):
            nxt = (f"Demanda!{get_column_letter(5 + m)}{rd}" if m < NM
                   else f"(Demanda!I{rd}*(1+Parametros!$C$6))")
            cc = ws.cell(row=r, column=4 + m, value=f"=MAX(0,{nxt}*{mult})")
            cc.font = F_TXT; cc.number_format = CAJ
            if m in TEMP_MESES: cc.fill = FILL_TEMP
        r += 1
    ws.cell(row=r, column=2, value="TOTAL CAJAS").font = F_HDR
    ws.cell(row=r, column=2).fill = FILL_TOT
    for m in range(1, NM + 1):
        L = get_column_letter(4 + m)
        cc = ws.cell(row=r, column=4 + m, value=f"=SUM({L}5:{L}{r-1})")
        cc.font = F_HDR; cc.number_format = CAJ
        cc.fill = FILL_TEMP if m in TEMP_MESES else FILL_TOT
    sin_triangulos(ws, f"E5:{get_column_letter(4 + NM)}{r}")
    return ws

hoja_stock("Stock Minimo", "Stock MÍNIMO por SKU (cajas, fin de mes) — ene-26 a abr-27",
           "Parametros!$C$4/4.33")
hoja_stock("Stock Ideal", "Stock IDEAL por SKU (cajas, fin de mes) — ene-26 a abr-27",
           "(Parametros!$C$4+Parametros!$C$5)/4.33")

# ── CD Split ───────────────────────────────────────────────────────────────
ws = wb.create_sheet("CD Split"); ws.sheet_view.showGridLines = False
ws["B1"] = "Distribución por Centro de Distribución (% por SKU)"; ws["B1"].font = F_TITLE
ws["B2"] = ("Según venta 2026 por sucursal. 'C. Matriz' = CONCEPCION. Editables (azul) Concepción y Santiago; "
            "Temuco es el complemento (=100%−CCP−STGO) para que los tres sumen exacto.")
ws["B2"].font = F_NOTE
ws["B4"] = "Código"; ws["C4"] = "Producto"
for j, cd in enumerate(["CONCEPCION", "SANTIAGO", "TEMUCO"]):
    ws.cell(row=4, column=4 + j, value=cd)
for col in range(2, 7):
    ws.cell(row=4, column=col).font = F_HDR; ws.cell(row=4, column=col).fill = FILL_HDR
r = 5
for c in orden:
    ws.cell(row=r, column=2, value=c).font = F_TXT
    ws.cell(row=r, column=3, value=prod.loc[c, "nombre"]).font = F_TXT
    for j, cd in enumerate(["CONCEPCION", "SANTIAGO", "TEMUCO"]):
        vv = float(cd_share.loc[c, cd]) if c in cd_share.index else float(glob[cd])
        if cd == "TEMUCO":
            cc = ws.cell(row=r, column=4 + j, value=f"=1-D{r}-E{r}"); cc.font = F_TXT
        else:
            cc = ws.cell(row=r, column=4 + j, value=round(vv, 6)); cc.font = F_INPUT
        cc.number_format = PCT
    r += 1
ws.column_dimensions["B"].width = 12; ws.column_dimensions["C"].width = 38
for col in "DEF": ws.column_dimensions[col].width = 13
ws.freeze_panes = "D5"

# ── Hojas por CD (ahora incluyen galletas) ─────────────────────────────────
CD_COL = {"CONCEPCION": "D", "SANTIAGO": "E", "TEMUCO": "F"}
bloques_cd = {}
for cd_nombre, share_col in CD_COL.items():
    ws = wb.create_sheet(f"CD - {cd_nombre.title()}"); ws.sheet_view.showGridLines = False
    ws["B1"] = f"CD {cd_nombre.title()} — Demanda y Stock (cajas) — ene-26 a abr-27"; ws["B1"].font = F_TITLE
    ws["B2"] = ("DEMANDA (venta del mes) · STOCK MÍNIMO · STOCK IDEAL (estos miran el mes SIGUIENTE). "
                "Cada celda = valor nacional × % del SKU en este CD. Incluye galletas.")
    ws["B2"].font = F_NOTE
    ws["B3"] = f"Datos al {ULT_FECHA.strftime('%d-%m-%Y')} · REAL hasta {MESES[U-1]}"; ws["B3"].font = F_NOTE
    ws.column_dimensions["B"].width = 12; ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 8
    for j in range(NM): ws.column_dimensions[get_column_letter(5 + j)].width = 9
    r = 4
    filas_bloque = []
    for bloque, src in [("DEMANDA (venta proyectada)", "Demanda"),
                        ("STOCK MÍNIMO", "Stock Minimo"),
                        ("STOCK IDEAL", "Stock Ideal")]:
        ws.cell(row=r, column=2, value=bloque).font = F_HDR
        ws.cell(row=r, column=2).fill = FILL_TOT
        r += 1
        ws.cell(row=r, column=2, value="Código").font = F_HDR
        ws.cell(row=r, column=3, value="Producto").font = F_HDR
        ws.cell(row=r, column=4, value="% CD").font = F_HDR
        for j, mm in enumerate(MESES):
            c = ws.cell(row=r, column=5 + j, value=mm)
            c.font = F_HDR_T if (j + 1) in TEMP_MESES else F_HDR
            c.fill = FILL_TEMP if (j + 1) in TEMP_MESES else FILL_HDR
        for col in range(2, 5): ws.cell(row=r, column=col).fill = FILL_HDR
        r += 1
        r0 = r; filas_bloque.append(r0)
        for c in orden:
            rs = fila_sku[c]
            ws.cell(row=r, column=2, value=c).font = F_TXT
            ws.cell(row=r, column=3, value=prod.loc[c, "nombre"]).font = F_TXT
            pc = ws.cell(row=r, column=4, value=f"='CD Split'!{share_col}{rs}")
            pc.font = F_TXT; pc.number_format = PCT
            for m in range(1, NM + 1):
                L = get_column_letter(4 + m)
                cc = ws.cell(row=r, column=4 + m,
                             value=f"='{src}'!{L}{rs}*'CD Split'!${share_col}${rs}")
                cc.font = F_TXT; cc.number_format = CAJ
                if bloque.startswith("DEMANDA") and m <= U: cc.fill = FILL_REAL
                elif m in TEMP_MESES: cc.fill = FILL_TEMP
            r += 1
        ws.cell(row=r, column=2, value="TOTAL").font = F_HDR
        ws.cell(row=r, column=2).fill = FILL_TOT
        for m in range(1, NM + 1):
            L = get_column_letter(4 + m)
            cc = ws.cell(row=r, column=4 + m, value=f"=SUM({L}{r0}:{L}{r-1})")
            cc.font = F_HDR; cc.number_format = CAJ
            cc.fill = FILL_TEMP if m in TEMP_MESES else FILL_TOT
        r += 3
    sin_triangulos(ws, f"E6:{get_column_letter(4 + NM)}{r}")
    ws.freeze_panes = "E6"
    bloques_cd[cd_nombre] = tuple(filas_bloque)

# ── Ingreso $ nacional ─────────────────────────────────────────────────────
def fp(m):
    return "*(1+Parametros!$C$7)" if m > 12 else ""

ws = hoja_grid("Ingreso $ SKU", "Ingreso estimado por SKU ($) — ene-2026 a abr-2027",
               f"Verde = FACTURADO real (ene–{MESES[U-1]}). Resto = cajas × precio/caja; 2027 lleva el ajuste de precio. "
               "La fila TOTAL es la venta del mes.", ancho_mes=13)
ws.cell(row=4, column=COL_PRECIO, value="Precio/caja").font = F_HDR
ws.column_dimensions[get_column_letter(COL_PRECIO)].width = 13
ws.column_dimensions[get_column_letter(COL_TOT26)].width = 15
ws.column_dimensions[get_column_letter(COL_TEMP)].width = 16
r = 5
fila_dinero = {}
for c in orden:
    rd = fila_sku[c]; fila_dinero[c] = r
    ws.cell(row=r, column=2, value=c).font = F_TXT
    ws.cell(row=r, column=3, value=prod.loc[c, "nombre"]).font = F_TXT
    ws.cell(row=r, column=4, value=prod.loc[c, "categoria"]).font = F_TXT
    for m in range(1, NM + 1):
        L = get_column_letter(4 + m)
        if m <= U:   # mes cerrado: neto REALMENTE facturado
            valr = float(neto_real.loc[c, m]) if (c in neto_real.index and m in neto_real.columns) else 0.0
            cc = ws.cell(row=r, column=4 + m, value=round(valr))
            cc.font = F_INPUT; cc.fill = FILL_REAL
        else:
            cc = ws.cell(row=r, column=4 + m, value=f"=Demanda!{L}{rd}*Demanda!$X${rd}{fp(m)}")
            cc.font = F_TXT
            if m in TEMP_MESES: cc.fill = FILL_TEMP
        cc.number_format = CLP
    ws.cell(row=r, column=COL_TOT26, value=f"=SUM(E{r}:P{r})").number_format = CLP
    t2 = ws.cell(row=r, column=COL_TEMP, value=f"=SUM(M{r}:T{r})")
    t2.font = F_HDR; t2.number_format = CLP; t2.fill = FILL_TEMP
    pc = ws.cell(row=r, column=COL_PRECIO, value=f"=Demanda!$X${rd}")
    pc.font = F_TXT; pc.number_format = CLP
    r += 1
rt_d = r
ws.cell(row=rt_d, column=2, value="TOTAL VENTA DEL MES ($)").font = F_HDR
ws.cell(row=rt_d, column=2).fill = FILL_TOT
for m in range(1, NM + 1):
    L = get_column_letter(4 + m)
    cc = ws.cell(row=rt_d, column=4 + m, value=f"=SUM({L}5:{L}{rt_d-1})")
    cc.font = F_HDR; cc.number_format = CLP
    cc.fill = FILL_REAL if m <= U else (FILL_TEMP if m in TEMP_MESES else FILL_TOT)
for col in (COL_TOT26, COL_TEMP):
    L = get_column_letter(col)
    cc = ws.cell(row=rt_d, column=col, value=f"=SUM({L}5:{L}{rt_d-1})")
    cc.font = F_HDR; cc.number_format = CLP; cc.fill = FILL_TOT
# desglose helados / galletas
r_h = rt_d + 2
filas_hel = [fila_dinero[c] for c in orden if c not in es_gall]
filas_gal = [fila_dinero[c] for c in orden if c in es_gall]
for etq, filas in [("  de los cuales HELADOS ($)", filas_hel), ("  de los cuales GALLETAS ($)", filas_gal)]:
    ws.cell(row=r_h, column=2, value=etq).font = F_TXT
    for m in range(1, NM + 1):
        L = get_column_letter(4 + m)
        ref = "+".join(f"{L}{f}" for f in filas) if filas else "0"
        cc = ws.cell(row=r_h, column=4 + m, value=f"={ref}")
        cc.font = F_TXT; cc.number_format = CLP
    ws.cell(row=r_h, column=COL_TOT26, value=f"=SUM(E{r_h}:P{r_h})").number_format = CLP
    r_h += 1
ws.cell(row=r_h, column=2, value="PLAN OFICIAL total del mes ($)").font = F_HDR
for m in range(1, NM + 1):
    val = plan_total[m - 1] if m <= 12 else plan_total[m - 13] * (1 + CREC_REAL_2027)
    cc = ws.cell(row=r_h, column=4 + m,
                 value=round(val) if m <= 12 else f"={round(val)}*(1+Parametros!$C$7)")
    cc.font = F_INPUT if m <= 12 else F_TXT; cc.number_format = CLP
ws.cell(row=r_h, column=COL_TOT26, value=f"=SUM(E{r_h}:P{r_h})").number_format = CLP
ws.cell(row=r_h + 1, column=2, value="Desvío vs plan").font = F_TXT
for m in range(1, NM + 1):
    L = get_column_letter(4 + m)
    cc = ws.cell(row=r_h + 1, column=4 + m, value=f"={L}{rt_d}/{L}{r_h}-1")
    cc.font = F_TXT; cc.number_format = PCT
ws.cell(row=r_h + 3, column=2,
        value=f"Meses REALES (verde): neto efectivamente facturado, cuadra con el ERP. El desvío vs plan de esos "
              f"meses es la diferencia real del negocio (julio cerró bajo el plan). En los proyectados, el desvío "
              "positivo es el aporte extra de galletas, que van por sobre el plan original.").font = F_NOTE
ws.cell(row=r_h + 4, column=2,
        value=f"OJO con el precio/caja: cayó de $23.703 promedio ene–may a $20.883 en julio (−12%). Las proyecciones "
              f"usan el promedio ene–{MESES[U-1]} por SKU; si el precio sigue bajando, el $ proyectado queda alto "
              "aunque las cajas estén bien.").font = F_NOTE

# ── Ingreso $ por CD ───────────────────────────────────────────────────────
ws = wb.create_sheet("Ingreso $ por CD"); ws.sheet_view.showGridLines = False
ws["B1"] = "Ingreso estimado por SKU y CD ($) — ene-2026 a abr-2027"; ws["B1"].font = F_TITLE
ws["B2"] = ("Ingreso = cajas del CD × precio/caja del SKU. Al final, la reconciliación: "
            "CCP + STGO + TEM = venta nacional del mes.")
ws["B2"].font = F_NOTE
ws["B3"] = f"Datos al {ULT_FECHA.strftime('%d-%m-%Y')} · REAL hasta {MESES[U-1]}"; ws["B3"].font = F_NOTE
ws.column_dimensions["B"].width = 12; ws.column_dimensions["C"].width = 38
for j in range(NM): ws.column_dimensions[get_column_letter(5 + j)].width = 13
r = 4
filas_total_cd = {}
for cd_nombre, share_col in CD_COL.items():
    hoja = f"CD - {cd_nombre.title()}"; r_dem0 = bloques_cd[cd_nombre][0]
    ws.cell(row=r, column=2, value=f"INGRESO $ — {cd_nombre}").font = F_HDR
    ws.cell(row=r, column=2).fill = FILL_TOT
    r += 1
    ws.cell(row=r, column=2, value="Código").font = F_HDR
    ws.cell(row=r, column=3, value="Producto").font = F_HDR
    for j, mm in enumerate(MESES):
        c = ws.cell(row=r, column=5 + j, value=mm)
        c.font = F_HDR_T if (j + 1) in TEMP_MESES else F_HDR
        c.fill = FILL_TEMP if (j + 1) in TEMP_MESES else FILL_HDR
    for col in (2, 3): ws.cell(row=r, column=col).fill = FILL_HDR
    r += 1
    r0 = r
    for i, c in enumerate(orden):
        rd = fila_sku[c]
        ws.cell(row=r, column=2, value=c).font = F_TXT
        ws.cell(row=r, column=3, value=prod.loc[c, "nombre"]).font = F_TXT
        for m in range(1, NM + 1):
            L = get_column_letter(4 + m)
            if m <= U:   # mes cerrado: neto REALMENTE facturado en ese CD
                key = (c, cd_nombre)
                valr = (float(neto_real_cd.loc[key, m])
                        if (key in neto_real_cd.index and m in neto_real_cd.columns) else 0.0)
                cc = ws.cell(row=r, column=4 + m, value=round(valr))
                cc.font = F_INPUT; cc.fill = FILL_REAL
            else:
                cc = ws.cell(row=r, column=4 + m,
                             value=f"='{hoja}'!{L}{r_dem0 + i}*Demanda!$X${rd}{fp(m)}")
                cc.font = F_TXT
                if m in TEMP_MESES: cc.fill = FILL_TEMP
            cc.number_format = CLP
        r += 1
    ws.cell(row=r, column=2, value=f"TOTAL {cd_nombre} ($)").font = F_HDR
    ws.cell(row=r, column=2).fill = FILL_TOT
    for m in range(1, NM + 1):
        L = get_column_letter(4 + m)
        cc = ws.cell(row=r, column=4 + m, value=f"=SUM({L}{r0}:{L}{r-1})")
        cc.font = F_HDR; cc.number_format = CLP
        cc.fill = FILL_REAL if m <= U else (FILL_TEMP if m in TEMP_MESES else FILL_TOT)
    filas_total_cd[cd_nombre] = r
    r += 3
ws.cell(row=r, column=2, value="RECONCILIACIÓN CON LA VENTA NACIONAL").font = F_HDR
ws.cell(row=r, column=2).fill = FILL_TOT
r += 1
for cd in CD_COL:
    ws.cell(row=r, column=2, value=f"Total {cd} ($)").font = F_TXT
    for m in range(1, NM + 1):
        L = get_column_letter(4 + m)
        cc = ws.cell(row=r, column=4 + m, value=f"={L}{filas_total_cd[cd]}")
        cc.font = F_TXT; cc.number_format = CLP
    r += 1
ws.cell(row=r, column=2, value="VENTA NACIONAL DEL MES ($)").font = F_HDR
ws.cell(row=r, column=2).fill = FILL_TOT
for m in range(1, NM + 1):
    L = get_column_letter(4 + m)
    partes = "+".join(f"{L}{filas_total_cd[cd]}" for cd in CD_COL)
    cc = ws.cell(row=r, column=4 + m, value=f"={partes}")
    cc.font = F_HDR; cc.number_format = CLP
    cc.fill = FILL_TEMP if m in TEMP_MESES else FILL_TOT
r_nac = r; r += 1
ws.cell(row=r, column=2, value="Chequeo vs hoja 'Ingreso $ SKU' (debe ser 0)").font = F_NOTE
for m in range(1, NM + 1):
    L = get_column_letter(4 + m)
    cc = ws.cell(row=r, column=4 + m, value=f"={L}{r_nac}-'Ingreso $ SKU'!{L}{rt_d}")
    cc.font = F_NOTE; cc.number_format = CLP
ws.freeze_panes = "E6"

destino = SALIDA_DRIVE
try:
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
except Exception as e:
    print(f"\n  [!] No se pudo escribir en Drive ({type(e).__name__}).")
    print("      Causa habitual: el archivo está ABIERTO en Excel. Ciérralo y vuelve a correr.")
    destino = SALIDA_LOCAL
    wb.save(destino)
    print(f"      Se guardó una copia en: {destino}")

print(f"\nOK — {n_skus} SKUs ({len(es_gall)} galletas incluidas) · corte real: {MESES[U-1]}")
print(f"Archivo: {destino}")
print("(el .bat recalcula las fórmulas con Excel al terminar)")
# Marca legible por el .ps1 para que recalcule EL archivo que se escribió
# (si Drive estaba bloqueado, el bueno es la copia local, no el de Drive).
print(f"##ARCHIVO##{destino}")
